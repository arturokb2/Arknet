import copy
from django.db.models import Q
from services.hospital.annual_reports import AnnualReportABC, get_rez_apr_1
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from openpyxl import load_workbook,styles
import os
from datetime import datetime
from services.hospital.patient import PatientsData
from collections import OrderedDict
import json
import numpy
from services.hospital.reports import *
from operator import itemgetter, attrgetter, methodcaller
from hospital.models import Oper,Oslo,V001,Manpy,PR_OSOB
from okb2.models import Ds,Vra,otde,V036
import re
from services.hospital.reports import TerrSpecification
border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin', color='000000'))


nzI60 = list(Ds.objects.values('kod').filter(kod__range=('I60','I60.9')))
nzI60 = [k['kod'] for k in nzI60]

nzI61 = list(Ds.objects.values('kod').filter(kod__range=('I61','I61.9')))
nzI61 = [k['kod'] for k in nzI61]

nzI61_I62 = list(Ds.objects.values('kod').filter(kod__range=('I61','I62.9')))
nzI61_I62 = [k['kod'] for k in nzI61_I62]

nzI63 = list(Ds.objects.values('kod').filter(kod__range=('I63','I63.9')))
nzI63 = [k['kod'] for k in nzI63]

nzI67_I69 = list(Ds.objects.values('kod').filter(kod__range=('I67','I69.9')))
nzI67_I69 = [k['kod'] for k in nzI67_I69]

nzG00_G09 = list(Ds.objects.values('kod').filter(kod__range=('G00','G09.9')))
nzG00_G09 = [k['kod'] for k in nzG00_G09]

nzG92 = list(Ds.objects.values('kod').filter(kod__range=('G92','G92.9')))
nzG92 = [k['kod'] for k in nzG92]

nzC70C72_C32_C33 = list(Ds.objects.values('kod').filter(Q(kod__range=('C70','C72.9'))|Q(kod__range=('C32','C33.9'))))
nzC70C72_C32_C33 = [k['kod'] for k in nzC70C72_C32_C33]

nzG45_46= list(Ds.objects.values('kod').filter(kod__range=('G45','G46.9')))
nzG45_46 = [k['kod'] for k in nzG45_46]

T1 = ['S02.0','S02.1','S04.0S05.7','S06.1','S06.2','S06.3','S06.4','S06.5',
        'S06.6','S06.7','S07.0','S07.1','S07.8','S09.0','S11.0','S11.1','S11.1',
        'S11.2','S11.7','S15.0','S15.1','S15.2','S15.3','S15.7','S15.8','S15.9',
        'S17.0','S17.8','S18']

T2 = ['S12.0','S12.9','S13.0','S13.1','S13.3','S14.0','S14.3','S22.0','S23.0',
        'S23.1','S24.0','S32.0','S32.1','S33.0','S33.1','S33.2','S33.4','S34.0',
        'S34.3','S34.4']

T3 = ['S22.2','S22.4','S22.5','S25.0','S25.1','S25.2','S25.3','S25.4','S25.5',
        'S25.7','S25.8','S25.9','S26.0','S27.0','S27.1','S27.2','S27.4','S27.5',
        'S27.6','S27.8','S28.0','S28.1']

T4 = ['S35.0','S35.1','S35.2','S35.3','S35.4','S35.5','S35.7','S35.8','S35.9',
        'S36.0','S36.1','S36.2','S36.3','S36.4','S36.5','S36.8','S36.9','S37.0','S38.3']

T5 = ['S32.3','S32.4','S32.5','S36.6','S37.1','S37.2','S37.4','S37.5','S37.6',
        'S37.8','S38.0','S38.2']

T6 = ['S42.2','S42.3','S42.4','S42.8','S45.0','S45.1','S45.2','S45.7','S45.8',
        'S47','S48.0','S48.1','S48.9','S52.7','S55.0','S55.1','S55.7','S55.8',
        'S57.0','S57.8','S57.9','S58.0','S58.1','S58.9','S68.4','S71.7','S72.0',
        'S72.1','S72.2','S72.3','S72.4','S72.7','S75.0','S75.1','S75.2','S75.7',
        'S75.8','S77.0','S77.1','S77.2','S78.0','S78.1','S78.9','S79.7','S82.1',
        'S82.2','S82.3','S82.7','S85.0','S85.1','S85.5','S85.7','S87.0','S87.8',
        'S88.0','S88.1','S88.9','S95.7','S95.8','S95.9','S97.0','S97.8','S98.0']

T7 = ['S02.7','S12.7','S22.1','S27.7','S29.7','S31.7','S32.7','S36.7','S38.1',
        'S39.6','S39.7','S37.7','S42.7','S49.7','T01.1','T01.8','T01.9','T02.0',
        'T02.1','T02.2','T02.3','T02.4','T02.5','T02.6','T02.7','T02.8','T02.9',
        'T04.0','T04.1','T04.2','T04.3','T04.4','T04.7','T04.8','T04.9','T05.0',
        'T05.1','T05.2','T05.3','T05.4','T05.5','T05.6','T05.8','T05.9','T06.0',
        'T06.1','T06.2','T06.3','T06.4','T06.5','T06.8','T07']

T8 = ['T01.1','T01.8','T01.9','T02.0','T02.1','T02.2','T02.3','T02.4','T02.5',
      'T02.6','T02.7','T02.8','T02.9','T04.0','T04.1','T04.2','T04.3','T04.4',
      'T04.7','T04.8','T04.9','T05.0','T05.1','T05.2','T05.3','T05.4','T05.5',
      'T05.6','T05.8','T05.9','T06.0','T06.1','T06.2','T06.3','T06.4','T06.5',
      'T06.8','T07']


def get_pop_oper(patient):
    pop = patient.sluchay.oper.filter(pop=True) if patient.sluchay.oper else 0
    if pop.count() > 0:
        return pop[0]
    else:
        return None

def get_opers(patient):
    if patient.sluchay.oper.count() > 0:
        opers = patient.sluchay.oper.values('id')
        return [Oper.objects.get(id=o['id']) for o in opers]
    else:
        return None

def get_list_otd(data,ym=None):
    otds_list = []
    for d in data:
        try:
            if d.sluchay.otd.naim not in otds_list:
                otds_list.append(d.sluchay.otd.naim)
        except Exception:
            otds_list.append(None)


    otds_sl = []
    try:
        for o in otds_list:
            temp = [[],[]]
            temp[0].append(o)
            for d in data:
                if d.sluchay.otd:
                    if o == d.sluchay.otd.naim:
                        temp[1].append(d)
                else:
                    if o == d.sluchay.otd:
                        temp[1].append(d)
            otds_sl.append(temp)
    except Exception as er:
        print(er)

    return otds_sl



def get_list_pers(data):
    pers_list = []
    for d in data:
        if [d.patient.im,d.patient.fam,d.patient.ot] not in pers_list:
            pers_list.append([d.patient.im,d.patient.fam,d.patient.ot])
    sl_list = []
    for o in pers_list:
        temp = [[],[]]
        temp[0].append(o)
        for d in data:
            if o == [d.patient.im,d.patient.fam,d.patient.ot]:
                temp[1].append(d)
        sl_list.append(temp)
    return sl_list

def get_list_lpu(data,all=None,typ_lpy=None):
    lpu_list = []
    for d in data:
        if typ_lpy is  None:
            if d.sluchay.lpy and d.sluchay.lpy.naim not in lpu_list:
                lpu_list.append(d.sluchay.lpy.naim)
            else:
                if d.sluchay.pmg and d.sluchay.pmg.naim not in lpu_list:
                    lpu_list.append(d.sluchay.pmg.naim)
    if all is not None:
        lpu_list.append(None)

    sl = []
    for o in lpu_list:
        temp = [[],[]]
        temp[0].append(o)
        for d in data:
            if typ_lpy is None:
                if d.sluchay.lpy and o == d.sluchay.lpy.naim:
                    temp[1].append(d)
            else:
                if d.sluchay.pmg and o == d.sluchay.pmg.naim:
                    temp[1].append(d)
            if all is not None:
                if o == d.sluchay.pmg:
                    temp[1].append(d)
        sl.append(temp)
    return sl
def get_list_ds(data):
    ds_list = []
    ds_list_kod = []

    for d in data:
        if d.sluchay.dspat == None and d.sluchay.dskz and d.sluchay.dskz not in ds_list:
            ds_list.append(d.sluchay.dskz)
            ds_list_kod.append(d.sluchay.dskz.kod)
        else:
            if d.sluchay.dspat and d.sluchay.dspat not in ds_list:
                ds_list.append(d.sluchay.dspat)
                ds_list_kod.append(d.sluchay.dspat.kod)
    ds_list_kod.sort()
    _=[]
    for kod in ds_list_kod:
        for ds in ds_list:
            if kod == ds.kod:
                _.append(ds)
    ds_list = _
    ds = []
    for o in ds_list:
        temp = [[], []]
        temp[0].append(o)
        for d in data:
            if d.sluchay.dspat == None:
                if d.sluchay.dskz and o.kod == d.sluchay.dskz.kod:
                    temp[1].append(d)
            else:
                if d.sluchay.dspat and o.kod == d.sluchay.dspat.kod:
                    temp[1].append(d)
        ds.append(temp)
    return ds

def get_list_ds_oper(data):
    ds_list = []
    for d in data:
        if d.sluchay.dskz and d.sluchay.dskz.kod not in ds_list:
            ds_list.append(d.sluchay.dskz.kod)
    ds_oper = []
    for o in ds_list:
        temp = [[],[],[]]
        temp[0].append(o)
        for d in data:
            if d.sluchay.dskz and o == d.sluchay.dskz.kod:
                if d.sluchay.oper.count() > 0:
                    opers = d.sluchay.oper.values('id')
                    for op in opers:
                        oper = Oper.objects.get(id=op['id'])
                        temp[1].append(oper.kod_op.kod if oper.kod_op else None)
        ds_oper.append(temp)

    for v in ds_oper:
        v[1] = list(set(v[1]))

    for v in ds_oper:
        for d in data:
            if d.sluchay.dskz and v[0][0] == d.sluchay.dskz.kod:
                if d.sluchay.oper.count() > 0:
                    opers = d.sluchay.oper.values('id')
                    for op in opers:
                        oper = Oper.objects.get(id=op['id'])
                        if oper.kod_op and oper.kod_op.kod in v[1]:
                            if d not in v[2]:
                                v[2].append(d)
                else:
                    if d not in v[2]:
                        v[2].append(d)
    return ds_oper

def get_list_oper(data):
    oper_kod_list = []
    for d in data:
        if d.sluchay.oper.count() > 0:
            opers = d.sluchay.oper.values('id')
            for o in opers:
                oper = Oper.objects.get(id=o['id'])
                if oper.kod_op and oper.kod_op.kod not in oper_kod_list:
                    oper_kod_list.append(oper.kod_op.kod)

    opers_list = []
    for o in oper_kod_list:
        temp = [[], []]
        temp[0].append(o)
        for d in data:
            if d.sluchay.oper.count() > 0:
                opers = d.sluchay.oper.values('id')
                for op in opers:
                    oper = Oper.objects.get(id=op['id'])
                    if oper.kod_op and oper.kod_op.kod == o:
                        temp[1].append(d)
        opers_list.append(temp)
    return opers_list


def get_list_otd_prof(data):
    otds_list = []
    for d in data:
        if d.sluchay.otd.naim not in otds_list:
            otds_list.append(d.sluchay.otd.naim)

    otds_prof_list = []
    for o in otds_list:
        temp = [[],[]]
        temp[0].append(o)
        for d in data:
            if o == d.sluchay.otd.naim:
                if d.sluchay.le_vr and d.sluchay.le_vr.prof_k:
                    if d.sluchay.le_vr.prof_k.k_prname not in temp[1]:
                        temp[1].append(d.sluchay.le_vr.prof_k.k_prname)
        otds_prof_list.append(temp)

    for otds_prof in otds_prof_list:
        for o in range(len(otds_prof[1])):
            otds_prof[1][o] = [otds_prof[1][o],[]]

    for otds_prof in range(len(otds_prof_list)):
        for d in data:
            if otds_prof_list[otds_prof][0][0] == d.sluchay.otd.naim:
                if d.sluchay.le_vr and d.sluchay.le_vr.prof_k:
                    for p in range(len(otds_prof_list[otds_prof][1])):
                        if otds_prof_list[otds_prof][1][p][0] == d.sluchay.le_vr.prof_k.k_prname:
                            otds_prof_list[otds_prof][1][p][1].append(d)

    return otds_prof_list

def get_rez_rep_4(data,d=None):
    if d == None:
        bf = BetterFilter()
        sp = CountSluchaySpecification() ^ GocEkSpecification() ^ ProfKNSpecification() ^ OperCountSpecification() ^ \
             OperAllKdSpecification() ^ OperAllCountSpecification() ^ PredOperKdAllSpecification() ^ RezUmerSpecification() ^ \
             RezUmerKdSpecification() ^ RezUmerOperSpecification() ^ RezUmerOperKDSpecification()
        all_temp = []
        for patient in data:
            for p in bf.filter(patient, sp):
                temp = bf.format_list(p)
                for t in range(len(temp)):
                    if temp[t] == 'None':
                        temp[t] = 0
                all_temp.append([int(i) for i in temp])
        all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(11)]
        all_temp.insert(11,copy.deepcopy(all_temp[2]))
        all_temp.insert(12,copy.deepcopy(all_temp[4]))
        all_temp.insert(13,copy.deepcopy(all_temp[8]))
        all_temp.insert(14,copy.deepcopy(all_temp[10]))

        try:
            all_temp[2]= float('{0:.2f}'.format(all_temp[2]/all_temp[0]))
        except ZeroDivisionError:
            all_temp[2] = 0

        try:
            all_temp[4]= float('{0:.2f}'.format(all_temp[4]/all_temp[3]))
        except ZeroDivisionError:
            all_temp[4] = 0

        try:
            all_temp[8]= float('{0:.2f}'.format(all_temp[8]/all_temp[7]))
        except ZeroDivisionError:
            all_temp[8] = 0

        try:
            all_temp[10]= float('{0:.2f}'.format(all_temp[10]/all_temp[9]))
        except ZeroDivisionError:
            all_temp[10] = 0

        return all_temp
    else:
        try:
            data[2]= float('{0:.2f}'.format(data[11]/data[0]))
        except ZeroDivisionError:
            data[2] = 0

        try:
            data[4]= float('{0:.2f}'.format(data[12]/data[3]))
        except ZeroDivisionError:
            data[4] = 0

        try:
            data[8]= float('{0:.2f}'.format(data[13]/data[7]))
        except ZeroDivisionError:
            data[8] = 0

        try:
            data[10]= float('{0:.2f}'.format(data[14]/data[9]))
        except ZeroDivisionError:
            data[10] = 0
        return data
def get_rez_rep_5_help(data,rez,nz,t,s):
    for d in data:
        if d.sluchay.dskz:
            ds = d.sluchay.dskz.kod
            year = d.patient_year
            pol = d.patient.pol.id_pol if d.patient.pol and d.patient.pol.id_pol else None
            if s == True:
                if t == True:
                    if nz == nzG00_G09:
                        print(d.sluchay.nib)
                    if ds in nz:
                        rez[0] += 1
                        if 15 <= year <= 17:
                            rez[1] += 1
                        elif 18 <= year <= 19:
                            rez[2] += 1
                        elif 20 <= year <= 29:
                            rez[3] += 1
                        elif 30 <= year <= 39:
                            rez[4] += 1
                        elif 40 <= year <= 49:
                            rez[5] += 1
                        elif 50 <= year <= 59:
                            rez[6] += 1
                        elif 60 <= year <= 69:
                            rez[7] += 1
                        elif 70 <= year <= 79:
                            rez[8] += 1
                        elif year >= 80:
                            rez[9] += 1

                        if pol == 1:
                            if year <= settings.OLD_M:
                                rez[10] += 1
                            elif year >= settings.OLD_M:
                                rez[11] += 1
                        if pol == 2:
                            if year <= settings.OLD_G:
                                rez[10] += 1
                            elif year >= settings.OLD_G:
                                rez[11] += 1
                else:
                    if ds not in nz:
                        rez[0] += 1
                        if 15 <= year <= 17:
                            rez[1] += 1
                        elif 18 <= year <= 19:
                            rez[2] += 1
                        elif 20 <= year <= 29:
                            rez[3] += 1
                        elif 30 <= year <= 39:
                            rez[4] += 1
                        elif 40 <= year <= 49:
                            rez[5] += 1
                        elif 50 <= year <= 59:
                            rez[6] += 1
                        elif 60 <= year <= 69:
                            rez[7] += 1
                        elif 70 <= year <= 79:
                            rez[8] += 1
                        elif year >= 80:
                            rez[9] += 1

                        if pol == 1:
                            if year <= settings.OLD_M:
                                rez[10] += 1
                            elif year >= settings.OLD_M:
                                rez[11] += 1
                        if pol == 2:
                            if year <= settings.OLD_G:
                                rez[10] += 1
                            elif year >= settings.OLD_G:
                                rez[11] += 1
            else:
                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                    if t == True:
                        if ds in nz:
                            rez[0] += 1
                            if 15 <= year <= 17:
                                rez[1] += 1
                            elif 18 <= year <= 19:
                                rez[2] += 1
                            elif 20 <= year <= 29:
                                rez[3] += 1
                            elif 30 <= year <= 39:
                                rez[4] += 1
                            elif 40 <= year <= 49:
                                rez[5] += 1
                            elif 50 <= year <= 59:
                                rez[6] += 1
                            elif 60 <= year <= 69:
                                rez[7] += 1
                            elif 70 <= year <= 79:
                                rez[8] += 1
                            elif year >= 80:
                                rez[9] += 1
                            if pol == 1:
                                if year <= settings.OLD_M:
                                    rez[10] += 1
                                elif year >= settings.OLD_M:
                                    rez[11] += 1
                            if pol == 2:
                                if year <= settings.OLD_G:
                                    rez[10] += 1
                                elif year >= settings.OLD_G:
                                    rez[11] += 1
                    else:
                        if ds not in nz:
                            rez[0] += 1
                            if 15 <= year <= 17:
                                rez[1] += 1
                            elif 18 <= year <= 19:
                                rez[2] += 1
                            elif 20 <= year <= 29:
                                rez[3] += 1
                            elif 30 <= year <= 39:
                                rez[4] += 1
                            elif 40 <= year <= 49:
                                rez[5] += 1
                            elif 50 <= year <= 59:
                                rez[6] += 1
                            elif 60 <= year <= 69:
                                rez[7] += 1
                            elif 70 <= year <= 79:
                                rez[8] += 1
                            elif year >= 80:
                                rez[9] += 1

                            if pol == 1:
                                if year <= settings.OLD_M:
                                    rez[10] += 1
                                elif year >= settings.OLD_M:
                                    rez[11] += 1
                            if pol == 2:
                                if year <= settings.OLD_G:
                                    rez[10] += 1
                                elif year >= settings.OLD_G:
                                    rez[11] += 1
    # rez[0]=sum(rez)
    return rez
def get_rez_rep_5(data,t):
    I60 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    I61_I62 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    I63 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    I67_I69 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    G00_G09 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    G92 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    C70C72_C32_C33 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    _ = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    not_nz = nzI60 + nzI61_I62 + nzI63 + nzI67_I69 + nzG00_G09 + nzC70C72_C32_C33
    if t == 1:
        I60 = get_rez_rep_5_help(data, I60, nzI60, True, True)
        I61_I62 = get_rez_rep_5_help(data, I61_I62, nzI61_I62, True, True)
        I63 = get_rez_rep_5_help(data, I63, nzI63, True, True)
        I67_I69 = get_rez_rep_5_help(data, I67_I69, nzI67_I69, True, True)
        G00_G09 = get_rez_rep_5_help(data, G00_G09, nzG00_G09, True, True)
        G92 = get_rez_rep_5_help(data, G92, nzG92, True, True)
        C70C72_C32_C33 = get_rez_rep_5_help(data, C70C72_C32_C33, nzC70C72_C32_C33, True, True)
        not_nz = get_rez_rep_5_help(data, _, not_nz, False, True)
    else:
        I60 = get_rez_rep_5_help(data, I60, nzI60, True, False)
        I61_I62 = get_rez_rep_5_help(data, I61_I62, nzI61_I62, True, False)
        I63 = get_rez_rep_5_help(data, I63, nzI63, True, False)
        I67_I69 = get_rez_rep_5_help(data, I67_I69, nzI67_I69, True, False)
        G00_G09 = get_rez_rep_5_help(data, G00_G09, nzG00_G09, True, False)
        G92 = get_rez_rep_5_help(data, G92, nzG92, True, False)
        C70C72_C32_C33 = get_rez_rep_5_help(data, C70C72_C32_C33, nzC70C72_C32_C33, True, False)
        not_nz = get_rez_rep_5_help(data, _, not_nz, False, False)

    _ = [I60, I61_I62, I63, I67_I69, G00_G09, G92, C70C72_C32_C33, not_nz]
    r = None
    for o in range(len(_)):
        if o == 0:
            r = numpy.array(_[o])
        else:
            r += numpy.array(_[o])
    _ = r.tolist()
    return [
        I60, I61_I62, I63, I67_I69, G00_G09, G92, C70C72_C32_C33, not_nz, _
    ]

    #
    # if t == 1:
    #     I60 = [0,0,0,0,0,0,0,0,0,0,0,0]
    #     I61_I62 = [0,0,0,0,0,0,0,0,0,0,0,0]
    #     I63 = [0,0,0,0,0,0,0,0,0,0,0,0]
    #     I67_I69 = [0,0,0,0,0,0,0,0,0,0,0,0]
    #     G00_G09 = [0,0,0,0,0,0,0,0,0,0,0,0]
    #     G92 = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
    #     C70C72_C32_C33 = [0,0,0,0,0,0,0,0,0,0,0,0]
    #     _ = [0,0,0,0,0,0,0,0,0,0,0,0]
    #     not_nz = nzI60 + nzI61_I62 + nzI63 + nzI67_I69 + nzG00_G09 + nzC70C72_C32_C33
    #
    #     I60 = get_rez_rep_5_help(data,I60,nzI60,True,t)
    #     I61_I62 = get_rez_rep_5_help(data,I61_I62,nzI61_I62,True,t)
    #     I63 = get_rez_rep_5_help(data,I63,nzI63,True,t)
    #     I67_I69 = get_rez_rep_5_help(data,I67_I69,nzI67_I69,True,t)
    #     G00_G09 = get_rez_rep_5_help(data,G00_G09,nzG00_G09,True,t)
    #     G92 = get_rez_rep_5_help(data,G92,nzG92,True,t)
    #     C70C72_C32_C33 = get_rez_rep_5_help(data,C70C72_C32_C33,nzC70C72_C32_C33,True,t)
    #     not_nz = get_rez_rep_5_help(data,_,not_nz,False,t)
    #
    #     _ = [I60,I61_I62,I63,I67_I69,G00_G09,G92,C70C72_C32_C33,not_nz]
    #     r = None
    #     for o in range(len(_)):
    #         if o == 0:
    #             r = numpy.array(_[o])
    #         else:
    #             r += numpy.array(_[o])
    #     _ = r.tolist()
    #     return [
    #         I60,I61_I62,I63,I67_I69,G00_G09,G92,C70C72_C32_C33,not_nz,_
    #     ]
    #
    # else:
    #     pass
def get_rez_rep_5_5(data,ds):
    ds[0]=len(data)
    for d in data:
        if d.sluchay.goc and d.sluchay.goc.tip_name == 'Экстренная':
            ds[1]+=1
            ds[4]+=d.sluchay.le_vr.kd if d.sluchay.le_vr.kd != ''and d.sluchay.le_vr.kd != None else 0
        elif d.sluchay.goc and d.sluchay.goc.tip_name == 'Плановая':
            ds[2]+=1
            ds[5] += d.sluchay.le_vr.kd if d.sluchay.le_vr.kd != '' and d.sluchay.le_vr.kd != None else 0
        ds[3]+=d.sluchay.le_vr.kd if d.sluchay.le_vr.kd != ''and d.sluchay.le_vr.kd != None else 0
        if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105,106]:
            ds[9]+=1
            if d.sluchay.goc and d.sluchay.goc.tip_name == 'Экстренная':
                ds[10]+=1
            elif d.sluchay.goc and d.sluchay.goc.tip_name == 'Плановая':
                ds[11] += 1
    try:
        ds[6] = float('{0:.2f}'.format(ds[3]/ds[0]))
    except ZeroDivisionError:
        ds[6] = 0

    try:
        ds[7] = float('{0:.2f}'.format(ds[4]/ds[1]))
    except ZeroDivisionError:
        ds[7] = 0
    try:
        ds[8] = float('{0:.2f}'.format(ds[5] / ds[2]))
    except ZeroDivisionError:
        ds[8] = 0
    return ds
def get_rez_rep_5_6(data,ds):
    for d in data:
        if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105,106]:
            ds[0]+=1
            kd = d.sluchay.le_vr.kd if d.sluchay.le_vr.kd != '' and d.sluchay.le_vr.kd != None else 0
            ds[1] += kd
            if kd == 1:
                ds[3]+=1
            if 1<=kd<=3:
                ds[4]+=1
            if 4<=kd<=6:
                ds[5]+=1
            if 7<=kd<=9:
                ds[6]+=1
            if kd>=10:
                ds[7]+=1
    return ds
def get_rez_rep_5_6_umer(data,ds):
    for d in data:
        if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
            ds[0] += 1
            year = d.patient_year
            if 20<=year<=29:
                ds[1]+=1
            elif 40<=year<=49:
                ds[2]+=1
            elif 50<=year<=59:
                ds[3]+=1
            elif 60<=year<=69:
                ds[4]+=1
            elif 70<=year<=79:
                ds[5]+=1
            elif year>=80:
                ds[6]+=1
    return ds

class PatientsDataFiltrs(PatientsData):
    def __init__(self,date_1,date_2,user,request):
        super().__init__(date_1,date_2,user)
        pats = Reports(user.id,request)
        pats.get_sluchays()

        filters = json.loads(request.get('filters'))
        sl = pats.filter(filters['filter'],pats.patients)
        for s in sl:
            self.get_data(s['sluchay'].id)



def insert_sheet_P1(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = kwargs['data']
    count = kwargs['count']
    ym_count = kwargs['ym_count']
    otdels = kwargs['otdels']


    sheet.cell(row=3, column=1).value = str(name).capitalize()
    sheet.cell(row=4,column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    row_ = 7
    font = styles.Font(size=16, name='Arial')
    if len(otdels)>0:
        filter = []
        height =20
        for o in otdels:
            filter.append(o)
        sheet.row_dimensions[5].height = height
        sheet.cell(row=5, column=1).value = 'Отделение: ' + ','.join(filter)
    for dat in data:
        row_ += 1
        sheet.row_dimensions[row_].height = 20
        sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:J{sheet.cell(row=row_, column=9).row}")
        sheet.cell(row=row_, column=1).value = dat[0]
        sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=1).font = font
        sheet.cell(row=row_, column=1).border = styles.Border(
            bottom=styles.Side(border_style='thin', color='000000'))
        for d in dat[1]:
            row_ += 1
            sheet.row_dimensions[row_].height = 33
            if (row_ % 64 == 0 and row_ <= 64) or (row_ > 64 and row_ % 61 == 3):
                sheet.cell(row=row_, column=1).value = 'История'
                sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=1).font = font
                sheet.cell(row=row_, column=2).value = 'Фамилия.И.О'
                sheet.cell(row=row_, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=2).font = font
                sheet.cell(row=row_, column=3).value = 'Дата рождения'
                sheet.cell(row=row_, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=3).font = font
                sheet.cell(row=row_, column=4).value = 'Дата поступ.'
                sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=4).font = font
                sheet.cell(row=row_, column=5).value = 'Дата выбытия'
                sheet.cell(row=row_, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=5).font = font
                sheet.cell(row=row_, column=6).value = 'Исх леч'
                sheet.cell(row=row_, column=6).alignment = styles.Alignment(horizontal="center", vertical="center",
                                                                            wrap_text=True)
                sheet.cell(row=row_, column=6).font = font
                sheet.cell(row=row_, column=7).value = 'Ds'
                sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=7).font = font
                sheet.merge_cells(
                    f"H{sheet.cell(row=row_, column=8).row}:I{sheet.cell(row=row_, column=9).row}")
                sheet.cell(row=row_, column=8).value = 'Профиль койки'
                sheet.cell(row=row_, column=8).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=8).font = font
                sheet.cell(row=row_, column=10).value = 'Оплат.'
                sheet.cell(row=row_, column=10).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=10).font = font
                row_ += 1
                sheet.row_dimensions[row_].height = 20
                sheet.merge_cells(
                    f"A{sheet.cell(row=row_, column=1).row}:J{sheet.cell(row=row_, column=9).row}")
                sheet.cell(row=row_, column=1).value = dat[0]
                sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center",
                                                                            vertical="center")
                sheet.cell(row=row_, column=1).font = font
                sheet.cell(row=row_, column=1).border = styles.Border(
                    bottom=styles.Side(border_style='thin', color='000000'))
                row_ += 1
            sheet.row_dimensions[row_].height = 20
            sheet.cell(row=row_, column=1).value = d['nib']
            sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center")
            sheet.cell(row=row_, column=1).font = font
            sheet.cell(row=row_, column=2).value = d['fio']
            sheet.cell(row=row_, column=2).alignment = styles.Alignment(horizontal="left", vertical="center")
            sheet.cell(row=row_, column=2).font = font
            sheet.cell(row=row_, column=3).value = d['datr']
            sheet.cell(row=row_, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_, column=3).font = font
            sheet.cell(row=row_, column=4).value = d['datp']
            sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_, column=4).font = font
            sheet.cell(row=row_, column=5).value = d['datv']
            sheet.cell(row=row_, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_, column=5).font = font
            sheet.cell(row=row_, column=6).value = d['isx']
            sheet.cell(row=row_, column=6).alignment = styles.Alignment(horizontal="center", vertical="center",
                                                                        wrap_text=True)
            sheet.cell(row=row_, column=6).font = font
            sheet.cell(row=row_, column=7).value = d['dskz']
            sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_, column=7).font = font
            sheet.merge_cells(f"H{sheet.cell(row=row_, column=8).row}:I{sheet.cell(row=row_, column=9).row}")
            sheet.cell(row=row_, column=8).value = d['prof_k']
            sheet.cell(row=row_, column=8).font = font
            sheet.cell(row=row_, column=10).value = d['vds']
            sheet.cell(row=row_, column=10).font = font
    else:
        row_ += 3
        sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:B{sheet.cell(row=row_, column=2).row}")
        sheet.cell(row=row_, column=1).value = f'Итого по отделению - {count}'
        sheet.cell(row=row_, column=1).font = font
        sheet.cell(row=row_, column=3).value = f'Умерших - {ym_count}'
        sheet.cell(row=row_, column=3).font = font
        sheet.row_dimensions[row_].height = 20
        row_ += 1
        sheet.cell(row=row_, column=1).value = 'Из них:'
        sheet.cell(row=row_, column=1).font = font
        sheet.row_dimensions[row_].height = 20
        for dat in data:
            row_ += 1
            sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:C{sheet.cell(row=row_, column=3).row}")
            sheet.cell(row=row_, column=1).value = f'{dat[0]} - {len(dat[1])}'
            sheet.cell(row=row_, column=1).font = font
            sheet.row_dimensions[row_].height = 20
        return sheet

def insert_sheet_implants(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = []
    for n, p in enumerate(kwargs['data'], 1):
        temp = OrderedDict()
        temp['n'] = n
        temp['nib'] = p.sluchay.nib
        temp['otd'] = p.sluchay.otd.naim if p.sluchay.otd else None
        temp['fio'] = f'{p.patient.fam} {p.patient.im[0] if len(p.patient.im) > 0 else ""}.{p.patient.ot[0] if len(p.patient.ot) > 0 else ""}'
        temp['datr'] = p.patient.datr
        temp['datp'] = p.sluchay.datp
        temp['datv'] = p.sluchay.datv
        oper = get_pop_oper(p)
        temp['oper_kod'] = oper.kod_op.kod if oper is not None and oper.kod_op else None
        if p.sluchay.med_dev.count() > 0:
            med_dev = p.sluchay.med_dev.first()
            temp['date_med'] = med_dev.date
            temp['code_med'] = med_dev.code.rzn if med_dev.code is not None else None
            temp['number_ser'] = med_dev.number_ser
        else:
            temp['date_med'] = None
            temp['code_med'] = None
            temp['number_ser'] = None
        data.append(temp)
    data = sorted(data, key=itemgetter('fio'))
    row = 8
    sheet.cell(row=4, column=1).value = str(name).capitalize()
    sheet.cell(row=5,column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    for n,d in enumerate(data,1):
        row += 1
        sheet.row_dimensions[row].height = 20
        sheet.cell(row=row, column=1).value = n
        sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row, column=2).value = d['nib']
        sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row, column=3).value = d['otd']
        sheet.cell(row=row, column=3).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        sheet.cell(row=row, column=4).value = d['fio']
        sheet.cell(row=row, column=4).alignment = styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.cell(row=row, column=5).value = f"{d['datr'].strftime('%d.%m.%Y')}" if d['datr'] else None
        sheet.cell(row=row, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row, column=6).value = f"{d['datp'].strftime('%d.%m.%Y')}" if d['datp'] else None
        sheet.cell(row=row, column=6).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row, column=7).value = d['oper_kod']
        sheet.cell(row=row, column=7).alignment = styles.Alignment(horizontal="left", vertical="center", wrap_text=True)
        sheet.cell(row=row, column=8).value = f"{d['date_med'].strftime('%d.%m.%Y')}" if d['date_med'] else None
        sheet.cell(row=row, column=8).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row, column=9).value = d['code_med']
        sheet.cell(row=row, column=9).alignment = styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.cell(row=row, column=10).value = d['number_ser']
        sheet.cell(row=row, column=10).alignment = styles.Alignment(horizontal="center", vertical="center", wrap_text=True)
        for c in range(1, 11):
            sheet.cell(row=row, column=c).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
def insert_sheet_P2(**kwargs):
    sheet = kwargs['sheet']
    # data = kwargs['data']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    filters = kwargs['filters']

    data = list()
    for n,p in enumerate(kwargs['data'],1):
        temp = OrderedDict()
        temp['n'] = n
        temp['fam'] = p.patient.fam
        temp['nib'] = p.sluchay.nib
        temp['im'] = p.patient.im
        temp['ot'] = p.patient.ot
        temp['datr'] = p.patient.datr
        temp['age'] = p.patient_year
        temp['nvs'] = p.patient.nvs
        temp['datp'] = p.sluchay.datp
        temp['tm_otd'] = p.sluchay.tm_otd
        temp['datv'] = p.sluchay.datv
        temp['adr'] = p.patient.m_roj
        temp['rab'] = p.patient.rab
        temp['lpy'] = p.sluchay.lpy.naim[:20] if p.sluchay.lpy else None
        temp['otd'] = p.sluchay.otd.naim if p.sluchay.otd else None
        if p.sluchay.dspat:
            temp['dskz'] = f'{p.sluchay.dspat.kod}-{p.sluchay.dspat.naim}'
        else:
            temp['dskz'] = f'{p.sluchay.dskz.kod}-{p.sluchay.dskz.naim}' if p.sluchay.dskz else None

        temp['ds_osl'] = f'{p.sluchay.ds_osl.kod}-{p.sluchay.ds_osl.naim}' if p.sluchay.ds_osl else None
        temp['dsc'] = f'{p.sluchay.dsc.kod}-{p.sluchay.dsc.naim}' if p.sluchay.dsc else None
        temp['dson'] = f'{p.sluchay.dson.kod}-{p.sluchay.dson.naim}' if p.sluchay.dson else None
        temp['ds_dspat'] =  f'{p.sluchay.dspat.kod}-{p.sluchay.dspat.naim}' if p.sluchay.dspat else None
        temp['vra'] = f'{p.sluchay.le_vr.kod.naim} {p.sluchay.le_vr.kod.ini}' if p.sluchay.le_vr and p.sluchay.le_vr.kod else None
        temp['nib'] = f'N МЭК - {p.sluchay.nib}'
        temp['icx'] = p.sluchay.icx.iz_name if p.sluchay.icx else None
        temp['rslt'] = p.sluchay.rslt.tip_name if p.sluchay.rslt else None
        temp['goc'] = p.sluchay.goc.tip_name if p.sluchay.goc else None
        temp['vds'] = p.sluchay.vds.vds.naim if p.sluchay.vds and p.sluchay.vds.vds else None
        temp['metod_hmp'] = p.sluchay.metod_hmp
        temp['vid_hmp'] = p.sluchay.vid_hmp
        temp['code_usl_vt'] = p.sluchay.code_usl_vt.kod if p.sluchay.code_usl_vt else ''
        temp['ctkom'] = p.sluchay.vds.ctkom.naim if p.sluchay.vds and p.sluchay.vds.ctkom else None
        temp['vrez'] = p.sluchay.vrez.naim if p.sluchay.vrez else None
        if p.sluchay.wskr:
            temp['wskr'] = p.sluchay.wskr_display()
        else:
            temp['wskr'] = None

        opers = list()
        if p.sluchay.oper.count() > 0:
            for o in p.sluchay.oper.all():
                opers.append(o)
        temp['oper'] = opers
        data.append(temp)

    data = sorted(data,key=itemgetter('fam'))
    sheet.cell(row=3, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()
    row_ = 7
    font = styles.Font(size=15, name='Arial')
    if filters:
        filter = []
        height = 0
        for f in filters:
            filter.append(f"{f['filter']}:{f['value']}")
            height +=20
        sheet.row_dimensions[6].height = height
        sheet.cell(row=6, column=1).value = '\n'.join(filter)
    for n, d in enumerate(data, 1):
        row_ += 1
        sheet.row_dimensions[row_].height = 20
        sheet.cell(row=row_, column=1).value = n
        sheet.cell(row=row_, column=1).font = font
        sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=2).value = d['fam']
        sheet.cell(row=row_, column=2).font = font
        sheet.cell(row=row_, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=3).value = f"{d['age']} {d['nvs']}"
        sheet.cell(row=row_, column=3).font = font
        sheet.cell(row=row_, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=4).value = f"{d['datp'].strftime('%d.%m.%Y')}" if d['datp'] else None
        sheet.cell(row=row_, column=4).font = font
        sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=5).value = f"{d['datv'].strftime('%d.%m.%Y')}" if d['datv'] else None
        sheet.cell(row=row_, column=5).font = font
        sheet.cell(row=row_, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.merge_cells(f"F{sheet.cell(row=row_, column=6).row}:F{sheet.cell(row=row_ + 2, column=6).row}")
        sheet.cell(row=row_, column=6).value = d['adr']
        sheet.cell(row=row_, column=6).alignment = styles.Alignment(vertical="top", wrap_text=True)
        sheet.cell(row=row_, column=6).font = font
        sheet.merge_cells(f"G{sheet.cell(row=row_, column=7).row}:G{sheet.cell(row=row_ + 1, column=7).row}")
        sheet.cell(row=row_, column=7).value = d['otd']
        sheet.cell(row=row_, column=7).font = font
        sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
        row_ += 3
        sheet.row_dimensions[row_-2].height = 20
        sheet.cell(row=row_ - 2, column=2).value = f"{d['im']} {d['ot']}"
        sheet.cell(row=row_ - 2, column=2).font = font
        sheet.cell(row=row_ - 2, column=2).alignment = styles.Alignment(vertical="center")
        sheet.cell(row=row_ - 2, column=3).value = f"{d['datr'].strftime('%d.%m.%Y')}" if d['datr'] else None
        sheet.cell(row=row_ - 2, column=3).font = font
        sheet.cell(row=row_ - 2, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_ - 2, column=4).value = d['tm_otd']
        sheet.cell(row=row_ - 2, column=4).font = font
        sheet.cell(row=row_ - 2, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[row_].height = 20
        sheet.cell(row=row_, column=6).value = d['rab']
        sheet.cell(row=row_, column=6).font = font
        sheet.cell(row=row_, column=6).alignment = styles.Alignment(vertical="center")
        sheet.row_dimensions[row_-1].height = 20
        sheet.cell(row=row_ - 1, column=7).value = d['vra']
        sheet.cell(row=row_ - 1, column=7).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_ - 1, column=7).font = font
        row_ += 1
        sheet.row_dimensions[row_].height = 20
        sheet.cell(row=row_, column=2).value = d['nib']
        sheet.cell(row=row_, column=2).font = font
        sheet.cell(row=row_, column=2).alignment = styles.Alignment(vertical="center")
        sheet.merge_cells(f"C{sheet.cell(row=row_, column=3).row}:D{sheet.cell(row=row_, column=4).row}")
        sheet.cell(row=row_, column=3).value = d['icx']
        sheet.cell(row=row_, column=3).font = font
        sheet.merge_cells(f"E{sheet.cell(row=row_, column=5).row}:F{sheet.cell(row=row_, column=6).row}")
        sheet.cell(row=row_, column=5).value = d['rslt']
        sheet.cell(row=row_, column=5).alignment = styles.Alignment(vertical="center")
        sheet.cell(row=row_, column=5).font = font
        sheet.cell(row=row_, column=7).value = d['wskr']
        sheet.cell(row=row_, column=7).font = font
        row_ += 1
        sheet.row_dimensions[row_].height = 20
        sheet.cell(row=row_, column=2).value = d['vrez']
        sheet.cell(row=row_, column=2).alignment = styles.Alignment(vertical="center")
        sheet.cell(row=row_, column=2).font = font
        sheet.cell(row=row_, column=3).value = d['goc']
        sheet.cell(row=row_, column=3).font = font
        sheet.cell(row=row_, column=3).alignment = styles.Alignment(vertical="center")
        sheet.merge_cells(f"D{sheet.cell(row=row_, column=4).row}:F{sheet.cell(row=row_, column=6).row}")
        sheet.cell(row=row_, column=4).value = d['lpy']
        sheet.cell(row=row_, column=4).font = font
        row_ += 2
        sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:B{sheet.cell(row=row_, column=2).row}")
        sheet.merge_cells(f"A{sheet.cell(row=row_ - 1, column=1).row}:B{sheet.cell(row=row_, column=2).row}")
        sheet.row_dimensions[row_-1].height = 25
        sheet.cell(row=row_ - 1, column=1).value = d['vds']
        sheet.cell(row=row_ - 1, column=1).font = font
        sheet.cell(row=row_ - 1, column=1).alignment = styles.Alignment(vertical="center", wrap_text=True)
        sheet.merge_cells(f"C{sheet.cell(row=row_ - 1, column=3).row}:E{sheet.cell(row=row_, column=5).row}")
        sheet.merge_cells(f"C{sheet.cell(row=row_, column=3).row}:E{sheet.cell(row=row_, column=5).row}")
        sheet.cell(row=row_ - 1, column=3).value = d['ctkom']
        sheet.cell(row=row_ - 1, column=3).alignment = styles.Alignment(vertical="center", wrap_text=True)
        sheet.cell(row=row_ - 1, column=3).font = font
        if d['dskz']:
            sheet.merge_cells(f"F{sheet.cell(row=row_ - 1, column=6).row}:G{sheet.cell(row=row_, column=7).row}")
            sheet.row_dimensions[row_ - 1].height = 20
            sheet.cell(row=row_ - 1, column=6).value = f"(осн.){d['dskz']}"
            sheet.cell(row=row_ - 1, column=6).alignment = styles.Alignment(vertical="center", wrap_text=True)
            sheet.cell(row=row_ - 1, column=6).font = font
        if d['dsc']:
            row_ += 1
            sheet.merge_cells(f"F{sheet.cell(row=row_, column=6).row}:G{sheet.cell(row=row_, column=7).row}")
            sheet.row_dimensions[row_].height = 20
            sheet.cell(row=row_, column=6).value = f"(спо.){d['dsc']}"
            sheet.cell(row=row_, column=6).alignment = styles.Alignment(vertical="center")
            sheet.cell(row=row_, column=6).font = font
        if d['ds_osl']:
            row_ += 1
            sheet.merge_cells(f"F{sheet.cell(row=row_, column=6).row}:G{sheet.cell(row=row_, column=7).row}")
            sheet.row_dimensions[row_].height = 20
            sheet.cell(row=row_, column=6).value = f"(осл.){d['ds_osl']}"
            sheet.cell(row=row_, column=6).alignment = styles.Alignment(vertical="center")
            sheet.cell(row=row_, column=6).font = font
        if d['dson']:
            row_ += 1
            sheet.merge_cells(f"F{sheet.cell(row=row_, column=6).row}:G{sheet.cell(row=row_, column=7).row}")
            sheet.row_dimensions[row_].height = 20
            sheet.cell(row=row_, column=6).value = f"(онк.){d['dson']}"
            sheet.cell(row=row_, column=6).alignment = styles.Alignment(vertical="center")
            sheet.cell(row=row_, column=6).font = font
        if len(d['oper']) > 0:
            if d['vds'] in ['ВТМП баз.программа ОМС','ВТМП сверхбаз.программа']:
                row_ += 1
                sheet.merge_cells(f"B{sheet.cell(row=row_, column=2).row}:E{sheet.cell(row=row_, column=5).row}")
                sheet.cell(row=row_, column=2).value = f"{d['metod_hmp']} - {d['vid_hmp']} : {d['code_usl_vt']}"
                sheet.cell(row=row_, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=2).font = font
                row_ -= 1
            for o in d['oper']:
                row_ += 1
                sheet.merge_cells(f"F{sheet.cell(row=row_, column=3).row}:G{sheet.cell(row=row_, column=4).row}")
                sheet.row_dimensions[row_].height = 20
                sheet.cell(row=row_,column=6).value = f"{o.dato.strftime('%d.%m.%Y') if o.dato else None} {o.kod_op.kod if o.kod_op else ''} {o.kod_op.naim if o.kod_op else ''}"
                sheet.cell(row=row_, column=6).alignment = styles.Alignment(vertical="center")
                sheet.cell(row=row_, column=6).font = font
        else:
            if d['vds'] in ['ВТМП баз.программа ОМС','ВТМП сверхбаз.программа']:
                row_ += 1
                sheet.merge_cells(f"B{sheet.cell(row=row_, column=2).row}:E{sheet.cell(row=row_, column=5).row}")
                sheet.cell(row=row_, column=2).value = f"{temp['metod_hmp']} - {temp['vid_hmp']} : {temp['code_usl_vt']}"
                sheet.cell(row=row_, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=2).font = font
        sheet.cell(row=row_, column=1).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row_, column=2).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row_, column=3).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row_, column=4).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row_, column=5).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row_, column=6).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row_, column=7).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))

def insert_sheet_P19(**kwargs):
    sheet = kwargs['sheet']
    # data = kwargs['data']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = []
    for p in kwargs['data']:
        ord = OrderedDict()
        ord['fio'] = f'{p.patient.fam} {p.patient.im} {p.patient.ot}'
        ord['datr'] = p.patient.datr.strftime('%d.%m.%Y') if p.patient.datr else None
        ord['adr'] = p.patient.m_roj
        ord['datp'] = p.sluchay.datp.strftime('%d.%m.%Y') if p.sluchay.datp else None
        ord['goc'] = p.sluchay.goc.tip_name if p.sluchay.goc else None
        ord['pri'] = p.sluchay.pri.naim if p.sluchay.pri else None
        ord['lpy'] = p.sluchay.lpy.naim[:20] if p.sluchay.lpy else None
        data.append(ord)

    sheet.cell(row=4, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()
    row_ = 7
    font = styles.Font(size=14, name='Arial')
    for n, d in enumerate(data, 1):
        row_ += 1
        sheet.row_dimensions[row_].height = 65
        sheet.cell(row=row_, column=1).value = n
        sheet.cell(row=row_, column=1).font = font
        sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=1).border = border
        sheet.cell(row=row_, column=2).value = d['fio']
        sheet.cell(row=row_, column=2).font = font
        sheet.cell(row=row_, column=2).alignment = styles.Alignment(horizontal="left",vertical="center",wrap_text=True)
        sheet.cell(row=row_, column=2).border = border
        sheet.cell(row=row_, column=3).value = d['datr']
        sheet.cell(row=row_, column=3).border = border
        sheet.cell(row=row_, column=3).font = font
        sheet.cell(row=row_, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=4).value = d['adr']
        sheet.cell(row=row_, column=4).font = font
        sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="left",wrap_text=True)
        sheet.cell(row=row_, column=4).border = border
        sheet.cell(row=row_, column=5).value = d['datp']
        sheet.cell(row=row_, column=5).font = font
        sheet.cell(row=row_, column=5).border = border
        sheet.cell(row=row_, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=6).value = d['goc']
        sheet.cell(row=row_, column=6).font = font
        sheet.cell(row=row_, column=6).border = border
        sheet.cell(row=row_, column=6).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=7).value = d['pri']
        sheet.cell(row=row_, column=7).font = font
        sheet.cell(row=row_, column=7).border = border
        sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        sheet.cell(row=row_, column=8).value = d['lpy']
        sheet.cell(row=row_, column=8).border = border
        sheet.cell(row=row_, column=8).font = font
        sheet.cell(row=row_, column=8).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
def insert_sheet_P22(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    otd = kwargs['otd']
    data = []
    otd_list = []
    otdels = kwargs['otdels']
    for p in kwargs['data']:
        ord = OrderedDict()
        ord['fio_nib'] = f'{p.patient.fam} {p.patient.im} {p.patient.ot} {p.sluchay.nib}'
        ord['datp'] = p.sluchay.datp.strftime('%d.%m.%Y') if p.sluchay.datp else None
        ord['datv'] = p.sluchay.datv.strftime('%d.%m.%Y') if p.sluchay.datv else None
        if p.sluchay.vds:
            if p.sluchay.vds.vds and p.sluchay.vds.vds.kod == 'Д':
                ord['vds'] = 'ВМП БП'
            elif p.sluchay.vds.vds and p.sluchay.vds.vds.kod == '5':
                ord['vds'] = 'ВМП СБ'
            else:
                ord['vds'] = None
        else:
            ord['vds'] = None
        ord['vra_fio'] = f'{p.sluchay.le_vr.kod.kod} {p.sluchay.le_vr.kod.naim} {p.sluchay.le_vr.kod.ini}' \
            if p.sluchay.le_vr and p.sluchay.le_vr.kod else None
        if ord['vds'] == 'ВМП БП':
            ord['metod_vid'] = f'1.{p.sluchay.metod_hmp}.{p.sluchay.vid_hmp}'
        elif ord['vds'] == 'ВМП СБ':
            ord['metod_vid'] = f'2.{p.sluchay.metod_hmp}.{p.sluchay.vid_hmp}'
        else:
            ord['metod_vid'] = None
        oper = get_pop_oper(p)
        ord['oper_kod'] = oper.kod_op.kod if oper is not None and oper.kod_op else None
        ord['vra_oper'] = f'{oper.kodx.kod} {oper.kodx.naim} {oper.kodx.ini}' if oper is not None and oper.kodx else None
        ord['otd'] = p.sluchay.otd.naim if p.sluchay.otd else None
        otd_list.append(p.sluchay.otd.naim if p.sluchay.otd else None)
        ord['an'] = f'{oper.kodan.kod} {oper.kodan.naim} {oper.kodan.ini}' if oper is not None and oper.kodan else None
        data.append(ord)
    row_ = 6
    font = styles.Font(size=14, name='Arial')
    sheet.cell(row=3, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    # if otd:
    #     sheet.cell(row=5, column=1).value = f'Отделение - {otd}'
    
    otd_list.sort()
    for en,o in enumerate(set(otd_list)):
        nn=0
        # if en != 0:
        d = ['N п/п','Фамилия И.О. пациента (N ист.)',' Дата пост.    Дата выбыт.','Источник финанс-я','Таб.N ФИО леча-щего врача',
        'Код вида ВТМП метод','Код операции','Таб.N ФИО Хирурга','Отделение','Таб.N ФИО анестези.']
        if en == 0:
            row_+=1
        else:
            row_ +=3
        sheet.row_dimensions[row_].height = 60
        for n,f in enumerate(d,1):
            sheet.cell(row=row_, column=n).value = f
            sheet.cell(row=row_, column=n).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
            sheet.cell(row=row_, column=n).font = font
            sheet.cell(row=row_,column=n).border = border
        else:

            sheet.merge_cells(f"A{sheet.cell(row=row_-1, column=1).row}:J{sheet.cell(row=row_ -1, column=13).row}")
            sheet.cell(row=row_-1, column=1).value = o
            sheet.cell(row=row_-1, column=1).alignment =  styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row_-1, column=1).font = font


        for n,d in enumerate(data,1):
            if o == d['otd']:
                nn+=1
                row_ += 1
                sheet.row_dimensions[row_].height = 45
                sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:A{sheet.cell(row=row_ + 1, column=1).row}")
                sheet.cell(row=row_, column=1).value = nn
                sheet.cell(row=row_, column=1).font = font
                sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.merge_cells(f"B{sheet.cell(row=row_, column=2).row}:B{sheet.cell(row=row_ + 1, column=2).row}")
                sheet.cell(row=row_, column=2).value = d['fio_nib']
                sheet.cell(row=row_, column=2).alignment = styles.Alignment(wrap_text=True)
                sheet.cell(row=row_, column=2).font = font
                sheet.cell(row=row_, column=3).value = d['datp']
                sheet.cell(row=row_, column=3).alignment =  styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_, column=3).font = font
                sheet.cell(row=row_ + 1,column=3).value = d['datv']
                sheet.cell(row=row_ + 1,column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row_ + 1, column=3).font = font
                sheet.merge_cells(f"D{sheet.cell(row=row_, column=4).row}:D{sheet.cell(row=row_ + 1, column=4).row}")
                sheet.cell(row=row_, column=4).value = d['vds']
                sheet.cell(row=row_, column=4).font = font
                sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.merge_cells(f"E{sheet.cell(row=row_, column=5).row}:E{sheet.cell(row=row_ + 1, column=5).row}")
                sheet.cell(row=row_, column=5).value = d['vra_fio']
                sheet.cell(row=row_, column=5).font = font
                sheet.cell(row=row_, column=5).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
                sheet.merge_cells(f"F{sheet.cell(row=row_, column=6).row}:F{sheet.cell(row=row_ + 1, column=6).row}")
                sheet.cell(row=row_, column=6).value = d['metod_vid']
                sheet.cell(row=row_, column=6).font = font
                sheet.cell(row=row_, column=6).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.merge_cells(f"G{sheet.cell(row=row_, column=7).row}:G{sheet.cell(row=row_ + 1, column=7).row}")
                sheet.cell(row=row_, column=7).value = d['oper_kod']
                sheet.cell(row=row_, column=7).font = font
                sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.merge_cells(f"H{sheet.cell(row=row_, column=8).row}:H{sheet.cell(row=row_ + 1, column=8).row}")
                sheet.cell(row=row_, column=8).value = d['vra_oper']
                sheet.cell(row=row_, column=8).font = font
                sheet.cell(row=row_, column=8).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
                sheet.merge_cells(f"I{sheet.cell(row=row_, column=9).row}:I{sheet.cell(row=row_ + 1, column=9).row}")
                sheet.cell(row=row_, column=9).font = font
                sheet.cell(row=row_, column=9).value = d['otd']
                sheet.cell(row=row_, column=9).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
                sheet.merge_cells(f"J{sheet.cell(row=row_, column=10).row}:J{sheet.cell(row=row_ + 1, column=10).row}")
                sheet.cell(row=row_, column=10).font = font
                sheet.cell(row=row_, column=10).value = d['an']
                sheet.cell(row=row_, column=10).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
                for c in range(1,11):
                    sheet.cell(row=row_+1, column=c).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
                row_+=1
def insert_sheet_P24(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    #kwargs['data']
    data = []
    for n,p in enumerate(kwargs['data'],1):
        ord = OrderedDict()
        ord['n'] = n
        ord['nib'] = p.sluchay.nib
        ord['fio'] = f'{p.patient.fam} {p.patient.im} {p.patient.ot}'
        ord['datr'] = p.patient.datr.strftime('%d.%m.%Y') if p.patient.datr else None
        ord['datv'] = p.sluchay.datv.strftime('%d.%m.%Y') if p.sluchay.datv else None
        ord['icx'] = p.sluchay.icx.iz_name if p.sluchay.icx else None
        ord['adr'] = p.patient.m_roj
        data.append(ord)
    row = 7
    font = styles.Font(size=14, name='Arial')
    for d in data:
        row+=1
        sheet.row_dimensions[row].height = 33
        for n,i in enumerate(d.values(),1):
            sheet.cell(row=row, column=n).value = i
            if n == 1:
                sheet.cell(row=row, column=n).alignment = styles.Alignment(horizontal="center", vertical="center")
            elif n == 7:
                sheet.cell(row=row, column=n).alignment = styles.Alignment(horizontal="left",vertical="center",wrap_text=True)
            else:
                sheet.cell(row=row, column=n).alignment = styles.Alignment(horizontal="left", vertical="center")
            sheet.cell(row=row, column=n).font = font
            sheet.cell(row=row, column=n).border = border
def insert_sheet_nnn(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = []
    ym = 0
    s1 = 0
    for p in kwargs['data']:
        ord = OrderedDict()
        if p.sluchay.otd_y and p.sluchay.otd_y.naim in ['АРО N1','АРО N2','АРО N3 (ЛДО)','ПРИЕМНОЕ']:
            ym +=1
        if p.sluchay.icx.id_iz == 106:
            s1 +=1
        ord['fio'] = f'{p.sluchay.nib} {p.patient.fam} {p.patient.im[0] if len(p.patient.im) > 0 else ""}.{p.patient.ot[0] if len(p.patient.ot) > 0 else ""}'
        # ord['prof_k'] = p.sluchay.le_vr.prof_k.k_prname if p.sluchay.le_vr and p.sluchay.le_vr.prof_k else None
        ord['prof_k'] = p.sluchay.otd.naim if p.sluchay.otd else None
        ord['age'] = f'{p.patient_year} {p.patient.nvs}'
        ord['datp'] = p.sluchay.datp.strftime('%d.%m.%Y') if p.sluchay.datp else None
        ord['datv'] = p.sluchay.datv.strftime('%d.%m.%Y') if p.sluchay.datv else None
        if p.sluchay.goc and p.sluchay.goc.tip_name == 'Экстренная':
            ord['goc'] = 'экс'
        elif p.sluchay.goc and p.sluchay.goc.tip_name == 'Плановая':
            ord['goc'] = 'пл'
        else:
            ord['goc'] = ''
        ord['vra'] = p.sluchay.le_vr.kod.naim if p.sluchay.le_vr and p.sluchay.le_vr.kod else None
        ord['dskz'] = f'{p.sluchay.dskz.kod}-{p.sluchay.dskz.naim}' if p.sluchay.dskz else ''
        ord['dspat'] = f'{p.sluchay.dspat.kod}-{p.sluchay.dspat.naim}' if p.sluchay.dspat else ''
        ord['pri'] = p.sluchay.pri.naim if p.sluchay.pri else ''
        opers = list()
        if p.sluchay.oper.count() > 0:
            for o in p.sluchay.oper.all().order_by('-pop'):
                opers.append(o)
        ord['oper'] = opers
        # oper = get_pop_oper(p)
        # ord['dato_oper'] = oper.dato.strftime('%d.%m.%Y') if oper is not None else None
        # oper['goc'] = p.sluchay.goc.tip_name if p.sluchay.goc else None
        # ord['vra_oper'] = oper.kod.naim if oper is not None and oper.kod else None
        data.append(ord)

    sheet.cell(row=3, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()
    row_ = 8
    font = styles.Font(size=11, name='Arial')
    for n, d in enumerate(data, 1):
        row_ += 1
        sheet.row_dimensions[row_].height = 50
        sheet.cell(row=row_, column=1).value = f'{n}   {d["fio"]}'
        sheet.cell(row=row_, column=1).font = font
        sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        sheet.cell(row=row_, column=2).value = d['age']
        sheet.cell(row=row_, column=2).font = font
        sheet.cell(row=row_, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=3).value = d['datp']
        sheet.cell(row=row_, column=3).font = font
        sheet.cell(row=row_, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=4).value = d['datv']
        sheet.cell(row=row_, column=4).font = font
        sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=5).value = d['goc']
        sheet.cell(row=row_, column=5).font = font
        sheet.cell(row=row_, column=5).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
        sheet.cell(row=row_, column=6).value = d['vra']
        sheet.cell(row=row_, column=6).font = font
        sheet.cell(row=row_, column=6).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row_, column=7).value = f'1.{d["dskz"]}'
        sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        sheet.cell(row=row_, column=7).font = font
        row_+=1
        sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:F{sheet.cell(row=row_, column=6).row}")
        sheet.row_dimensions[row_].height = 50
        sheet.cell(row=row_, column=1).value = d['prof_k']
        sheet.cell(row=row_, column=1).font = font
        sheet.cell(row=row_, column=1).alignment = styles.Alignment(horizontal="left",vertical="center",wrap_text=True)
        sheet.cell(row=row_, column=7).value = f'2.{d["dspat"]}'
        sheet.cell(row=row_, column=7).font = font
        sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        row_+=1
        sheet.cell(row=row_, column=7).value = f'3.{d["pri"]}'
        sheet.cell(row=row_, column=7).font = font
        sheet.cell(row=row_, column=7).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        sheet.row_dimensions[row_].height = 35
        if len(d['oper']) > 0:
            for o in d['oper']:
                row_+=1
                sheet.cell(row=row_,column=4).value = o.dato.strftime('%d.%m.%Y') if o.dato else ''
                sheet.cell(row=row_, column=4).font = font
                sheet.cell(row=row_, column=4).alignment = styles.Alignment(horizontal="left")
                sheet.cell(row=row_, column=6).value = o.kodx.naim if o.kodx else ''
                sheet.cell(row=row_, column=6).font = font
                sheet.cell(row=row_, column=6).alignment = styles.Alignment(horizontal="left")
                sheet.cell(row=row_, column=7).value = f'4.{o.kod_op.kod if o.kod_op else ""}{"+" if o.pop == True else ""}'
                sheet.cell(row=row_, column=7).font = font
        for c in range(1,8):
            sheet.cell(row=row_, column=c).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))

    row_ += 3
    sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:B{sheet.cell(row=row_, column=2).row}")
    sheet.cell(row=row_, column=1).value = f'Умерло всего - {len(data)}'
    sheet.cell(row=row_, column=1).font = font
    row_+=1
    sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:B{sheet.cell(row=row_, column=2).row}")
    sheet.cell(row=row_, column=1).value = f'в том числе'
    sheet.cell(row=row_, column=1).font = font
    row_ += 1
    sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:C{sheet.cell(row=row_, column=3).row}")
    sheet.cell(row=row_, column=1).value = f'в реанимации и прием.отд-ии - {ym}'
    sheet.cell(row=row_, column=1).font = font
    row_ += 1
    sheet.merge_cells(f"A{sheet.cell(row=row_, column=1).row}:B{sheet.cell(row=row_, column=2).row}")
    sheet.cell(row=row_, column=1).value = f'в 1-е сутки - {s1}'
    sheet.cell(row=row_, column=1).font = font
def insert_sheet_a_oth_5(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    user = kwargs['user']
    filters = kwargs['filters']

    data = get_list_otd(kwargs['data'])
    data_filter = get_list_otd(kwargs['data_filter'])


    row = 8
    font = styles.Font(size=14, name='Arial')
    sheet.cell(row=3, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()

    if filters:
        filter = []
        height = 0
        for f in filters:
            filter.append(f"{f['filter']}:{f['value']}")
            height += 20
        sheet.row_dimensions[6].height = height
        sheet.cell(row=6, column=1).value = '\n'.join(filter)

    all_sl = []
    all_sl_filter = []
    for d in data:
        row+=1
        rez_data = get_rez_a_oth_5(d[1])
        all_sl.append(rez_data) 
        sheet.cell(row=row, column=1).value = d[0][0]
        fields_data = [i for i in range(1,15) if i % 2 == 1]
        for f in range(len(fields_data)):
            sheet.cell(row=row, column=1+fields_data[f]).value = rez_data[f] if rez_data[f] != 0 else None
            sheet.cell(row=row, column=1+fields_data[f]).alignment = styles.Alignment(horizontal="center", vertical="center")
        if len(data_filter) > 0:
            for df in data_filter:
                if d[0] == df[0]:
                    rez_filter = get_rez_a_oth_5(df[1])
                    all_sl_filter.append(rez_filter) 
                    fields_data = [i for i in range(1,15) if i % 2 == 0]
                    for f in range(len(fields_data)):
                        sheet.cell(row=row, column=1+fields_data[f]).value = rez_filter[f] if rez_filter[f] != 0 else None
                        sheet.cell(row=row, column=1+fields_data[f]).alignment = styles.Alignment(horizontal="center", vertical="center")
    else:
        age_data = 0
        age_data_filter = 0
        row+=1
        sheet.cell(row=row, column=1).value = 'ВСЕГО'
        if len(all_sl) > 0:
            r = None
            for o in range(len(all_sl)):
                if o == 0:
                    r = numpy.array(all_sl[o])
                else:
                    r+= numpy.array(all_sl[o])
            rez_data = r.tolist()
            fields_data = [i for i in range(1,15) if i % 2 == 1]
            age_data = float('{0:.2f}'.format(rez_data[7]/rez_data[0]))
            for f in range(len(fields_data)):
                sheet.cell(row=row, column=1+fields_data[f]).value = rez_data[f] if rez_data[f] != 0 else None
                sheet.cell(row=row, column=1+fields_data[f]).alignment = styles.Alignment(horizontal="center", vertical="center")
            if len(all_sl_filter) != 0:
                r = None
                for o in range(len(all_sl_filter)):
                    if o == 0:
                        r = numpy.array(all_sl_filter[o])
                    else:
                        r+= numpy.array(all_sl_filter[o])
                rez_filter = r.tolist()
                age_data_filter = float('{0:.2f}'.format(rez_filter[7]/rez_filter[0]))
                fields_data = [i for i in range(1,15) if i % 2 == 0]
                for f in range(len(fields_data)):
                    sheet.cell(row=row, column=1+fields_data[f]).value = rez_filter[f] if rez_filter[f] != 0 else None
                    sheet.cell(row=row, column=1+fields_data[f]).alignment = styles.Alignment(horizontal="center", vertical="center")
            row+=1
            sheet.cell(row=row, column=1).value = 'Средний возраст (кол-во лет)'
            sheet.cell(row=row, column=2).value = age_data if age_data != 0 else None
            sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row, column=3).value = age_data_filter if age_data_filter != 0 else None
            sheet.cell(row=row, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")



def get_rez_a_oth_5(data):
    bf = BetterFilter()
    sp = CountSluchaySpecification() ^ GocEkSpecification() ^ ProfKNSpecification() ^ \
         OperCountSpecification() ^ RezUmerSpecification() ^ PoslOperKdSpecification() ^ AgeSpecification()
    all_temp = []
    for patient in data:
        for p in bf.filter(patient,sp):
            temp = bf.format_list(p)
            for t in range(len(temp)):
                if temp[t] == 'None':
                    temp[t] = 0
            all_temp.append([int(i) for i in temp])
    all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(7)]
    all_temp.insert(3,0)
    all_temp[3] = float('{0:.2f}'.format(all_temp[2]/all_temp[0]))
    return all_temp

def insert_sheet_a_oth_7(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    user = kwargs['user']
    sheet.cell(row=4, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=6, column=1).value = str(name).capitalize()    
    data = get_list_otd_prof(kwargs['data'])

    row = 9
    font = styles.Font(size=10, name='Arial')
    fs = ['Выбыло','Койко/день','Сред.к/день','Оперировано','К-во операц','Умерло']
    #Всего пациентов
    sl_all = []

    for d in data:
        for b in range(1,13):
            sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        row+=1
        sheet.row_dimensions[row].height = 25
        sheet.cell(row=row, column=1).value = d[0][0]
        sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        #Список итог всех пациентов в отделении
        otd_prof_sl_all = []
        for prof in d[1]:
            sheet.cell(row=row, column=2).value = prof[0]
            sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
            rez = get_rez_7(prof[1])
            otd_prof_sl_all.append(rez)
            sl_all.append(rez)
            for f in range(len(fs)):
                sheet.cell(row=row, column=3).value = fs[f]
                for b in range(3,13):
                    sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
                for v in range(len(rez[f])):
                    sheet.cell(row=row, column=4+v).value = rez[f][v] if rez[f][v] != 0 else None
                    sheet.cell(row=row, column=4+v).alignment = styles.Alignment(horizontal="center", vertical="center")
                row+=1
            else:
                row-=1
                for b in range(2,13):
                    sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))

            if len(d[1]) > 1:
                row+=1
        else:
            if len(d[1]) > 1:
                r = None
                for o in range(len(otd_prof_sl_all)):
                    if o == 0:
                        r = numpy.array(otd_prof_sl_all[o])
                    else:
                        r+= numpy.array(otd_prof_sl_all[o])
                rez = r.tolist()
                rez[2] = (numpy.array(rez[1])/numpy.array(rez[0])).tolist()
                rez[2][1] = rez[2][2] + rez[2][3] + rez[2][4] 

                sheet.cell(row=row, column=2).value = 'Итого'
                for f in range(len(fs)):
                    sheet.cell(row=row, column=3).value = fs[f]
                    for b in range(3,13):
                        sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
                    for v in range(len(rez[f])):
                        sheet.cell(row=row, column=4+v).value =  float('{0:.2f}'.format(rez[f][v])) if rez[f][v] != None and rez[f][v] != 0 else None
                        sheet.cell(row=row, column=4+v).alignment = styles.Alignment(horizontal="center", vertical="center")
                    row+=1
                else:
                    row-=1
                    for b in range(2,13):
                        sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
    else:
        row+=1
        for b in range(1,13):
            sheet.cell(row=row, column=b).border = styles.Border(top=styles.Side(border_style='thin', color='000000'))
        if len(sl_all)>0:
            r = None
            for o in range(len(sl_all)):
                if o == 0:
                    r = numpy.array(sl_all[o])
                else:
                    r+= numpy.array(sl_all[o])
            
            rez = r.tolist()
            rez[2] = (numpy.array(rez[1])/numpy.array(rez[0])).tolist()
            rez[2][1] = rez[2][2] + rez[2][3] + rez[2][4] 

            sheet.cell(row=row, column=1).value = 'Всего'
            for f in range(len(fs)):
                    sheet.cell(row=row, column=3).value = fs[f]
                    for b in range(3,13):
                        sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
                    for v in range(len(rez[f])):
                        sheet.cell(row=row, column=4+v).value =  float('{0:.2f}'.format(rez[f][v])) if rez[f][v] != None and rez[f][v] != 0 else None
                        sheet.cell(row=row, column=4+v).alignment = styles.Alignment(horizontal="center", vertical="center")
                    row+=1

def get_rez_7(data):
    bf = BetterFilter()
    all_sp = IsfinKastSpecification('all') ^ IsfinKastSpecification('all','tym_oms') ^ IsfinKastSpecification('all','bez_polis') ^ IsfinKastSpecification('all','dr_oms') ^ \
        IsfinKastSpecification('all','vmp_sb') ^ IsfinKastSpecification('all','vmp_sv') ^ IsfinKastSpecification('all','dms') ^ IsfinKastSpecification('all','fss') 
    all_temp = []
    for patient in data:
        for p in bf.filter(patient,all_sp):
            all_temp.append([int(i) for i in bf.format_list(p)])
    all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(8)]

    kd_sp = IsfinKastSpecification('kd') ^ IsfinKastSpecification('kd','tym_oms') ^ IsfinKastSpecification('kd','bez_polis') ^ IsfinKastSpecification('kd','dr_oms') ^ \
        IsfinKastSpecification('kd','vmp_sb') ^ IsfinKastSpecification('kd','vmp_sv') ^ IsfinKastSpecification('kd','dms') ^ IsfinKastSpecification('kd','fss') 
    kd_temp = []
    for patient in data:
        for p in bf.filter(patient,kd_sp):
            temp = temp = bf.format_list(p)
            for t in range(len(temp)):
                if temp[t] == 'None':
                    temp[t] = 0
            kd_temp.append([int(i) for i in temp])
            # kd_temp.append([int(i) for i in bf.format_list(p)])


    kd_temp = [sum([kd_temp[i][x] for i in range(len(kd_temp))]) for x in range(8)]

    oper_sp = IsfinKastSpecification('oper') ^ IsfinKastSpecification('oper','tym_oms') ^ IsfinKastSpecification('oper','bez_polis') ^ IsfinKastSpecification('oper','dr_oms') ^ \
        IsfinKastSpecification('oper','vmp_sb') ^ IsfinKastSpecification('oper','vmp_sv') ^ IsfinKastSpecification('oper','dms') ^ IsfinKastSpecification('oper','fss') 
    oper_temp = []
    for patient in data:
        for p in bf.filter(patient,oper_sp):
            oper_temp.append([int(i) for i in bf.format_list(p)])
    oper_temp = [sum([oper_temp[i][x] for i in range(len(oper_temp))]) for x in range(8)]

    oper_count_sp = IsfinKastSpecification('oper_count') ^ IsfinKastSpecification('oper_count','tym_oms') ^ IsfinKastSpecification('oper_count','bez_polis') ^ IsfinKastSpecification('oper_count','dr_oms') ^ \
        IsfinKastSpecification('oper_count','vmp_sb') ^ IsfinKastSpecification('oper_count','vmp_sv') ^ IsfinKastSpecification('oper_count','dms') ^ IsfinKastSpecification('oper_count','fss') 
    oper_count_temp = []
    for patient in data:
        for p in bf.filter(patient,oper_count_sp):
            oper_count_temp.append([int(i) for i in bf.format_list(p)])
    oper_count_temp = [sum([oper_count_temp[i][x] for i in range(len(oper_count_temp))]) for x in range(8)]

    ymer_sp = IsfinKastSpecification('ymer') ^ IsfinKastSpecification('ymer','tym_oms') ^ IsfinKastSpecification('ymer','bez_polis') ^ IsfinKastSpecification('ymer','dr_oms') ^ \
        IsfinKastSpecification('ymer','vmp_sb') ^ IsfinKastSpecification('ymer','vmp_sv') ^ IsfinKastSpecification('ymer','dms') ^ IsfinKastSpecification('ymer','fss') 
    ymer_temp = []
    for patient in data:
        for p in bf.filter(patient,ymer_sp):
            ymer_temp.append([int(i) for i in bf.format_list(p)])
    ymer_temp = [sum([ymer_temp[i][x] for i in range(len(ymer_temp))]) for x in range(8)]
    medium_kd = []
    for k in range(len(kd_temp)):
        try:
            medium_kd.append(float('{0:.2f}'.format(kd_temp[k]/all_temp[k])))
        except ZeroDivisionError:
            medium_kd.append(0)

    # all_temp.insert(1,all_temp[2]+all_temp[3]+all_temp[4])
    # kd_temp.insert(1,kd_temp[2]+kd_temp[3]+kd_temp[4])
    # medium_kd.insert(1,medium_kd[2]+medium_kd[3]+medium_kd[4])
    # oper_temp.insert(1,oper_temp[2]+oper_temp[3]+oper_temp[4])
    # oper_count_temp.insert(1,oper_count_temp[2]+oper_count_temp[3]+oper_count_temp[4])
    # ymer_temp.insert(1,ymer_temp[2]+ymer_temp[3]+ymer_temp[4])

    all_temp.insert(1,0)
    kd_temp.insert(1,0)
    medium_kd.insert(1,0)
    oper_temp.insert(1,0)
    oper_count_temp.insert(1,0)
    ymer_temp.insert(1,0)

    all_temp[1] = all_temp[2]+all_temp[3]+all_temp[4]
    kd_temp[1] = kd_temp[2]+kd_temp[3]+kd_temp[4]
    medium_kd[1] = medium_kd[2]+medium_kd[3]+medium_kd[4]
    oper_temp[1] = oper_temp[2]+oper_temp[3]+oper_temp[4]
    oper_count_temp[1] = oper_count_temp[2]+oper_count_temp[3]+oper_count_temp[4]
    ymer_temp[1] = ymer_temp[2]+ymer_temp[3]+ymer_temp[4]
    


    return [all_temp,kd_temp,medium_kd,oper_temp,oper_count_temp,ymer_temp]

def insert_sheet_a_oth_19(**kwargs):
    sheet = kwargs['sheet']
    data = kwargs['data']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    list_data_1 = [0,0,0,0,0,0,0,0,0,0]
    list_data_2 = [0,0,0,0,0,0,0,0,0,0,0,0,0]

    sheet.cell(row=4, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()

    sheet.cell(row=26, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=27, column=1).value = str(name).capitalize()    

    for d in data:
        ds = d.sluchay.dskz.kod if d.sluchay.dskz else None
        if ds in T8:
            list_data_1[0] += 1
            if 0 <= d.patient_year <= 17:
                list_data_1[1] += 1
        if d.sluchay.oper.count() > 0:
            list_data_1[2] += 1
            if ds in T7:
                list_data_1[3] += 1
            if ds in T8:
                list_data_1[4] += 1
            if ds in T1 or ds in T2:
                list_data_1[5] += 1
        if d.sluchay.rslt and d.sluchay.rslt.id_tip  in [105, 106]:
            list_data_1[6] += 1
            if 0 <= d.patient_year <= 17:
                list_data_1[7] += 1
            try:
                day = d.sluchay.datv - d.sluchay.datp
                if 0 <= day.days <= 30:
                    list_data_1[8] += 1
                if 0 <= day.days <= 7:
                    list_data_1[9] += 1

            except:
                pass

        if d.sluchay.rslt and d.sluchay.rslt.id_tip  not in [105, 106]:        
            list_data_2[0]+=1
            if d.patient.pol and d.patient.pol.id_pol == 1:
                if d.patient_year >= settings.OLD_M:
                    list_data_2[1] += 1
            if d.patient.pol and d.patient.pol.id_pol == 2:
                if d.patient_year >= settings.OLD_G:
                    list_data_2[1] += 1
            if 0 <= d.patient_year <= 18:
                list_data_2[2] += 1
                
        if d.sluchay.goc:
            list_data_2[3] += 1
            if d.patient.pol and d.patient.pol.id_pol == 1:
                if d.patient_year >= settings.OLD_M:
                    list_data_2[4] += 1
            if d.patient.pol and d.patient.pol.id_pol == 2:
                if d.patient_year >= settings.OLD_G:
                    list_data_2[4] += 1
        adr = TerrSpecification()
        adr.t = 'Др.регионы Россий'
        adr.c = 643
        result = adr.is_satisfied(d)
        if result != 0:
            list_data_2[5] += 1
        if d.sluchay.rslt and d.sluchay.rslt.id_tip  in [105, 106]:
            list_data_2[6] += 1
            if d.patient.pol and d.patient.pol.id_pol == 1:
                if d.patient_year >= settings.OLD_M:
                    list_data_2[7] += 1
            if d.patient.pol and d.patient.pol.id_pol == 2:
                if d.patient_year >= settings.OLD_G:
                    list_data_2[7] += 1
            list_data_2[8] += 1
            if d.patient.pol and d.patient.pol.id_pol == 1:
                if d.patient_year >= settings.OLD_M:
                    list_data_2[9] += 1
            if d.patient.pol and d.patient.pol.id_pol == 2:
                if d.patient_year >= settings.OLD_G:
                    list_data_2[9] += 1
            if result != 0:
                list_data_2[10]
        
        try:
            day = d.sluchay.datv - d.sluchay.datp
            if 1 == day.days:
                list_data_1[11] += 1
                if d.patient.pol and d.patient.pol.id_pol == 1:
                    if d.patient_year >= settings.OLD_M:
                        list_data_2[12] += 1
                if d.patient.pol and d.patient.pol.id_pol == 2:
                    if d.patient_year >= settings.OLD_G:
                        list_data_2[12] += 1
        except:
            pass

    row = 9
    for l1 in list_data_1:
        row+=1
        sheet.cell(row=row, column=4).value = l1 if l1 != 0 else None
    
    row = 30
    for l1 in list_data_2:
        row+=1
        sheet.cell(row=row, column=3).value = l1 if l1 != 0 else None



def insert_sheet_a_oth_20(**kwargs):
    sheet = kwargs['sheet']
    # data = kwargs['data']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    filters = kwargs['filters']

    sheet.cell(row=3, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()

    if filters:
        filter = []
        height = 0
        for f in filters:
            filter.append(f"{f['filter']}:{f['value']}")
            height += 20
        sheet.row_dimensions[6].height = height
        sheet.cell(row=6, column=1).value = '\n'.join(filter)

    bf = BetterFilter()
    sp = OtdSpecification() ^ ProfkSpecification() ^ OperCountSpecification() ^ EndosOper() ^ OsloCountAllSpecification()

    temp = []
    for d in kwargs['data']:
        for p in bf.filter(d, sp):
            temp.append(bf.format_list(p))
            
    res_set = list(set([t[0] for t in temp]))
    res = []
    for r in res_set:
        res.append([
            r,
            []
        ])
    for r in res:
        for t in temp:
            if r[0] == t[0]:
                r[1].append(t[1])
            r[1] = list(set(r[1]))

    res_set.clear()
    for r in res:
        tt = []
        tt.append(r[0])
        for rr in r[1]:
            tt.append([rr,0,0,0,0])
        res_set.append(tt)
    # print(res_set)
    # for t in temp:
    #     x = (t[0],(t[1],None,None,None,None))
    #     res_set.add(x)
    # res_set = [list(r[rr]) for r in res_set for rr in range(len(r)) if r!=0 ]

    # print(res_set)
    # for t in temp:
    #     for r in res_set:
    #         if r[0] == t[0] and r[1] == t[1]:
    #             pass
    #             r[2] = r[2] + t[2] if r[2] != None else t[2]
    #             r[3] = r[3] + t[3] if r[3] != None else t[3]
    #             r[4] = r[4] + t[4] if r[4] != None else t[4]
    row = 7

    for i, res in enumerate(res_set):
        row +=1
        for l, r in enumerate(res):
            if l == 0:
                sheet.cell(row=row, column=1 + l).value = r if r != 0 else ''
            else:
                for n,prof in enumerate(r):
                    if n == 0:
                        sheet.cell(row=row, column=2).value = r[0] if r[0] != 0 else ''
                    else:
                        row += 1
                        sheet.cell(row=row, column=2).value = r[0] if r[0] != 0 else ''

def get_rez_a_oth_23_list(data):
    v1 = 0
    v2 = 0
    v3 = 0
    v4 = 0
    v5 = 0
    v6 = 0
    for d in data:
        le_trv = d.sluchay.le_trv
        if le_trv.details and le_trv.details.kod[:3] in ['V10','V19']:
            v1+=1
        if le_trv.details and le_trv.details.kod[:3] in ['V20','V29']:
            v2+=1
        if le_trv.details and le_trv.details.kod[:3] in ['V01','V09']:
            v3+=1
        if le_trv.details and le_trv.details.kod[:3] in ['V30', 'V39']:
            v4 += 1
        if le_trv.details and le_trv.details.kod[:3] in ['V40', 'V69']:
            v5 += 1
        if le_trv.details and le_trv.details.kod[:3] in ['V70', 'V79']:
            v6 += 1
    return [v1,v2,v3,v4,v5,v6]

def insert_sheet_a_oth_23(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']

    data_ymer = 0
    data_17_ym = []
    data_65_ym = []
    data_60_ym = []
    data_old_ym = []
    all_count = 0
    data_17 = []
    data_65 = []
    data_60 = []
    data_old = []
    data_sl = []
    for d in kwargs['data']:
        if d.sluchay.le_trv:
            le_trv = d.sluchay.le_trv
            if le_trv.t_trv and le_trv.t_trv.kod == '7':
                data_sl.append(d)
                year = d.patient_year
                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105,106]:
                    data_ymer+=1
                    if 0 < year <= 17:
                        data_17_ym.append(d)
                    if d.patient.pol and d.patient.pol.id_pol == 1:
                        if 16 < year <= 65:
                            data_65_ym.append(d)
                        elif year > 65:
                            data_old_ym.append(d)
                    if d.patient.pol and d.patient.pol.id_pol == 2:
                        if 16 < year <= 60:
                            data_60_ym.append(d)
                        elif year > 60:
                            data_old_ym.append(d)
                else:
                    all_count += 1
                    if 0 < year <= 17:
                        data_17.append(d)
                    if d.patient.pol and d.patient.pol.id_pol == 1:
                        if 16 < year <= 65:
                            data_65.append(d)
                        elif year > 65:
                            data_old.append(d)
                    if d.patient.pol and d.patient.pol.id_pol == 2:
                        if 16 < year <= 60:
                            data_60.append(d)
                        elif year > 60:
                            data_old.append(d)


    sheet.cell(row=8, column=3).value = data_ymer if data_ymer != 0 else None
    sheet.cell(row=9, column=3).value = len(data_17_ym) if len(data_17_ym) != 0 else None
    rez = get_rez_a_oth_23_list(data_17_ym)
    for n,r in enumerate(rez):
        sheet.cell(row=10+n, column=3).value = r if r != 0 else None
    sheet.cell(row=16, column=3).value = len(data_65_ym) + len(data_60_ym) if (len(data_65_ym) + len(data_60_ym)) != 0 else None
    rez65 = get_rez_a_oth_23_list(data_65_ym)
    rez60 = get_rez_a_oth_23_list(data_60_ym)
    rez_65_60 = []
    rez_65_60.append(rez65)
    rez_65_60.append(rez60)
    r = None
    for o in range(len(rez_65_60)):
        if o == 0:
            r = numpy.array(rez_65_60[o])
        else:
            r += numpy.array(rez_65_60[o])
    rez = r.tolist()
    for n, r in enumerate(rez):
        sheet.cell(row=17 + n, column=3).value = r if r != 0 else None
    rez_old = get_rez_a_oth_23_list(data_old_ym)
    sheet.cell(row=23, column=3).value = len(data_old_ym) if len(data_old_ym) != 0 else None 
    for n,r in enumerate(rez_old):
        sheet.cell(row=24+n, column=3).value = r if r != 0 else None

    sheet.cell(row=30, column=3).value = all_count if all_count != 0 else None
    sheet.cell(row=31, column=3).value = len(data_17) if len(data_17) != 0 else None
    rez = get_rez_a_oth_23_list(data_17)
    for n, r in enumerate(rez):
        sheet.cell(row=32 + n, column=3).value = r if r != 0 else None
    sheet.cell(row=38, column=3).value = len(data_65) + len(data_60) if (len(data_65) + len(data_60)) != 0 else None
    rez65 = get_rez_a_oth_23_list(data_65)
    rez60 = get_rez_a_oth_23_list(data_60)
    rez_65_60 = []
    rez_65_60.append(rez65)
    rez_65_60.append(rez60)
    r = None
    for o in range(len(rez_65_60)):
        if o == 0:
            r = numpy.array(rez_65_60[o])
        else:
            r += numpy.array(rez_65_60[o])
    rez = r.tolist()
    for n, r in enumerate(rez):
        sheet.cell(row=39 + n, column=3).value = r if r != 0 else None
    rez_old = get_rez_a_oth_23_list(data_old)
    sheet.cell(row=45, column=3).value = len(data_old) if len(data_old) != 0 else None
    for n,r in enumerate(rez_old):
        sheet.cell(row=46+n, column=3).value = r if r != 0 else None
    row = 60
    # print(len(data_sl))
    for sl in data_sl:
        sheet.cell(row=row, column=1).value = sl.sluchay.nib
        sheet.cell(row=row, column=2).value = f'{sl.patient.fam} {sl.patient.im} {sl.patient.ot}'
        sheet.cell(row=row, column=3).value = sl.sluchay.datv.strftime('%d.%m.%Y') if sl.sluchay.datv else None
        sheet.cell(row=row, column=4).value = sl.sluchay.otd.naim if sl.sluchay.otd else None
        sheet.cell(row=row, column=5).value = sl.sluchay.dskz.kod if sl.sluchay.dskz != None else None
        for i in range(1,6):
            sheet.cell(row=row, column=i).alignment = styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row, column=i).border = border
        row+=1
def get_rez_32_list(data,ds):
    temp = []
    ds_list = []
    ds_list_T51 = ['T51','T51.0','T51.1','T51.2','T51.3','T51.4','T51.5','T51.6','T51.7','T51.8','T51.9']
    ds_list_T40 = ['T40','T40.0','T40.1','T40.2','T40.3','T40.4','T40.5','T40.6','T40.7','T40.8','T40.9']
    ds_list_T40_6 = ['T40.6']
    ds_list_T43 = ['T43','T43.0','T43.1','T43.2','T43.3','T43.4','T43.5','T43.6','T43.7','T43.8','T43.9']
    ds_list_T39 = ['T39','T39.0','T39.1','T39.2','T39.3','T39.4','T39.5','T39.6','T39.7','T39.8','T39.9']
    not_ds = ds_list_T51+ds_list_T40+ds_list_T40_6+ds_list_T43+ds_list_T39

    if ds == 'T51':
        ds_list = ds_list_T51
    elif ds == 'T40':
        ds_list = ds_list_T40
    elif ds == 'T40_6':
        ds_list = ds_list_T40_6
    elif ds == 'T43':
        ds_list = ds_list_T43
    elif ds == 'T39':
        ds_list = ds_list_T39
    

    for d in data:
        if ds != '':
            if d.sluchay.dskz and d.sluchay.dskz.kod in ds_list:
                temp.append(rez_32_list_ord(d))
        else:
            if d.sluchay.dskz and d.sluchay.dskz.kod not in not_ds:
             temp.append(rez_32_list_ord(d))


    return temp
def set_rez_32_list(sheet,row,n,d):
    sheet.row_dimensions[row].height = 35
    sheet.cell(row=row, column=1).value = n
    sheet.cell(row=row, column=1).border = border

    sheet.cell(row=row, column=2).value = d['fio']
    sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
    sheet.cell(row=row, column=2).border = border

    sheet.cell(row=row, column=3).value = d['m']
    sheet.cell(row=row, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=3).border = border

    sheet.cell(row=row, column=4).value = d['datr']
    sheet.cell(row=row, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=4).border = border

    sheet.cell(row=row, column=5).value = d['datp']
    sheet.cell(row=row, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=5).border = border

    sheet.cell(row=row, column=6).value = d['datv']
    sheet.cell(row=row, column=6).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=6).border = border

    sheet.cell(row=row, column=7).value = d['dskz']
    sheet.cell(row=row, column=7).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=7).border = border

    sheet.cell(row=row, column=8).value = d['ds_osl']
    sheet.cell(row=row, column=8).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=8).border = border

    sheet.cell(row=row, column=9).value = d['isx']
    sheet.cell(row=row, column=9).alignment = styles.Alignment(horizontal="center", vertical="center")
    sheet.cell(row=row, column=9).border = border

    sheet.cell(row=row, column=10).value = d['adr_1']
    sheet.cell(row=row, column=10).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
    sheet.cell(row=row, column=10).border = border
    
    sheet.cell(row=row, column=11).value = d['adr_2']
    sheet.cell(row=row, column=11).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
    sheet.cell(row=row, column=11).border = border

def rez_32_list_ord(d):
    ord = OrderedDict()
    f = d.patient.fam
    i = d.patient.im if len(d.patient.im) > 0 else ''
    ot = d.patient.ot if len(d.patient.ot) > 0 else ''
    ord['fio'] = f'{f} {i}.{ot}'
    ord['m'] = d.patient.pol.polname[0] if d.patient.pol else None
    ord['datr'] = d.patient.datr.strftime('%d.%m.%Y') if d.patient.datr else None
    ord['datp'] = d.sluchay.datp.strftime('%d.%m.%Y') if d.sluchay.datp else None
    ord['datv'] = d.sluchay.datv.strftime('%d.%m.%Y') if d.sluchay.datv else None
    ord['dskz'] = d.sluchay.dskz.kod if d.sluchay.dskz else None
    ord['ds_osl'] = d.sluchay.ds_osl.kod if d.sluchay.ds_osl else None
    if d.sluchay.icx:
        if d.sluchay.icx.id_iz == 101:
            ord['isx'] = 'выздор'
        elif d.sluchay.icx.id_iz == 102:
            ord['isx'] = 'улучшен'
        elif d.sluchay.icx.id_iz == 103:
            ord['isx'] = 'без.пер'
        elif d.sluchay.icx.id_iz == 104:
            ord['isx'] = 'ухудшен'
        elif d.sluchay.icx.id_iz == 105:
            ord['isx'] = 'умер'
        elif d.sluchay.icx.id_iz == 106:
            ord['isx'] = 'ум.1-е.сут.'
    else:
        ord['isx'] = ''
    # if 'Тюмень' in d.patient.m_roj:
    if 'р-н' in d.patient.m_roj:
        ord['adr_1'] = 'Тюм.р-н'
    else:
        ord['adr_1'] = 'г.Тюмень'
    # else:
    #     ord['adr_1'] = ''
    ord['adr_2'] = d.patient.m_roj

    return ord

def get_rez_32_info(data,ds):
    ds_list_T51 = ['T51','T51.0','T51.1','T51.2','T51.3','T51.4','T51.5','T51.6','T51.7','T51.8','T51.9']
    ds_list_T40 = ['T40','T40.0','T40.1','T40.2','T40.3','T40.4','T40.5','T40.6','T40.7','T40.8','T40.9']
    ds_list_T40_6 = ['T40.6']
    ds_list_T40_7 = ['T40.7']
    ds_list_T43 = ['T43','T43.0','T43.1','T43.2','T43.3','T43.4','T43.5','T43.6','T43.7','T43.8','T43.9']
    ds_list_T39 = ['T39','T39.0','T39.1','T39.2','T39.3','T39.4','T39.5','T39.6','T39.7','T39.8','T39.9']
    not_ds = ds_list_T51+ds_list_T40+ds_list_T40_6+ds_list_T43+ds_list_T39+ds_list_T40_7

    ds_list = []
    if ds == 'T51':
        ds_list = ds_list_T51
    elif ds == 'T40':
        ds_list = ds_list_T40
    elif ds == 'T40.6':
        ds_list = ds_list_T40_6
    elif ds == 'T40.7':
        ds_list = ds_list_T40_7
    elif ds == 'T43':
        ds_list = ds_list_T43
    elif ds == 'T39':
        ds_list = ds_list_T39
    
    temp1 = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    temp2 = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    temp3 = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]

    for d in data:
        if ds != '':
            if d.sluchay.dskz and d.sluchay.dskz.kod in ds_list:
                set_rez_32_info(d,temp1,temp2,temp3)
        else:
            if d.sluchay.dskz and d.sluchay.dskz.kod not in not_ds:
                set_rez_32_info(d,temp1,temp2,temp3)
    
    return [temp1,temp2,temp3]

def set_rez_32_info(d,temp1,temp2,temp3):
    if 'р-н' not in d.patient.m_roj:
        temp1[0]+=1
        if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
            temp2[0]+=1
        if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
            temp3[0]+=1
    if 'р-н' in d.patient.m_roj: 
        temp1[1]+=1
        if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
            temp2[1]+=1
        if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
            temp3[1]+=1

    if d.patient.pol and d.patient.pol.id_pol == 2:
        if 'р-н' not in d.patient.m_roj:
            temp1[2]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[2]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[2]+=1
    if d.patient.pol and d.patient.pol.id_pol == 2:
        if 'р-н' in d.patient.m_roj: 
            temp1[3]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[3]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[3]+=1
    
    if d.patient_year < 15:
        if 'р-н' not in d.patient.m_roj:
            temp1[4]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[4]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[4]+=1
    if d.patient_year < 15:
        if 'р-н' in d.patient.m_roj: 
            temp1[5]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[5]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[5]+=1
    
    if d.patient_year <= 17:
        if 'р-н' not in d.patient.m_roj:
            temp1[6]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[6]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[6]+=1
    if d.patient_year <= 17:
        if 'р-н' in d.patient.m_roj: 
            temp1[7]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[7]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[7]+=1
    
    if 18 <= d.patient_year <= 39:
        if 'р-н' not in d.patient.m_roj:
            temp1[8]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[8]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[8]+=1
    if 18 <= d.patient_year <= 39:
        if 'р-н' in d.patient.m_roj: 
            temp1[9]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[9]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[9]+=1
    
    if 18 <= d.patient_year<= 19:
        if 'р-н' not in d.patient.m_roj:
            temp1[10]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[10]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[10]+=1
    if 18 <= d.patient_year <= 19:
        if 'р-н' in d.patient.m_roj: 
            temp1[11]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[11]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[11]+=1
    
    if 20 <= d.patient_year <= 39:
        if 'р-н' not in d.patient.m_roj:
            temp1[12]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[12]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[12]+=1
    if 20 <= d.patient_year <= 39:
        if 'р-н' in d.patient.m_roj: 
            temp1[13]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[13]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[13]+=1
    
    if d.patient_year >= 40:
        if 'р-н' not in d.patient.m_roj:
            temp1[14]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[14]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[14]+=1
    if d.patient_year >= 40:
        if 'р-н' in d.patient.m_roj: 
            temp1[15]+=1
            if d.sluchay.ds_osl and d.sluchay.ds_osl.kod == 'R40.2':
                temp2[15]+=1
            if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                temp3[15]+=1

def insert_sheet_a_oth_32(**kwargs):
    sheet = kwargs['sheet'][0]
    sheet1 = kwargs['sheet'][1]
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    ###Инфо
    rez_T51 = get_rez_32_info(kwargs['data'],'T51')
    rez_T40 = get_rez_32_info(kwargs['data'],'T40')
    rez_T40_7 = get_rez_32_info(kwargs['data'],'T40.7')
    rez_T40_6 = get_rez_32_info(kwargs['data'],'T40.6')
    rez_T43 = get_rez_32_info(kwargs['data'],'T43')
    rez_T39 = get_rez_32_info(kwargs['data'],'T39')
    rez_none = get_rez_32_info(kwargs['data'],'')



    row = 6
    for data in rez_T51:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")
   

    for data in rez_T40:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")
    
    for data in rez_T40_7:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")


    for data in rez_T40_6:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")

    
    for data in rez_T43:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")
    
    for data in rez_T39:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")
    
    for data in rez_none:
        row+=1
        for d in range(len(data)):
            sheet.cell(row=row, column=2+d).value = data[d] if data[d] != 0 else None
            sheet.cell(row=row, column=2+d).alignment = styles.Alignment(horizontal="center", vertical="center")
    
    ### Список
    rez_T51 = get_rez_32_list(kwargs['data'],'T51')
    rez_T40 = get_rez_32_list(kwargs['data'],'T40')
    rez_T40_6 = get_rez_32_list(kwargs['data'],'T40_6')
    rez_T43 = get_rez_32_list(kwargs['data'],'T43')
    rez_T39 = get_rez_32_list(kwargs['data'],'T39')
    rez_none = get_rez_32_list(kwargs['data'],'')

    row = 3
    row+=1
    sheet1.cell(row=row, column=1).value = 'Отравления алкоголем и спиртами (Т51-Т51.9)'
    sheet1.merge_cells(f"A{sheet.cell(row=row, column=1).row}:K{sheet.cell(row=row, column=11).row}")
    sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
    for n,r in enumerate(sorted(rez_T51,key=itemgetter('fio')),1):
        row+=1
        set_rez_32_list(sheet1,row,n,r)

    row+=1
    sheet1.cell(row=row, column=1).value = 'Отравления наркотическими средствами (Т40-Т40.9)'
    sheet1.merge_cells(f"A{sheet.cell(row=row, column=1).row}:K{sheet.cell(row=row, column=11).row}")
    sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
    for n,r in enumerate(sorted(rez_T40,key=itemgetter('fio')),1):
        row+=1
        set_rez_32_list(sheet1,row,n,r)
    
    row+=1
    sheet1.cell(row=row, column=1).value = 'Отравления курительными смесями (Т40.6)'
    sheet1.merge_cells(f"A{sheet.cell(row=row, column=1).row}:K{sheet.cell(row=row, column=11).row}")
    sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
    for n,r in enumerate(sorted(rez_T40_6,key=itemgetter('fio')),1):
        row+=1
        set_rez_32_list(sheet1,row,n,r)
        
    
    row+=1
    sheet1.cell(row=row, column=1).value = 'Отравления психотропными средствами смесями (Т43-Т43.9)'
    sheet1.merge_cells(f"A{sheet.cell(row=row, column=1).row}:K{sheet.cell(row=row, column=11).row}")
    sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
    for n,r in enumerate(sorted(rez_T43,key=itemgetter('fio')),1):
        row+=1
        set_rez_32_list(sheet1,row,n,r)
    
    row+=1
    sheet1.cell(row=row, column=1).value = 'Отравления (Т39-Т39.9)'
    sheet1.merge_cells(f"A{sheet.cell(row=row, column=1).row}:K{sheet.cell(row=row, column=11).row}")
    sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
    for n,r in enumerate(sorted(rez_T39,key=itemgetter('fio')),1):
        row+=1
        set_rez_32_list(sheet1,row,n,r)
    
    row+=1
    sheet1.cell(row=row, column=1).value = 'Другие отравления и токсич.воздействия веществ'
    sheet1.merge_cells(f"A{sheet.cell(row=row, column=1).row}:K{sheet.cell(row=row, column=11).row}")
    sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
    for n,r in enumerate(sorted(rez_none,key=itemgetter('fio')),1):
        row+=1
        set_rez_32_list(sheet1,row,n,r)

def insert_sheet_a_oth_36(**kwargs):
    ds60_64 = list(Ds.objects.values('kod').filter(kod__range=('I60', 'I64.9')))
    ds60_64 = [k['kod'] for k in ds60_64]
    ds63 = list(Ds.objects.values('kod').filter(kod__range=('I63', 'I63.9')))
    ds63 = [k['kod'] for k in ds63]
    ds60_69 = list(Ds.objects.values('kod').filter(kod__range=('I60', 'I69.9')))
    ds60_69 = [k['kod'] for k in ds60_69]
    ds60_66 = list(Ds.objects.values('kod').filter(kod__range=('I60', 'I66.9')))
    ds60_66 = [k['kod'] for k in ds60_66]
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = kwargs['data']
    data = get_list_ds(data)
    tab1 = [[0,0,0],[0,0,0],[0,0,0]]
    # tab2 = [[0,0,0],[0,0,0],[0,0,0],[0,0,0],
    #         [0,0,0],[0,0,0],[0,0,0],[0,0,0],
    #         [0,0,0],[0,0,0]]
    tab2 = [[0, 0], [0, 0], [0, 0], [0, 0],
            [0, 0], [0, 0], [0, 0], [0, 0],
            [0, 0], [0, 0]]

    for d in data:
        ds = d[0][0]
        if ds.kod in ds60_64:
            tab1[0][0] += len(d[1])
            for sl in d[1]:
                year = sl.patient_year
                if sl.patient.cj:
                    if sl.patient.cj.kod == 2:
                        tab1[1][0] += 1
                if sl.patient.pol and sl.patient.pol.id_pol == 1:
                    if year < settings.OLD_M:
                        tab1[2][0] += 1
                if sl.patient.pol and sl.patient.pol.id_pol == 2:
                    if year < settings.OLD_G:
                        tab1[2][0] += 1
        if ds.kod in ds63:
            tab1[0][1] += len(d[1])
            for sl in d[1]:
                year = sl.patient_year
                if sl.patient.cj:
                    if sl.patient.cj.kod == 2:
                        tab1[1][1] += 1
                if sl.patient.pol and sl.patient.pol.id_pol == 1:
                    if year < settings.OLD_M:
                        tab1[2][1] += 1
                if sl.patient.pol and sl.patient.pol.id_pol == 2:
                    if year < settings.OLD_G:
                        tab1[2][1] += 1
        if ds.kod in ds60_69:
            tab1[0][2] += len(d[1])
            for sl in d[1]:
                year = sl.patient_year
                if sl.patient.cj:
                    if sl.patient.cj.kod == 2:
                        tab1[1][2] += 1
                if sl.patient.pol and sl.patient.pol.id_pol == 1:
                    if year < settings.OLD_M:
                        tab1[2][2] += 1
                if sl.patient.pol and sl.patient.pol.id_pol == 2:
                    if year < settings.OLD_G:
                        tab1[2][2] += 1
        if ds.kod in ds60_64:
            tab2[0][0] += len(d[1])
            for sl in d[1]:
                if sl.sluchay.vrez:
                    if sl.sluchay.vrez.kod in [1,2,3,4,5,6,7,8,9,10]:
                        tab2[1][0] += 1
                    if sl.sluchay.vrez.kod == 1:
                        tab2[2][0] += 1
                    if sl.sluchay.vrez.kod == 2:
                        tab2[3][0] += 1
                    if sl.sluchay.vrez.kod == 3:
                        tab2[4][0] += 1
                    if sl.sluchay.vrez.kod == 4:
                        tab2[5][0] += 1
                    if sl.sluchay.vrez.kod == 5:
                        tab2[6][0] += 1
                    if sl.sluchay.vrez.kod in [1,2,3,4,41]:
                        tab2[7][0] += 1
                    if sl.sluchay.vrez.kod == 6:
                        tab2[8][0] += 1
                    if sl.sluchay.vrez.kod == 7:
                        tab2[9][0] += 1
        if ds.kod in ds63:
            tab2[0][1] += len(d[1])
            for sl in d[1]:
                if sl.sluchay.vrez:
                    if sl.sluchay.vrez.kod in [1,2,3,4,5,6,7,8,9,10]:
                        tab2[1][1] += 1
                    if sl.sluchay.vrez.kod == 1:
                        tab2[2][1] += 1
                    if sl.sluchay.vrez.kod == 2:
                        tab2[3][1] += 1
                    if sl.sluchay.vrez.kod == 3:
                        tab2[4][1] += 1
                    if sl.sluchay.vrez.kod == 4:
                        tab2[5][1] += 1
                    if sl.sluchay.vrez.kod == 5:
                        tab2[6][1] += 1
                    if sl.sluchay.vrez.kod in [1,2,3,4,41]:
                        tab2[7][1] += 1
                    if sl.sluchay.vrez.kod == 6:
                        tab2[8][1] += 1
                    if sl.sluchay.vrez.kod == 7:
                        tab2[9][1] += 1

        # if ds.kod in ds60_66:
        #     tab2[0][2] += len(d[1])
        #     for sl in d[1]:
        #         if sl.sluchay.vrez:
        #             if sl.sluchay.vrez.kod in [1,2,3,4,5,6,7,8,9,10]:
        #                 tab2[1][2] += 1
        #             if sl.sluchay.vrez.kod == 1:
        #                 tab2[2][2] += 1
        #             if sl.sluchay.vrez.kod == 2:
        #                 tab2[3][2] += 1
        #             if sl.sluchay.vrez.kod == 3:
        #                 tab2[4][2] += 1
        #             if sl.sluchay.vrez.kod == 4:
        #                 tab2[5][2] += 1
        #             if sl.sluchay.vrez.kod == 5:
        #                 tab2[6][2] += 1
        #             if sl.sluchay.vrez.kod in [1,2,3,4,41]:
        #                 tab2[7][2] += 1
        #             if sl.sluchay.vrez.kod == 6:
        #                 tab2[8][2] += 1
        #             if sl.sluchay.vrez.kod == 7:
        #                 tab2[9][2] += 1
    sheet.cell(row=4, column=1).value = str(name).capitalize()
    sheet.cell(row=5, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    row = 8
    for result in tab1:
        row += 1
        for n,res in enumerate(result):
            sheet.cell(row=row, column=2+n).value = res if res != 0 else None
    row = 16
    for result in tab2:
        row += 1
        for n,res in enumerate(result):
            sheet.cell(row=row, column=2+n).value = res if res != 0 else None



def rez_oth_a_oth_29(data,all=None):
    rez = [0,0,0,0]
    for d in data:
        if d.sluchay.dskz:
            year = d.patient_year
            pol = d.patient.pol.id_pol if d.patient.pol and d.patient.pol.id_pol else None
            if year >= 18:
                rez[0]+=1
                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                    rez[1]+=1
                if pol == 1:
                    if year >= 65:
                        rez[2] += 1
                        if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                            rez[3] += 1
                if pol == 2:
                    if year >= 60:
                        rez[2] += 1
                        if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                            rez[3] += 1
    return rez

def insert_sheet_a_oth_29(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    filters = kwargs['filters']
    sheet.cell(row=2, column=1).value = f'Справка о пациентах, за период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")}'
    sheet.cell(row=3, column=1).value = f'ГБУЗ ТО "Областная клиническая больница № 2" {str(name).capitalize()}'
    if filters:
        filter = []
        for f in filters:
            filter.append(f"{f['filter']}:{f['value']}")
        sheet.cell(row=4, column=1).value = '\n'.join(filter)
    list_lpu = get_list_lpu(kwargs['data'],all=True,typ_lpy=True)
    row = 7
    rez_all = []
    tyumen = []
    total = []
    for l in list_lpu:
        if l[0][0] is None:
            for d in l[1]:
                if (re.search('Тюменс',d.patient.m_roj)) or (re.search('ТЮМЕНС',d.patient.m_roj)):
                    tyumen.append(d)
                else:
                    total.append(d)
    for l in list_lpu:
        row += 1
        if l[0][0] != None:
            if len(l[0][0]) >= 25:
                sheet.row_dimensions[row].height = 25
            sheet.cell(row=row, column=1).value = l[0][0]
            sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
            rez = rez_oth_a_oth_29(l[1])
            rez_all.append(rez)
            for r in range(len(rez)):
                sheet.cell(row=row, column=2+r).value = rez[r] if rez[r] != 0 else None
                sheet.cell(row=row, column=2+r).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
            else:
                for c in range(1, 6):
                    sheet.cell(row=row, column=c).border = styles.Border(
                        bottom=styles.Side(border_style='thin', color='000000'))
        else:
            if len(tyumen) > 0:
                sheet.cell(row=row, column=1).value = 'ТЮМЕНСКИЙ Р-Н'
                rez = rez_oth_a_oth_29(tyumen)
                rez_all.append(rez)
                for r in range(len(rez)):
                    sheet.cell(row=row, column=2+r).value = rez[r] if rez[r] != 0 else None
                    sheet.cell(row=row, column=2+r).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
                else:
                    for c in range(1, 6):
                        sheet.cell(row=row, column=c).border = styles.Border(
                            bottom=styles.Side(border_style='thin', color='000000'))
            if len(total) > 0:
                row+=1
                sheet.cell(row=row, column=1).value = 'ПРОЧИЕ'
                rez = rez_oth_a_oth_29(total)
                rez_all.append(rez)
                for r in range(len(rez)):
                    sheet.cell(row=row, column=2 + r).value = rez[r] if rez[r] != 0 else None
                    sheet.cell(row=row, column=2 + r).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
                else:
                    for c in range(1, 6):
                        sheet.cell(row=row, column=c).border = styles.Border(
                            bottom=styles.Side(border_style='thin', color='000000'))

    else:
        if len(rez_all) > 0:
            row+=1
            r = None
            for o in range(len(rez_all)):
                if o == 0:
                    r = numpy.array(rez_all[o])
                else:
                    r += numpy.array(rez_all[o])
            rez_all = r.tolist()
            sheet.cell(row=row, column=1).value = 'ИТОГО'
            for r in range(len(rez_all)):
                sheet.cell(row=row, column=2 + r).value = rez_all[r] if rez_all[r] != 0 else None
                sheet.cell(row=row, column=2 + r).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
            else:
                for c in range(1, 6):
                    sheet.cell(row=row, column=c).border = styles.Border(
                        bottom=styles.Side(border_style='thin', color='000000'))
class GroupP1(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user,request)
        self.group_st_ot = request.get('group_st_ot')
        self.otdels = json.loads(request.get('otdels'))
        self.user_group_name = 'hospital_reports_%s' % user
        self.fil_child = request.get('fil_child')
    def create(self):
        file = self.is_file('group_p1.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            if self.group_st_ot == 'st':
                otds = list(set([p.sluchay.otd.naim for p in patients.patients if p.sluchay.otd]))
            elif self.group_st_ot == 'ot':
                otds = list(set([p.sluchay.otd.naim for p in patients.patients if p.sluchay.otd and p.sluchay.otd.naim in self.otdels]))
            data = list()
            count = 0
            ym_count = 0

            for o in sorted(otds):
                temp = []
                for p in patients.patients:
                    if p.sluchay.otd and p.sluchay.otd.naim == o:
                        ord = OrderedDict()
                        count += 1
                        ord['nib'] = p.sluchay.nib[4:]
                        f = p.patient.fam
                        i = p.patient.im[0] if len(p.patient.im) > 0 else ''
                        ot = p.patient.ot[0] if len(p.patient.ot) > 0 else ''
                        ord['fio'] = f'{f} {i}.{ot}'
                        ord['datr'] = p.patient.datr.strftime('%d.%m.%Y') if p.patient.datr else None
                        ord['datp'] = p.sluchay.datp.strftime('%d.%m.%Y') if p.sluchay.datp else None
                        ord['datv'] = p.sluchay.datv.strftime('%d.%m.%Y') if p.sluchay.datv else None
                        ord['isx'] = 'умр.' if p.sluchay.icx and p.sluchay.icx.id_iz in [105, 106] else None
                        if ord['isx']:
                            ym_count += 1
                        ord['dskz'] = f'{p.sluchay.dskz.kod}' if p.sluchay.dskz else None
                        ord['prof_k'] = p.sluchay.le_vr.prof_k.k_prname if p.sluchay.le_vr and p.sluchay.le_vr.prof_k else None
                        ord['vds'] = p.sluchay.vds.vds.naim[:5] if p.sluchay.vds and p.sluchay.vds.vds else None
                        temp.append(ord)
                data.append([o, temp])

            for d in data:
                for n,i in enumerate(d):
                    if n != 0:
                        if self.fil_child == 'По фамилии':
                            d[n] = sorted(d[n],key=itemgetter('fio'))
                        else:
                            d[n] = sorted(d[n], key=itemgetter('datv'),reverse=False)



            dic = dict([('sheet',sheet),('name',self.user.statistics_type.name),
                        ('date_1',self.date_1),('date_2',self.date_2),('data',data),
                        ('count',count),('ym_count',ym_count),('otdels',self.otdels)])
            sheet = insert_sheet_P1(**dic)
            wb.save(self.path() + f'group_p1_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download', 'text': self.path() + f'group_p1_{self.user.user.id}.xlsx',
                                                           'name': 'Отчет о выбывших'})

class Implants(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user,request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('implants.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays(update_user=False)
            implants = []
            s_code = V036.objects.values('s_code').filter(dateend=None, parameter__in=[1, 3]).all()
            code_oper = [code['s_code'] for code in s_code]
            for t in patients.patients:
                oper = get_pop_oper(t)
                if oper is not None:
                    if oper.kod_op is not None:
                        if oper.kod_op.kod in code_oper:
                            implants.append(t)

            dic = dict([('sheet', sheet), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2),('data',implants)])
            sheet = insert_sheet_implants(**dic)
            wb.save(self.path() + f'implants_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'implants_{self.user.user.id}.xlsx',
                                                           'name': 'Отчет о имплантах'})
class GroupP2(AnnualReportABC):
    def __init__(self, user, request):
        super().__init__(user,request)
        self.request = request
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('group_p2_2.xlsx')
        if file:
            wb = load_workbook(file)
            os.remove(file)
            sheet = wb.active
            patients = PatientsDataFiltrs(self.date_1,self.date_2,self.user,self.request)
            # print(len(patients.patients))
            dic = dict([('sheet',sheet),('data',patients.patients),('name',self.user.statistics_type.name),
                        ('date_1',self.date_1),('date_2',self.date_2),('filters',self.filters_list())])
            insert_sheet_P2(**dic)
            wb.save(self.path() + f'group_p2_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'group_p2_{self.user.user.id}.xlsx',
                                                           'name': 'Отчет о выбывших расшир.'})
        else:
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'text': 'Отсутствует шаблон - group_p1.xlsx'})
            raise FileNotFoundError('Отсутствует шаблон - group_p1.xlsx')
class GroupP3(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP4(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP5(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP6(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP7(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP8(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP9(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP10(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP11(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP12(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP13(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP14(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP15(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP16(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP17(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP18(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP19(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('group_p19.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            data = []
            for p in patients.patients:
                if (p.sluchay.icx and p.sluchay.icx.id_iz in(105,106))\
                        or (p.sluchay.rslt and p.sluchay.rslt.id_tip in (105,106)):
                    if (p.patient_year) <= 65:
                        ## в первые сутки ???
                        if p.sluchay.datp == p.sluchay.datv:
                            data.append(p)
            dic = dict([('sheet', sheet), ('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_P19(**dic)
            wb.save(self.path() + f'group_p19_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                         {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'group_p19_{self.user.user.id}.xlsx',
                                                           'name': ''})

        else:
            pass
class GroupP20(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP21(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP22(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        filters = json.loads(request.get('filters'))['filter']
        self.otd = filters['otd']['otd'] if filters['otd']['otd'] else None
        ### Уточнить фильтруем по этим полям ?
        self.metod_hmp = filters['metod_hmp']['metod_hmp'] if filters['metod_hmp']['metod_hmp'] else None
        self.vid_hmp = filters['vid_hmp']['vid_hmp'] if filters['vid_hmp']['vid_hmp'] else None
        self.user_group_name = 'hospital_reports_%s' % user
        self.otdels = json.loads(request.get('otdels'))
    def create(self):
        file = self.is_file('group_p22.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsDataFiltrs(self.date_1,self.date_2,self.user,self.request)
            data = []
            if self.otd != None:
                for p in patients.patients:
                    if p.sluchay.otd and p.sluchay.otd.naim == self.otd and p.sluchay.tip_oms == '2':
                        data.append(p)
            else:
                for p in patients.patients:
                    if p.sluchay.tip_oms == '2':
                        data.append(p)
            dic = dict([('sheet', sheet), ('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2),('otd',self.otd),('otdels',self.otdels)])
            sheet = insert_sheet_P22(**dic)
            wb.save(self.path() + f'group_p22_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'group_p22_{self.user.user.id}.xlsx'})
class GroupP23(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class GroupP24(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('group_p24.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsDataFiltrs(self.date_1, self.date_2, self.user, self.request)
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2), ('user', self.user)])
            insert_sheet_P24(**dic)
            wb.save(self.path() + f'group_p24_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'group_p24_{self.user.user.id}.xlsx',
                                                           'name': ''})
class GroupP25(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        pass


class VaultOtd(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.request = request
        self.user_group_name = 'hospital_reports_%s' % user

    def create(self):
        typ = self.request.get('type',None)
        patients = PatientsData(self.date_1, self.date_2, self.user)
        patients.sluchays()
        data = []
        otdel = self.request.get('otdel',None)
        otd = ''
        filter_otd = []
        oth_typ = None
        if typ != 'null':
            for p in patients.patients:
                if typ == 'ttt':
                    filter_otd = ['ТРАВМА N1','ТРАВМА N2','ТРАВМА N3']
                    otd = ' '.join(['ТРАВМА N1','ТРАВМА N2','ТРАВМА N3'])
                    if p.sluchay.otd and p.sluchay.otd.naim in ['ТРАВМА N1','ТРАВМА N2','ТРАВМА N3']:
                        data.append(p)
                elif typ == 'nnn':
                    filter_otd = ['НЕВРОЛОГИЯ N1','НЕВРОЛОГИЯ N2','НЕВРОЛОГИЯ N3']
                    otd = ' '.join(['НЕВРОЛОГИЯ N1','НЕВРОЛОГИЯ N2','НЕВРОЛОГИЯ N3'])
                    if p.sluchay.otd and p.sluchay.otd.naim in ['НЕВРОЛОГИЯ N1','НЕВРОЛОГИЯ N2','НЕВРОЛОГИЯ N3']:
                        data.append(p)
                elif typ == 'hh':
                    filter_otd = ['ХИРУРГИЧЕСКОЕ N1','ХИРУРГИЧЕСКОЕ N2(гн)']
                    otd = ' '.join(['ХИРУРГИЧЕСКОЕ N1','ХИРУРГИЧЕСКОЕ N2(гн)'])
                    if p.sluchay.otd and p.sluchay.otd.naim in ['ХИРУРГИЧЕСКОЕ N1','ХИРУРГИЧЕСКОЕ N2(гн)']:
                        data.append(p)
            oth_typ=False
        else:
            otd = otdel
            for p in patients.patients:
                if otdel != '':
                    if p.sluchay.otd and p.sluchay.otd.naim == otdel:
                        data.append(p)
                else:
                    data.append(p)
            oth_typ=True

        n = self.request.get('n',None)
        if n == '1':
            self.oth_1(data,oth_typ,otd,filter_otd)
        elif n == '2':
            self.oth_2(data)
        elif n == '3':
            self.oth_3(data,otd)
        elif n == '4':
            self.oth_4(data,otd)
        elif n == '5':
            self.oth_5(data,otd)
        elif n == '6':
            self.oth_6(data)
        elif n == '7':
            self.oth_7(data)
        elif n == '8':
            self.oth_8(data)
        elif n == '9':
            self.oth_9(data)
        elif n == 'a':
            self.oth_a(data)
        elif n == 'b':
            self.oth_b(data,otd)
        elif n == 'v':
            self.oth_v(data,otd)
        elif n == 'g':
            self.oth_g(data,otd)
        elif n == 'd':
            self.oth_d(data,otd)

    def rez_oth_1_2(self,data):
        bf = BetterFilter()
        sp = CountSluchaySpecification() ^ ProfKNSpecification() ^ IshUmerSpecification() ^ PolSpecification(1) ^ PolSpecification(2)
        all_temp = []
        for patient in data:
            for p in bf.filter(patient,sp):
                temp = bf.format_list(p)
                for t in range(len(temp)):
                    if temp[t] == 'None':
                        temp[t] = 0
                all_temp.append([int(i) for i in temp])

        all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(5)]
        return all_temp

    def rez_oth_1_4(self,data,t,c_oksm):
        # _ = [0,0,0,0,0]
        _ = []
        for d in data:
            adr = d.patient.m_roj
            c_oksm_p = d.patient.c_oksm.kod if d.patient.c_oksm else None
            if c_oksm_p == c_oksm == 643:
                if t == 'г.Тюменю':
                    if 'Тюмень' in adr:
                        _.append(d)
                elif t == 'Юг Тюм.обл.кроме Тюм.р-н':
                    if (('Тюменская обл' in adr) or ('обл. Тюменская' in adr) or ('ОБЛ ТЮМЕНСКАЯ' in adr)) \
                            and (('Тюменский р-н' not in adr) or ('р-н. Тюменский' not in adr)):
                        _.append(d)
                elif t == 'Тюменский р-н':
                    if 'Тюменский р-н' in adr or 'р-н. Тюменский' in adr:
                        _.append(d)
                elif t == 'Ханты-Мансйский АО':
                    if 'Ханты-Мансийский' in adr:
                        _.append(d)
                elif t == 'Ямало-Немецкий АО':
                    if 'Ямало-Ненецкий' in adr:
                        _.append(d)
                elif t == 'Др.регионы Россий':
                    if (('Тюменская обл' not in adr) and ('обл. Тюменская' not in adr) and ('ОБЛ ТЮМЕНСКАЯ' not in adr)) \
                            and (('Тюменский р-н' not in adr) and ('р-н. Тюменский' not in adr)):
                        _.append(d)
            else:
                _.append(d)
        return _

    def rez_oth_1_4_all(self,data):
        _ = [0, 0, 0, 0, 0]
        _[0] = len(data)
        for d in data:
            if d.sluchay.goc and d.sluchay.goc.tip_name == 'Экстренная':
                _[1] += 1
            if d.le_vr and d.le_vr.kd != None and d.le_vr.kd != '':
                _[2] += d.le_vr.kd
            if d.sluchay.oper.count() != 0:
                _[3] += 1
            if d.sluchay.icx and d.sluchay.icx.id_iz in [105, 106]:
                _[4] += 1
        return _

    def rez_oth_1_5(self,data):
        bf = BetterFilter()
        sp = CountSluchaySpecification() ^ GocEkSpecification() ^ ProfKNSpecification()^ \
             OperCountSpecification() ^ RezUmerSpecification()
        all_temp = []
        for patient in data:
            for p in bf.filter(patient,sp):
                temp = bf.format_list(p)
                for t in range(len(temp)):
                    if temp[t] == 'None':
                        temp[t] = 0
                all_temp.append([int(i) for i in temp])
        all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(5)]
        return all_temp

    def oth_1(self,data,oth_typ,otd,filter_otd):
        file = self.is_file('otd_rep_1.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            sheet1=wb.get_sheet_by_name('Лист1')
            sheet2=wb.get_sheet_by_name('Лист2')
            sheet3=wb.get_sheet_by_name('Лист3')
            sheet4=wb.get_sheet_by_name('Лист4')
            sheet5=wb.get_sheet_by_name('Лист5')

            if len(data) > 0:
                #Лист1
                sheet1.cell(row=3,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
                sheet1.cell(row=4, column=1).value = f'отделение {otd}'
                if oth_typ:
                    otd = otde.objects.values('id').filter(naim=otd)[0]['id']
                    otd = otde.objects.get(id=otd)
                    number_beds = otd.number_beds if otd.number_beds and otd.number_beds != 0 else 0
                else:
                    otds = otde.objects.filter(naim__in=filter_otd)
                    number_beds = 0
                    for o in otds:
                        number_beds += o.number_beds if o.number_beds and o.number_beds != 0 else 0
                sheet1.cell(row=10, column=2).value = number_beds
                sheet1.cell(row=11, column=2).value = number_beds

                datp_count = 0
                datv_count = 0
                rslt_umer = 0
                #не понятно
                otd_y = 0
                goc_ek_count=0
                goc_pl_count=0

                for d in data:
                    if self.date_1 <= d.sluchay.datp <= self.date_2:
                        datp_count+=1
                    if self.date_1 <= d.sluchay.datv <= self.date_2:
                        datv_count+=1
                    if d.sluchay.rslt and  d.sluchay.rslt.id_tip in [105,106]:
                        rslt_umer+=1
                    # if d.sluchay.otd_y and d.sluchay.otd_y.
                    if d.sluchay.goc:
                        if d.sluchay.goc.id_tip == 1:
                            goc_ek_count+=1
                        elif  d.sluchay.goc.id_tip == 3:
                            goc_pl_count+=1

                sheet1.cell(row=12, column=2).value = datp_count
                sheet1.cell(row=13, column=2).value = datv_count - rslt_umer
                sheet1.cell(row=14, column=2).value = rslt_umer
                try:
                    v = float('{0:.2f}'.format((datp_count-rslt_umer+datv_count) / 2))
                except ZeroDivisionError:
                    v = 0
                sheet1.cell(row=16, column=2).value = v if v != 0 else None
                sheet1.cell(row=18, column=2).value = number_beds * 29

                try:
                    v = float('{0:.2f}'.format((rslt_umer*100) / datv_count))
                except ZeroDivisionError:
                    v = 0
                sheet1.cell(row=29, column=2).value = f'{v}.%' if v != 0 else None

                sheet1.cell(row=32, column=3).value = goc_ek_count if goc_ek_count != 0 else None
                try:
                    v = float('{0:.2f}'.format((goc_ek_count*100) / len(data)))
                except ZeroDivisionError:
                    v = 0
                sheet1.cell(row=32, column=4).value = f'{v}%' if v != 0 else None
                sheet1.cell(row=33, column=3).value = goc_pl_count if goc_pl_count != 0 else None
                try:
                    v = float('{0:.2f}'.format((goc_pl_count*100) / len(data)))
                except ZeroDivisionError:
                    v = 0
                sheet1.cell(row=33, column=4).value = f'{v}%' if v != 0 else None
                #Лист2

                data_20_29 = []
                data_30_39 = []
                data_40_49 = []
                data_50_59 = []
                data_60_69 = []
                data_70_79 = []
                data_80 = []

                t_trv_1 = [0,0]
                t_trv_2 = [0,0]
                t_trv_3 = [0,0]
                t_trv_4 = [0,0]
                t_trv_all = [t_trv_1,t_trv_2,t_trv_3,t_trv_4]
                for d in data:
                    year = d.patient_year

                    if d.le_trv and d.le_trv.t_trv:
                        if d.le_trv.t_trv.kod == '6':
                            t_trv_1[0]+=1
                            if d.sluchay.alg and d.sluchay.alg in ['2']:
                                t_trv_1[1] += 1
                        elif d.le_trv.t_trv.kod == '5':
                            t_trv_2[0] += 1
                            if d.sluchay.alg and d.sluchay.alg in ['2']:
                                t_trv_2[1] += 1
                        elif d.le_trv.t_trv.kod == '7':
                            t_trv_3[0] += 1
                            if d.sluchay.alg and d.sluchay.alg in ['2']:
                                t_trv_3[1] += 1
                        elif d.le_trv.t_trv.kod == '10':
                            t_trv_4[0] += 1
                            if d.sluchay.alg and d.sluchay.alg in ['2']:
                                t_trv_4[1] += 1

                    if 20<= year <= 29:
                        data_20_29.append(d)
                    if 30 <= year <= 39:
                        data_30_39.append(d)
                    if 40 <= year <= 49:
                        data_40_49.append(d)
                    if 50 <= year <= 59:
                        data_50_59.append(d)
                    if 60 <= year <= 69:
                        data_60_69.append(d)
                    if 70 <= year <= 79:
                        data_70_79.append(d)
                    if 80 <= year:
                        data_80.append(d)

                rez_20 = self.rez_oth_1_2(data_20_29)
                rez_30 = self.rez_oth_1_2(data_30_39)
                rez_40 = self.rez_oth_1_2(data_40_49)
                rez_50 = self.rez_oth_1_2(data_50_59)
                rez_60 = self.rez_oth_1_2(data_60_69)
                rez_70 = self.rez_oth_1_2(data_70_79)
                rez_80 = self.rez_oth_1_2(data_80)

                rez_all = [rez_20,rez_30,rez_40,rez_50,rez_60,rez_70,rez_80]
                r = None
                for o in range(len(rez_all)):
                    if o == 0:
                        r = numpy.array(rez_all[o])
                    else:
                        r += numpy.array(rez_all[o])
                rez = r.tolist()
                row=3
                for r in rez_all:
                    row+=1
                    sheet2.cell(row=row, column=2).value = r[0] if r[0] != 0 else None
                    try:
                        v = float('{0:.2f}'.format((r[0] * 100) / rez[0]))
                    except ZeroDivisionError:
                        v = 0
                    sheet2.cell(row=row, column=3).value = v if v != 0 else None
                    sheet2.cell(row=row, column=4).value = r[1] if r[1] != 0 else None
                    try:
                        v = float('{0:.2f}'.format((r[1] * 100) / rez[1]))
                    except ZeroDivisionError:
                        v = 0
                    sheet2.cell(row=row, column=5).value = v if v != 0 else None
                    try:
                        v = float('{0:.2f}'.format(r[1] / r[0]))
                    except ZeroDivisionError:
                        v = 0
                    sheet2.cell(row=row, column=6).value = v if v != 0 else None
                    sheet2.cell(row=row, column=7).value = r[2] if r[2] != 0 else None
                    sheet2.cell(row=row, column=8).value = r[3] if r[3] != 0 else None
                    sheet2.cell(row=row, column=9).value = r[4] if r[4] != 0 else None
                else:
                    sheet2.cell(row=12, column=2).value = rez[0] if rez[0] != 0 else None
                    sheet2.cell(row=12, column=3).value = 100 if rez[0] != 0 else None
                    sheet2.cell(row=12, column=4).value = rez[1] if rez[1] != 0 else None
                    sheet2.cell(row=12, column=5).value = 100 if rez[1] != 0 else None
                    try:
                        v = float('{0:.2f}'.format(rez[1] / rez[0]))
                    except ZeroDivisionError:
                        v = 0
                    sheet2.cell(row=12, column=6).value = v if v != 0 else None
                    sheet2.cell(row=12, column=7).value = rez[2] if rez[2] != 0 else None
                    sheet2.cell(row=12, column=8).value = rez[3] if rez[3] != 0 else None
                    sheet2.cell(row=12, column=9).value = rez[4] if rez[4] != 0 else None

                #Лист3
                for o in range(len(t_trv_all)):
                    if o == 0:
                        r = numpy.array(t_trv_all[o])
                    else:
                        r += numpy.array(t_trv_all[o])
                rez = r.tolist()
                row=3
                for t in t_trv_all:
                    row+=1
                    sheet3.cell(row=row, column=3).value = t[0]
                    try:
                        v = float('{0:.2f}'.format((t[0] * 100) / rez[0]))
                    except ZeroDivisionError:
                        v = 0
                    sheet3.cell(row=row, column=4).value = v if v != 0 else None
                    sheet3.cell(row=row, column=5).value = t[1] if t[1] != 0 else None
                else:
                    row+=1
                    sheet3.cell(row=row, column=3).value = rez[0] if rez[0] != 0 else None
                    sheet3.cell(row=row, column=4).value = 100 if rez[0] != 0 else None
                    sheet3.cell(row=row, column=5).value = rez[1] if rez[1] != 0 else None

                #Лист4
                rez1 = self.rez_oth_1_4(data,t='г.Тюменю',c_oksm=643)
                rez2 = self.rez_oth_1_4(data, t='Юг Тюм.обл.кроме Тюм.р-н', c_oksm=643)
                rez3 = self.rez_oth_1_4(data, t='Тюменский р-н', c_oksm=643)
                rez4 = self.rez_oth_1_4(data, t='Ханты-Мансйский АО', c_oksm=643)
                rez5 = self.rez_oth_1_4(data, t='Ямало-Немецкий АО', c_oksm=643)
                rez6 = self.rez_oth_1_4(data, t='Др.регионы Россий', c_oksm=643)
                rez7 = self.rez_oth_1_4(data, t='', c_oksm=0)

                rez1_all = self.rez_oth_1_4_all(rez1)
                rez2_all = self.rez_oth_1_4_all(rez2)
                rez3_all = self.rez_oth_1_4_all(rez3)
                rez4_all = self.rez_oth_1_4_all(rez4)
                rez5_all = self.rez_oth_1_4_all(rez5)
                rez6_all = self.rez_oth_1_4_all(rez6)
                rez7_all = self.rez_oth_1_4_all(rez7)

                _ = [rez1_all,rez2_all,rez3_all,rez4_all,rez5_all,rez6_all,rez7_all]
                for o in range(len(_)):
                    if o == 0:
                        r = numpy.array(_[o])
                    else:
                        r += numpy.array(_[o])
                rez = r.tolist()
                _.append(rez)
                row=6
                for r in _:
                    row+=1
                    sheet4.cell(row=row, column=2).value = r[0] if r[0] != 0 else None
                    sheet4.cell(row=row, column=3).value = r[1] if r[1] != 0 else None
                    try:
                        v = float('{0:.2f}'.format(r[2] / r[0]))
                    except ZeroDivisionError:
                        v = 0
                    sheet4.cell(row=row, column=4).value = v if v != 0 else None
                    sheet4.cell(row=row, column=5).value = r[3] if r[3] != 0 else None
                    sheet4.cell(row=row, column=6).value = r[4] if r[4] != 0 else None

                sel = []
                for d in data:
                    if d.patient.cj and d.patient.cj.kod == 2:
                        sel.append(d)

                sel = self.rez_oth_1_4_all(sel)
                row=15
                sheet4.cell(row=row, column=2).value = sel[0] if sel[0] != 0 else None
                sheet4.cell(row=row, column=3).value = sel[1] if sel[1] != 0 else None
                try:
                    v = float('{0:.2f}'.format(sel[2] / sel[0]))
                except ZeroDivisionError:
                    v = 0
                sheet4.cell(row=row, column=4).value = v if v != 0 else None
                sheet4.cell(row=row, column=5).value = sel[3] if sel[3] != 0 else None
                sheet4.cell(row=row, column=6).value = sel[4] if sel[4] != 0 else None
                ing = []
                for d in data:
                    c_oksm_p = d.patient.c_oksm.kod if d.patient.c_oksm else None
                    adr = d.patient.m_roj
                    if c_oksm_p == 643:
                        if 'г.Тюмень' not in adr:
                            ing.append(d)

                ing = self.rez_oth_1_4_all(ing)
                row = 16
                sheet4.cell(row=row, column=2).value = ing[0] if ing[0] != 0 else None
                sheet4.cell(row=row, column=3).value = ing[1] if ing[1] != 0 else None
                try:
                    v = float('{0:.2f}'.format(ing[2] / ing[0]))
                except ZeroDivisionError:
                    v = 0
                sheet4.cell(row=row, column=4).value = v if v != 0 else None
                sheet4.cell(row=row, column=5).value = ing[3] if ing[3] != 0 else None
                sheet4.cell(row=row, column=6).value = ing[4] if ing[4] != 0 else None

                # Лист5
                list_lpu = get_list_lpu(data)
                row=6
                rez_all = []
                for l in list_lpu:
                    row+=1
                    sheet5.cell(row=row, column=1).value = l[0][0]
                    rez = self.rez_oth_1_5(l[1])
                    rez_all.append(rez)
                    sheet5.cell(row=row, column=2).value = rez[0] if rez[0] != 0 else None
                    sheet5.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
                    sheet5.cell(row=row, column=3).value = rez[1] if rez[1]!=0 else None
                    sheet5.cell(row=row, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
                    try:
                        v = float('{0:.2f}'.format(rez[2] / rez[0]))
                    except ZeroDivisionError:
                        v = 0
                    sheet5.cell(row=row, column=4).value = v if v != 0 else None
                    sheet5.cell(row=row, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
                    sheet5.cell(row=row, column=5).value = rez[3] if rez[3] != 0 else None
                    sheet5.cell(row=row, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
                    sheet5.cell(row=row, column=6).value = rez[4] if rez[4] != 0 else None
                    sheet5.cell(row=row, column=6).alignment = styles.Alignment(horizontal="center", vertical="center")
                    for r in range(6):
                        sheet5.cell(row=row, column=1+r).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
                else:
                    row+=1
                    sheet5.cell(row=row, column=1).value = 'Итого'
                    r = None
                    for o in range(len(rez_all)):
                        if o == 0:
                            r = numpy.array(rez_all[o])
                        else:
                            r += numpy.array(rez_all[o])
                    rez_all = r.tolist()
                    sheet5.cell(row=row, column=2).value = rez_all[0] if rez_all[0] != 0 else None
                    sheet5.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
                    sheet5.cell(row=row, column=3).value = rez_all[1] if rez_all[1]!=0 else None
                    sheet5.cell(row=row, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
                    try:
                        v = float('{0:.2f}'.format(rez_all[2] / rez_all[0]))
                    except ZeroDivisionError:
                        v = 0
                    sheet5.cell(row=row, column=4).value = v if v != 0 else None
                    sheet5.cell(row=row, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
                    sheet5.cell(row=row, column=5).value = rez_all[3] if rez_all[3] != 0 else None
                    sheet5.cell(row=row, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
                    sheet5.cell(row=row, column=6).value = rez_all[4] if rez_all[4] != 0 else None
                    sheet5.cell(row=row, column=6).alignment = styles.Alignment(horizontal="center", vertical="center")
                    for r in range(6):
                        sheet5.cell(row=row, column=1+r).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))

            wb.save(self.path() + f'otd_rep_1_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_vault_otd', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'download_vault_otd','text': self.path() + f'otd_rep_1_{self.user.user.id}.xlsx'})
    
    def oth_2(self,data):
         async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
    
    def oth_3(self,data,otd):
        file = self.is_file('otd_rep_3.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            sheet.cell(row=3, column=1).value = f'{otd}'
            sheet.cell(row=5,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'

            bf = BetterFilter()
            sp = CountSluchaySpecification() ^ OperCountSpecification() ^ OperAllCountSpecification() ^ OperAllCountGocEkSpecification() ^ \
                    OsloCountAllSpecification() ^ RezUmerOperSpecification() ^ OperAllKdSpecification() ^ PredOperKdSpecification() ^ CountSluchayPlanSpecification()
            all_temp = []
            for patient in data:
                for p in bf.filter(patient, sp):
                    temp = bf.format_list(p)
                    for t in range(len(temp)):
                        if temp[t] == 'None':
                            temp[t] = 0
                    all_temp.append([int(i) for i in temp])
            all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(9)]

            sheet.cell(row=6, column=2).value = all_temp[1] if all_temp[1] != 0 else None
            sheet.cell(row=7, column=2).value = all_temp[2] if all_temp[2] != 0 else None
            sheet.cell(row=8, column=2).value = all_temp[3] if all_temp[3] != 0 else None
            try:
                v = float('{0:.2f}'.format((all_temp[1]*100)/all_temp[0]))
            except ZeroDivisionError:
                v = 0
            sheet.cell(row=9, column=2).value = v if v != 0 else None
            sheet.cell(row=10, column=2).value = all_temp[4] if all_temp[4] != 0 else None
            try:
                v = float('{0:.2f}'.format((all_temp[4]*100)/all_temp[2]))
            except ZeroDivisionError:
                v = 0
            sheet.cell(row=11, column=2).value = v if v != 0 else None
            sheet.cell(row=12, column=2).value = all_temp[5] if all_temp[5] != 0 else None
            try:
                v = float('{0:.2f}'.format((all_temp[5]*100)/all_temp[1]))
            except ZeroDivisionError:
                v = 0
            sheet.cell(row=13, column=2).value = v if v != 0 else None
            try:
                v = float('{0:.2f}'.format((all_temp[6])/all_temp[1]))
            except ZeroDivisionError:
                v = 0
            sheet.cell(row=14, column=2).value = v if v != 0 else None
            try:
                v = float('{0:.2f}'.format((all_temp[7])/all_temp[8]))
            except ZeroDivisionError:
                v = 0
            sheet.cell(row=15, column=2).value = v if v != 0 else None

            opers = get_list_oper(data)
            row = 18
            count_oper = 0
            all_ek = 0
            for o in opers:
                row+=1
                v = V001.objects.get(kod=o[0][0])
                sheet.cell(row=row, column=1).value = f'{v.kod} {v.naim[:40]}'
                count_oper += len(o[1])
                sheet.cell(row=row, column=2).value = len(o[1]) if len(o[1]) != 0 else None
                sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center",vertical="center")
                sheet.cell(row=row, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
                ek = 0
                for d in o[1]:
                    if d.sluchay.goc and d.sluchay.goc.tip_name == 'Экстренная':
                        ek+=1
                all_ek+=ek
                sheet.cell(row=row, column=4).value = ek if ek != 0 else None
            else:
                row+=1
                sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")

            row = 18
            for o in opers:
                row+=1
                try:
                    v = float('{0:.2f}'.format((len(o[1])*100) / count_oper))
                except ZeroDivisionError:
                    v = 0
                sheet.cell(row=row, column=3).value = v if v != 0 else None

            row+=1
            sheet.cell(row=row, column=1).value = 'ИТОГО'
            sheet.cell(row=row, column=2).value = count_oper if count_oper != 0 else None
            sheet.cell(row=row, column=3).value = 100.0 if count_oper != 0 else None
            sheet.cell(row=row, column=4).value = all_ek if all_ek != 0 else None

            wb.save(self.path() + f'otd_rep_3_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_vault_otd', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'download_vault_otd','text': self.path() + f'otd_rep_3_{self.user.user.id}.xlsx'})
            
    def oth_4(self,data,otd):
        data = get_list_ds(data)
        file = self.is_file('otd_rep_4.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            sheet.cell(row=3,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=4, column=1).value = f'отделение {otd}'
            row = 7
            all_rez = []
            for d in data:
                row+=1
                rez = get_rez_rep_4(d[1])
                all_rez.append(rez)
                sheet.cell(row=row, column=1).value = f'{d[0][0].kod} {d[0][0].naim[:45]}'
                for n,v in enumerate(rez):
                    if n <= 10:
                        sheet.cell(row=row, column=2+n).value = v if v != 0 else None
                        sheet.cell(row=row, column=2+n).alignment = styles.Alignment(horizontal="center",vertical="center")
                else:
                    for c in range(1, 13):
                        sheet.cell(row=row, column=c).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
            else:
                row+=1
                sheet.cell(row=row, column=1).value = 'ВСЕГО'
                r = None
                for o in range(len(all_rez)):
                    if o == 0:
                        r = numpy.array(all_rez[o])
                    else:
                        r += numpy.array(all_rez[o])
                rez = r.tolist()
                rez = get_rez_rep_4(rez,d=True)
                for n,v in enumerate(rez):
                    if n <= 10:
                        sheet.cell(row=row, column=2+n).value = v if v != 0 else None
                        sheet.cell(row=row, column=2+n).alignment = styles.Alignment(horizontal="center",vertical="center")
            wb.save(self.path() + f'otd_rep_4_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_vault_otd', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'download_vault_otd','text': self.path() + f'otd_rep_4_{self.user.user.id}.xlsx'})

            async_to_sync(get_channel_layer().group_send)(self.user_group_name, {'type': 'download',
                                                                                 'text': self.path() + f'otd_rep_4_{self.user.user.id}.xlsx'})


    def oth_5(self,data,otd):
        file = self.is_file('otd_rep_5.xlsx')
        if file:
            wb = load_workbook(file)
            os.remove(file)

            sheet1 = wb.get_sheet_by_name('Лист1')
            sheet2 = wb.get_sheet_by_name('Лист2')
            sheet3 = wb.get_sheet_by_name('Лист3')
            sheet4 = wb.get_sheet_by_name('Лист4')
            sheet5 = wb.get_sheet_by_name('Лист5')
            sheet6 = wb.get_sheet_by_name('Лист6')

            #Лист1
            sheet1.cell(row=3, column=1).value = f'отделение {otd}'
            sheet1.cell(row=5,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            rez1 = get_rez_rep_5(data,t=1)
            row=8
            for r in rez1:
                row+=1
                for n,v in enumerate(r):
                    sheet1.cell(row=row, column=3+n).value = v if v != 0 else None
                    sheet1.cell(row=row, column=3+n).alignment = styles.Alignment(horizontal="center", vertical="center")
            #Лист2
            sheet2.cell(row=3, column=1).value = f'отделение {otd}'
            sheet2.cell(row=5,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            rez2 = get_rez_rep_5(data, t=2)
            row = 9
            for r2 in range(len(rez2)):
                row+=1
                _1 = [i for i in range(1,25) if i % 2 == 1]
                _0 = [i for i in range(1, 25) if i % 2 == 0]

                for n2,v2 in enumerate(rez2[r2]):
                    sheet2.cell(row=row, column=2+_1[n2]).value = v2 if v2 != 0 else None
                    sheet2.cell(row=row, column=2+_1[n2]).alignment = styles.Alignment(horizontal="center",vertical="center")
                    try:
                        v = float('{0:.2f}'.format((v2 * 100) / rez1[r2][n2]))
                    except ZeroDivisionError:
                        v = 0
                    sheet2.cell(row=row, column=2 + _0[n2]).value= v if v != 0 else None
                    sheet2.cell(row=row, column=2 + _0[n2]).alignment = styles.Alignment(horizontal="center",vertical="center")


            rez3_4 = get_list_pers(data)
            _ = []
            for r in rez3_4:
                if len(r[1]) > 1:
                    for n,v in enumerate(r[1]):
                        if n != 0:
                            _.append(v)

            # Лист3
            sheet3.cell(row=3, column=1).value = f'отделение {otd}'
            sheet3.cell(row=5,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            rez3 = get_rez_rep_5(data=_, t=1)
            row=8
            for r in rez3:
                row+=1
                for n,v in enumerate(r):
                    sheet3.cell(row=row, column=3+n).value = v if v != 0 else None
                    sheet3.cell(row=row, column=3+n).alignment = styles.Alignment(horizontal="center", vertical="center")
            #Лист4
            sheet4.cell(row=3, column=1).value = f'отделение {otd}'
            sheet4.cell(row=5,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            rez4 = get_rez_rep_5(data=_, t=2)
            row = 9
            for r2 in range(len(rez4)):
                row += 1
                _1 = [i for i in range(1, 25) if i % 2 == 1]
                _0 = [i for i in range(1, 25) if i % 2 == 0]

                for n2, v2 in enumerate(rez4[r2]):
                    sheet4.cell(row=row, column=2 + _1[n2]).value = v2 if v2 != 0 else None
                    sheet4.cell(row=row, column=2 + _1[n2]).alignment = styles.Alignment(horizontal="center",
                                                                                         vertical="center")
                    try:
                        v = float('{0:.2f}'.format((v2 * 100) / rez3[r2][n2]))
                    except ZeroDivisionError:
                        v = 0
                    sheet4.cell(row=row, column=2 + _0[n2]).value = v if v != 0 else None
                    sheet4.cell(row=row, column=2 + _0[n2]).alignment = styles.Alignment(horizontal="center",
                                                                                         vertical="center")
            #Лист5
            I60 = []
            I61 = []
            I63 = []
            G45_46 = []
            I67_I69 = []
            _ = []

            for d in data:
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzI60:
                    I60.append(d)
                elif d.sluchay.dskz and d.sluchay.dskz.kod in nzI61:
                    I61.append(d)
                elif d.sluchay.dskz and d.sluchay.dskz.kod in nzI63:
                    I63.append(d)
                elif d.sluchay.dskz and d.sluchay.dskz.kod in nzG45_46:
                    G45_46.append(d)
                elif d.sluchay.dskz and d.sluchay.dskz.kod in nzI67_I69:
                    I67_I69.append(d)
                else:
                    _.append(d)
            I60_data = [0,0,0,0,0,0,0,0,0,0,0,0]
            I61_data = [0,0,0,0,0,0,0,0,0,0,0,0]
            I63_data = [0,0,0,0,0,0,0,0,0,0,0,0]
            G45_46_data = [0,0,0,0,0,0,0,0,0,0,0,0]
            I67_I69_data = [0,0,0,0,0,0,0,0,0,0,0,0]
            _data = [0,0,0,0,0,0,0,0,0,0,0,0]

            I60 = get_rez_rep_5_5(I60,I60_data)
            I61 = get_rez_rep_5_5(I61,I61_data)
            I63 = get_rez_rep_5_5(I63,I63_data)
            G45_46 = get_rez_rep_5_5(G45_46,G45_46_data)
            I67_I69 = get_rez_rep_5_5(I67_I69, I67_I69_data)
            _ = get_rez_rep_5_5(_,_data)
            all_ds = [I60,I61,I63,G45_46,I67_I69,_]
            r = None
            for o in range(len(all_ds)):
                try:
                    if o == 0:
                        r = numpy.array(all_ds[o])
                    else:
                        r += numpy.array(all_ds[o])
                except:
                    pass

            rez = r.tolist()

            try:
                rez[6] = float('{0:.2f}'.format(rez[3] / rez[0]))
            except ZeroDivisionError:
                rez[6] = 0
            try:
                rez[7] = float('{0:.2f}'.format(rez[4] / rez[1]))
            except ZeroDivisionError:
                rez[7] = 0
            try:
                rez[8] = float('{0:.2f}'.format(rez[5] / rez[2]))
            except ZeroDivisionError:
                rez[8] = 0

            all_ds.append(rez)
            row=4
            for a in all_ds:
                row+=1
                for n,v in enumerate(a):
                    sheet5.cell(row=row, column=2+n).value = v if v != 0 else None

            #Лист6
            I61 = []
            I63 = []
            _ = []


            for d in data:
                if d.sluchay.dspat and d.sluchay.dspat.kod in nzI61:
                    I61.append(d)
                elif d.sluchay.dspat and d.sluchay.dspat.kod in nzI63:
                    I63.append(d)
                else:
                    _.append(d)
            I61_data = [0,0,0,0,0,0,0,0]
            I63_data = [0,0,0,0,0,0,0,0]
            _data = [0,0,0,0,0,0,0,0]

            I61r = get_rez_rep_5_6(I61,I61_data)
            I63r = get_rez_rep_5_6(I63,I63_data)
            _r = get_rez_rep_5_6(_,_data)
            all_ds = [I61r, I63r,_r]
            r = None
            for o in range(len(all_ds)):
                if o == 0:
                    r = numpy.array(all_ds[o])
                else:
                    r += numpy.array(all_ds[o])
            rez = r.tolist()
            try:
                rez[2] = float('{0:.2f}'.format(rez[1] / rez[0]))
            except ZeroDivisionError:
                rez[2] = 0
            all_ds.append(rez)
            row=4
            for a in all_ds:
                row+=1
                try:
                    a[2] = float('{0:.2f}'.format(a[1] / a[0]))
                except ZeroDivisionError:
                    a[2] = 0
                for n,v in enumerate(a):
                    sheet6.cell(row=row, column=2+n).value = v if v != 0 else None

            I61_data = [0,0,0,0,0,0,0]
            I63_data = [0,0,0,0,0,0,0]
            _data = [0,0,0,0,0,0,0]
            I61 = get_rez_rep_5_6_umer(I61,I61_data)
            I63 = get_rez_rep_5_6_umer(I63, I63_data)
            _ = get_rez_rep_5_6_umer(_,_data)
            all_ds = [I61,I63,_]
            r = None
            for o in range(len(all_ds)):
                if o == 0:
                    r = numpy.array(all_ds[o])
                else:
                    r += numpy.array(all_ds[o])
            rez = r.tolist()
            all_ds.append(rez)
            row=13
            for a in all_ds:
                row += 1
                for n, v in enumerate(a):
                    sheet6.cell(row=row, column=2 + n).value = v if v != 0 else None

            wb.save(self.path() + f'otd_rep_5_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_vault_otd', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'download_vault_otd','text': self.path() + f'otd_rep_5_{self.user.user.id}.xlsx'})

    def oth_6(self,data):
         async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})

    def oth_7(self,data):
         async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})

    def oth_8(self,data):
         async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})

    def oth_9(self,data):
        pass

    def oth_a(self,data):
        pass

    def oth_b(self,data,otd):
        file = self.is_file('otd_rep_b.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            dat = []
            ym = 0
            s1 = 0
            n = 0
            font = styles.Font(size=10, name='Arial')
            sheet.cell(row=3,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5,column=1).value = str(self.user.statistics_type.name).capitalize()
            sheet.cell(row=6, column=1).value = otd
            for d in data:
                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105,106]:
                    ord = OrderedDict()
                    n+=1
                    ord['otd'] = d.sluchay.otd.naim if d.sluchay.otd else None
                    ord['nib'] = f'{d.sluchay.nib}'
                    ord['fio'] = f'{d.patient.fam} {d.patient.im[0] if len(d.patient.im) > 0 else ""}.{d.patient.ot[0] if len(d.patient.ot) > 0 else ""}'
                    ord['age'] = f'{d.patient_year} {d.patient.nvs}'
                    ord['datp'] = d.sluchay.datp.strftime('%d.%m.%Y') if d.sluchay.datp else None
                    ord['datv'] = d.sluchay.datv.strftime('%d.%m.%Y') if d.sluchay.datv else None
                    ord['vra'] = d.sluchay.le_vr.kod.naim if d.sluchay.le_vr and d.sluchay.le_vr.kod else None
                    oper = get_pop_oper(d)
                    ord['dato'] = oper.dato.strftime('%d.%m.%Y') if oper else None
                    if d.sluchay.goc and d.sluchay.goc.tip_name == 'Экстренная':
                        ord['goc'] = 'экс'
                    elif d.sluchay.goc and d.sluchay.goc.tip_name == 'Плановая':
                        ord['goc'] = 'пл'
                    else:
                        ord['goc'] = ''
                    ord['vra_oper'] = f'{oper.kodx.kod} {oper.kodx.naim} {oper.kodx.ini}' if oper is not None and oper.kodx else None
                    ord['dskz'] = f'{d.sluchay.dskz.kod}-{d.sluchay.dskz.naim}' if d.sluchay.dskz else None
                    ord['dspat'] = f'{d.sluchay.dspat.kod}-{d.sluchay.dspat.naim}' if d.sluchay.dspat else None
                    ord['pri'] = d.sluchay.pri.naim if d.sluchay.pri else None
                    ord['vra_oper'] = f'{oper.kodx.naim}' if oper is not None and oper.kodx else None
                    ord['oper_kod'] = oper.kod_op.kod if oper is not None and oper.kod_op else None
                    if d.sluchay.oslo.count() > 0:
                        ord['oslo'] = 'oslo'
                    else:
                        ord['oslo'] = ''
                    if d.sluchay.otd_y and d.sluchay.otd_y.naim in ['АРО N1','АРО N2','АРО N3 (ЛДО)','ПРИЕМНОЕ']:
                        ym +=1
                    if d.sluchay.icx and d.sluchay.icx.id_iz == 106:
                        s1 +=1
                    dat.append(ord)
                    dat = sorted(dat,key=itemgetter('fio'))
            row = 7
            for d in dat:
                row+=1
                sheet.cell(row=row, column=1).value = f'{d["nib"][4:]} {d["fio"]}'
                sheet.cell(row=row, column=2).value = d['age']
                sheet.cell(row=row, column=3).value = d['datp']
                sheet.cell(row=row, column=4).value = d['datv']
                sheet.cell(row=row, column=5).value = d['goc']
                sheet.cell(row=row, column=6).value = d['vra']
                sheet.cell(row=row, column=7).value = f'1.{d["dskz"][:30]}' if d["dskz"] != None else ''
                row+=1
                sheet.cell(row=row, column=1).value = f'{d["otd"]}'
                sheet.cell(row=row, column=7).value = f'2.{d["dspat"][:30]}' if d["dspat"] != None else ''
                if d['pri'] != None:
                    row+=1
                    sheet.cell(row=row, column=7).value = f'3.{d["pri"][:20]}' if d['pri'] != None else ''
                if d['oper_kod'] != None:
                    row+=1
                    sheet.cell(row=row, column=4).value = d['dato']
                    sheet.cell(row=row, column=6).value = d['vra_oper']
                    sheet.cell(row=row, column=7).value = d['oper_kod']

                for c in range(1, 8):
                    sheet.cell(row=row, column=c).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
            else:
                row += 3
                sheet.merge_cells(f"A{sheet.cell(row=row, column=1).row}:B{sheet.cell(row=row, column=2).row}")
                sheet.cell(row=row, column=1).value = f'Умерло всего - {n}'
                sheet.cell(row=row, column=1).font = font
                row += 1
                sheet.merge_cells(f"A{sheet.cell(row=row, column=1).row}:B{sheet.cell(row=row, column=2).row}")
                sheet.cell(row=row, column=1).value = f'в том числе'
                sheet.cell(row=row, column=1).font = font
                row += 1
                sheet.merge_cells(f"A{sheet.cell(row=row, column=1).row}:C{sheet.cell(row=row, column=3).row}")
                sheet.cell(row=row, column=1).value = f'в реанимации и прием.отд-ии - {ym}'
                sheet.cell(row=row, column=1).font = font
                row += 1
                sheet.merge_cells(f"A{sheet.cell(row=row, column=1).row}:B{sheet.cell(row=row, column=2).row}")
                sheet.cell(row=row, column=1).value = f'в 1-е сутки - {s1}'
                sheet.cell(row=row, column=1).font = font
            wb.save(self.path() + f'otd_rep_b_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_vault_otd', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'download_vault_otd','text': self.path() + f'otd_rep_b_{self.user.user.id}.xlsx'})

    def oth_v(self,data,otd):
        file = self.is_file('otd_rep_v.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            dat = []
            font = styles.Font(size=10, name='Arial')
            sheet.cell(row=3,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = str(self.user.statistics_type.name).capitalize()
            sheet.cell(row=6, column=1).value = otd

            n = 0
            for d in data:
                if d.sluchay.oslo.count() > 0:
                    for o in d.sluchay.oslo.values('id'):
                        n += 1
                        oslo = Oslo.objects.get(id=o['id'])
                        ord = OrderedDict()
                        ord['n'] = n
                        ord['fio'] = f'{d.patient.fam} {d.patient.im[0] if len(d.patient.im) > 0 else ""}.{d.patient.ot[0] if len(d.patient.ot) > 0 else ""}'
                        ord['datp'] = d.sluchay.datp.strftime('%d.%m.%Y') if d.sluchay.datp else None
                        ord['datv'] = d.sluchay.datv.strftime('%d.%m.%Y') if d.sluchay.datv else None
                        ord['nib'] = f'{d.sluchay.nib}'
                        ord['age'] = f'{d.patient_year} {d.patient.nvs}'
                        ord['dskz'] = f'{d.sluchay.dskz.kod}-{d.sluchay.dskz.naim}' if d.sluchay.dskz else None
                        ord['osl'] = oslo.osl.naim if oslo.osl else None
                        ord['xosl'] = oslo.xosl.naim if oslo.xosl else None
                        ord['posl'] = oslo.posl.naim if oslo.posl else None
                        dat.append(ord)
            row = 7
            for d in dat:
                row+=1
                sheet.cell(row=row, column=1).value = d['n']
                sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=2).value = d['fio']
                sheet.cell(row=row, column=3).value = d['nib'][4:]
                sheet.cell(row=row, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=4).value = d['age']
                sheet.cell(row=row, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=5).value = f'1.{d["dskz"][:40]}' if d['dskz'] else None
                row+=1
                sheet.cell(row=row, column=2).value = f'{d["datp"]} - {d["datv"]}'
                sheet.cell(row=row, column=5).value = f'2.{d["osl"][:40]}'
                row+=1
                sheet.cell(row=row, column=5).value = f'3.{d["xosl"][:40]}'
                row+=1
                sheet.cell(row=row, column=5).value = f'4.{d["posl"][:40]}'
                for c in range(1,6):
                    sheet.cell(row=row, column=c).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
            wb.save(self.path() + f'otd_rep_v_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_vault_otd', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'download_vault_otd','text': self.path() + f'otd_rep_v_{self.user.user.id}.xlsx'})
    
    def oth_g(self,data,otd):
        file = self.is_file('otd_rep_g.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            sheet.cell(row=3,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = f'отделение {otd}'

            ds_oper = get_list_ds_oper(data)
            row = 8
            all_sl = []
            for d in ds_oper:
                row += 1
                ds = Ds.objects.values('id').filter(kod=d[0][0])
                ds = Ds.objects.get(id=ds[0]['id'])
                sheet.cell(row=row, column=1).value = f'{ds.kod} {ds.naim[:35]}'
                bf = BetterFilter()
                sp = CountSluchaySpecification() ^ ProfKNSpecification() ^ RezUmerSpecification() ^ OperAllCountSpecification()
                all_temp = []
                for patient in d[2]:
                    for p in bf.filter(patient, sp):
                        temp = bf.format_list(p)
                        for t in range(len(temp)):
                            if temp[t] == 'None':
                                temp[t] = 0
                        all_temp.append([int(i) for i in temp])
                all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(4)]
                all_sl.append(all_temp)

                sheet.cell(row=row, column=2).value = all_temp[0] if all_temp[0] != 0 else None
                sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=3).value = all_temp[1] if all_temp[1] != 0 else None
                sheet.cell(row=row, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
                try:
                    v = float('{0:.2f}'.format(all_temp[1]/all_temp[0]))
                except ZeroDivisionError:
                    v = 0
                sheet.cell(row=row, column=4).value = v if v != 0 else None
                sheet.cell(row=row, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=5).value = all_temp[2] if all_temp[2] != 0 else None
                sheet.cell(row=row, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=6).value = all_temp[3] if all_temp[3] != 0 else None
                sheet.cell(row=row, column=6).alignment = styles.Alignment(horizontal="center", vertical="center")
                if len(d[1]) > 0:
                    for n,o in enumerate(d[1],1):
                        if n != 1:
                            row+=1
                        k = V001.objects.get(kod=o)
                        sheet.cell(row=row, column=7).value = k.kod
                        sheet.cell(row=row, column=7).alignment = styles.Alignment(horizontal="center", vertical="center")
                        sheet.cell(row=row, column=8).value = k.naim[:25]
                        count_oper = 0
                        count_ym = 0
                        for c in d[2]:
                            opers = c.sluchay.oper.values('id')
                            for ope in opers:
                                oper = Oper.objects.get(id=ope['id'])
                                if oper.kod_op and oper.kod_op.kod == k.kod:
                                    count_oper +=1
                                    if c.sluchay.rslt and c.sluchay.rslt.id_tip in [105,106]:
                                        count_ym+=1

                        sheet.cell(row=row, column=9).value = count_oper if count_oper != 0 else None
                        sheet.cell(row=row, column=9).alignment = styles.Alignment(horizontal="center", vertical="center")
                        sheet.cell(row=row, column=10).value = count_ym if count_ym != 0 else None
                        sheet.cell(row=row, column=10).alignment = styles.Alignment(horizontal="center", vertical="center")
                    for c in range(1, 11):
                        sheet.cell(row=row, column=c).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
                else:
                    for c in range(1, 11):
                        sheet.cell(row=row, column=c).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
            else:
                row+=1
                sheet.cell(row=row, column=1).value = 'ИТОГО'
                sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
                r = None
                for o in range(len(all_sl)):
                    if o == 0:
                        r = numpy.array(all_sl[o])
                    else:
                        r += numpy.array(all_sl[o])
                rez = r.tolist()
                sheet.cell(row=row, column=2).value = rez[0] if rez[0] != 0 else None
                sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=3).value = rez[1] if rez[1] != 0 else None
                sheet.cell(row=row, column=3).alignment = styles.Alignment(horizontal="center", vertical="center")
                try:
                    v = float('{0:.2f}'.format(rez[1]/rez[0]))
                except ZeroDivisionError:
                    v = 0
                sheet.cell(row=row, column=4).value = v if v != 0 else None
                sheet.cell(row=row, column=4).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=5).value = rez[2] if rez[2] != 0 else None
                sheet.cell(row=row, column=5).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=6).value = rez[3] if rez[3] != 0 else None
                sheet.cell(row=row, column=6).alignment = styles.Alignment(horizontal="center", vertical="center")

            wb.save(self.path() + f'otd_rep_g_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_vault_otd', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'download_vault_otd','text': self.path() + f'otd_rep_g_{self.user.user.id}.xlsx'})

    def oth_d(self,data,otd):
        file = self.is_file('otd_rep_d.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            sheet.cell(row=3,column=1).value = f'За период с {self.date_1.strftime("%d.%m.%Y")} по {self.date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5, column=1).value = f'отделение {otd}'

            list_vra = []
            otdel = otde.objects.filter(naim=otd,dateend=None).first()

            for d in data:
                if d.sluchay.le_vr.kod and d.sluchay.le_vr.kod.kod not in list_vra:
                    if (otdel.kod_ot and d.sluchay.le_vr.kod.kod_ot) and (otdel.kod_ot.strip() == d.sluchay.le_vr.kod.kod_ot.strip()):
                        list_vra.append(d.sluchay.le_vr.kod.kod)

                if d.sluchay.oper.count() > 0:
                    opers = d.sluchay.oper.values('id')
                    for op in opers:
                        oper = Oper.objects.get(id=op['id'])
                        if oper.kodx and oper.kodx.kod not in list_vra:
                            if (otdel.kod_ot and oper.kodx.kod_ot) and (otdel.kod_ot.strip() == oper.kodx.kod_ot.strip()):
                                list_vra.append(oper.kodx.kod)


                if d.sluchay.manpy.count() > 0:
                    manpys = d.sluchay.manpy.values('id')
                    for m in manpys:
                        manpy = Manpy.objects.get(id=m['id'])
                        if manpy.tnvr and manpy.tnvr.kod not in list_vra:
                            if (otdel.kod_ot and manpy.tnvr.kod_ot) and (otdel.kod_ot.strip() == manpy.tnvr.kod_ot.strip()):
                                list_vra.append(manpy.tnvr.kod)

            row = 8
            count_all_vra = []
            for v in list_vra:
                row+=1
                vra = Vra.objects.values('id').filter(kod=v)[0]['id']
                vra = Vra.objects.get(id=vra)

                sheet.cell(row=row, column=1).value = vra.kod
                sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")

                sheet.cell(row=row, column=2).value = vra.naim
                sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
                sl_count = 0
                ek_count = 0
                lek_kd = 0
                oper_count = 0
                oper_op_count = 0
                oper_e_count = 0
                oper_kodxa_count = 0
                manpy_con_count = 0
                manpy_count = 0

                for d in data:
                    if d.sluchay.le_vr.kod and d.sluchay.le_vr.kod.kod == vra.kod:
                        sl_count+=1
                        if d.sluchay.goc and d.sluchay.goc.tip_name == 'Экстренная':
                            ek_count += 1

                        if d.sluchay.le_vr.kd != None and d.sluchay.le_vr.kd != '':
                            lek_kd+=d.sluchay.le_vr.kd

                    if d.sluchay.oper.count() > 0:
                        if d.sluchay.le_vr.kod and d.sluchay.le_vr.kod.kod == vra.kod:
                            oper_count+=1
                        opers = d.sluchay.oper.values('id')
                        for op in opers:
                            oper = Oper.objects.get(id=op['id'])
                            if oper.kodx and oper.kodx.kod == vra.kod:
                                oper_op_count+=1
                            if oper.pr_osob.count() > 0:
                                pr_osobs = oper.pr_osob.values('id')
                                for pr in pr_osobs:
                                    pr_osob = PR_OSOB.objects.get(id=pr['id'])
                                    if pr_osob.kod == 'Э':
                                        oper_e_count+=1
                            if oper.kodxa and oper.kodxa.kod == vra.kod:
                                oper_kodxa_count += 1
                            if oper.kodxa1 and oper.kodxa1.kod == vra.kod:
                                oper_kodxa_count += 1
                    if d.sluchay.manpy.count() > 0:
                        manpys = d.sluchay.manpy.values('id')
                        for m in manpys:
                            manpy = Manpy.objects.get(id=m['id'])
                            if manpy.tnvr and manpy.tnvr.kod == vra.kod:
                                manpy_count+=1
                                if manpy.kodmn and manpy.kodmn.kod == '01013':
                                    manpy_con_count+=1
                count_all_vra.append(
                    [sl_count,ek_count,lek_kd,oper_count,oper_op_count,oper_e_count,oper_kodxa_count,manpy_count,manpy_con_count])

                _ = [sl_count,ek_count,lek_kd,oper_count,oper_op_count,oper_e_count,oper_kodxa_count,manpy_count,manpy_con_count]
                for n,v in enumerate(_):
                    sheet.cell(row=row, column=3+n).value = v if v != 0 else None
                    sheet.cell(row=row, column=3 + n).alignment = styles.Alignment(horizontal="center", vertical="center")

                for c in range(1,13):
                    sheet.cell(row=row, column=c).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))

            r = None
            for o in range(len(count_all_vra)):
                if o == 0:
                    r = numpy.array(count_all_vra[o])
                else:
                    r += numpy.array(count_all_vra[o])
            rez = r.tolist()
            row+=1
            sheet.cell(row=row, column=2).value = 'ИТОГО'
            sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
            for n,v in enumerate(rez):
                sheet.cell(row=row, column=3 + n).value = v if v != 0 else None
                sheet.cell(row=row, column=3 + n).alignment = styles.Alignment(horizontal="center", vertical="center")
            for c in range(1, 13):
                sheet.cell(row=row, column=c).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
            wb.save(self.path() + f'otd_rep_d_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_vault_otd', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'download_vault_otd','text': self.path() + f'otd_rep_d_{self.user.user.id}.xlsx'})


# class VaultOtdTTT(AnnualReportABC):
#     def __init__(self,user, request):
#         super().__init__(user, request)
#     def create(self):
#         pass
# class VaultOtdNNN(AnnualReportABC):
#     def __init__(self,user, request):
#         super().__init__(user, request)
#         self.user_group_name = 'hospital_reports_%s' % user
#     def create(self):
#         file = self.is_file('vault_otd_rep_nnn.xlsx')
#         if file:
#             wb = load_workbook(file)
#             sheet = wb.active
#             os.remove(file)
#             patients = PatientsData(self.date_1, self.date_2, self.user)
#             patients.sluchays()
#             data = []
#             for p in patients.patients:
#                 if p.sluchay.otd and p.sluchay.otd.naim in ('НЕВРОЛОГИЯ N1','НЕВРОЛОГИЯ N2','НЕВРОЛОГИЯ N3'):
#                     if (p.sluchay.icx and p.sluchay.icx.id_iz in (105, 106)) \
#                             or (p.sluchay.rslt and p.sluchay.rslt.id_tip in (105, 106)):
#                         data.append(p)
#             dic = dict([('sheet', sheet), ('data', data), ('name', self.user.statistics_type.name),
#                         ('date_1', self.date_1), ('date_2', self.date_2)])
#             insert_sheet_nnn(**dic)
#             wb.save(self.path() + f'group_nnn_{self.user.user.id}.xlsx')
#             async_to_sync(get_channel_layer().group_send)(self.user_group_name,
#                                                           {'type': 'report_vault_otd', 'text': 'Отчет cфромирован'})
#             async_to_sync(get_channel_layer().group_send)(self.user_group_name,
#                                                           {'type': 'download_vault_otd',
#                                                            'text': self.path() + f'group_nnn_{self.user.user.id}.xlsx'})

# class VaultOtdHH(AnnualReportABC):
#     def __init__(self,user, request):
#         super().__init__(user, request)
#         self.user_group_name = 'hospital_reports_%s' % user
#     def create(self):
#         pass
#         # file = self.is_file('vault_otd_rep_nnn.xlsx')
#         # if file:
#         #     wb = load_workbook(file)
#         #     sheet = wb.active
#         #     os.remove(file)
#         #     patients = PatientsData(self.date_1, self.date_2, self.user)
#         #     patients.sluchays()
#         #     data = []
#         #     for p in patients.patients:
#         #         if p.sluchay.otd and p.sluchay.otd.naim in ['НЕВРОЛОГИЯ N1','НЕВРОЛОГИЯ N2','НЕВРОЛОГИЯ N3']:
#         #             if

class AOth1(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth2(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth3(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth4(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth5(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('a_oth_5.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()

            filters = json.loads(self.request.get('filters'))
            if len(filters['filter']) > 0:
                patients_filter = PatientsDataFiltrs(self.date_1, self.date_2, self.user, self.request)
                data_filter = patients_filter.patients
            else:
                data_filter = []
            dic = dict([('sheet', sheet), ('data', patients.patients),('data_filter',data_filter), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2),('user',self.user),('filters',self.filters_list())])
            insert_sheet_a_oth_5(**dic)
            wb.save(self.path() + f'a_oth_5_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download_a_oth',
                                                           'text': self.path() + f'a_oth_5_{self.user.user.id}.xlsx',
                                                           'btn':'download_a_oth_5'
                                                           },
                                                          )
class AOth6(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth7(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('a_oth_7.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2),('user',self.user)])
            insert_sheet_a_oth_7(**dic)
            wb.save(self.path() + f'a_oth_7_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download_a_oth',
                                                           'text': self.path() + f'a_oth_7_{self.user.user.id}.xlsx',
                                                           'btn':'download_a_oth_7'
                                                           },
                                                          )
class AOth8(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth9(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth10(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth11(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth12(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth13(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth14(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth15(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth16(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth17(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth18(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth19(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('a_oth_19.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsDataFiltrs(self.date_1, self.date_2, self.user,self.request)

            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2), ('user', self.user)])
            insert_sheet_a_oth_19(**dic)
            wb.save(self.path() + f'a_oth_19_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download_a_oth',
                                                           'text': self.path() + f'a_oth_19_{self.user.user.id}.xlsx',
                                                           'btn':'download_a_oth_19'
                                                           },
                                                          )
class AOth20(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('a_oth_20.xlsx')
        if file:
            wb = load_workbook(file)
            os.remove(file)
            sheet = wb.active
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2),('filters',self.filters_list())])
            insert_sheet_a_oth_20(**dic)
            wb.save(self.path() + f'a_oth_20_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download_a_oth',
                                                           'text': self.path() + f'a_oth_20_{self.user.user.id}.xlsx'
                                                           }
                                                          )
class AOth21(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth22(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth23(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('a_oth_23.xlsx')
        if file:
            wb = load_workbook(file)
            os.remove(file)
            sheet = wb.active
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_a_oth_23(**dic)
            wb.save(self.path() + f'a_oth_23_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download_a_oth',
                                                           'text': self.path() + f'a_oth_23_{self.user.user.id}.xlsx'
                                                           }
                                                          )
class AOth24(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth25(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth26(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth27(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth28(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth29(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('a_oth_29.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsDataFiltrs(self.date_1, self.date_2, self.user, self.request)
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2), ('filters', self.filters_list())])
            insert_sheet_a_oth_29(**dic)

            wb.save(self.path() + f'a_oth_29_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download', 'text': self.path() + f'a_oth_29_{self.user.user.id}.xlsx',
                                                           'name': 'Отчет о выбывших'})

class AOth30(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth31(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})
class AOth32(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        file = self.is_file('a_oth_32.xlsx')
        if file:
            wb = load_workbook(file)
            os.remove(file)
            # sheet = wb.active
            sheet=wb.get_sheet_by_name('инфо')
            sheet1=wb.get_sheet_by_name('список')
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            data = []
            for p in patients.patients:
                if p.sluchay.otd and p.sluchay.otd.naim == 'ТОКСИКОЛОГИЯ':
                    data.append(p) 

            dic = dict([('sheet', [sheet,sheet1]), ('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])

            insert_sheet_a_oth_32(**dic)
            wb.save(self.path() + f'a_oth_32_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download_a_oth',
                                                           'text': self.path() + f'a_oth_32_{self.user.user.id}.xlsx'
                                                           }
                                                          )
class AOth33(AnnualReportABC):
    def __init__(self,user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_reports_%s' % user
    def create(self):
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'error_messages'})

class AOth36(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user,request)
        self.user_group_name = 'hospital_reports_%s' % user

    def create(self):
        file = self.is_file('a_oth_36.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_a_oth_36(**dic)
            wb.save(self.path() + f'a_oth_36_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download_a_oth',
                                                           'text': self.path() + f'a_oth_36_{self.user.user.id}.xlsx'
                                                           }
                                                          )

















def ReferenceReport(user,request):
    print(request)
    type_fun = request.get('group_p_list')
    if type_fun == 'group_p1':
        report = GroupP1(user,request)
        report.create()
    elif type_fun == 'implants':
        report = Implants(user,request)
        report.create()
    elif type_fun == 'group_p2':
        report = GroupP2(user,request)
        report.create()
    elif type_fun == 'group_p3':
        report = GroupP3(user,request)
        report.create()
    elif type_fun == 'group_p4':
        report = GroupP4(user,request)
        report.create()
    elif type_fun == 'group_p5':
        report = GroupP5(user,request)
        report.create()
    elif type_fun == 'group_p6':
        report = GroupP6(user,request)
        report.create()
    elif type_fun == 'group_p7':
        report = GroupP7(user,request)
        report.create()
    elif type_fun == 'group_p8':
        report = GroupP8(user,request)
        report.create()
    elif type_fun == 'group_p9':
        report = GroupP9(user,request)
        report.create()
    elif type_fun == 'group_p10':
        report = GroupP10(user,request)
        report.create()
    elif type_fun == 'group_p11':
        report = GroupP11(user,request)
        report.create()
    elif type_fun == 'group_p12':
        report = GroupP12(user,request)
        report.create()
    elif type_fun == 'group_p13':
        report = GroupP13(user,request)
        report.create()
    elif type_fun == 'group_p14':
        report = GroupP14(user,request)
        report.create()
    elif type_fun == 'group_p15':
        report = GroupP15(user,request)
        report.create()
    elif type_fun == 'group_p16':
        report = GroupP16(user,request)
        report.create()
    elif type_fun == 'group_p17':
        report = GroupP17(user,request)
        report.create()
    elif type_fun == 'group_p18':
        report = GroupP18(user,request)
        report.create()
    elif type_fun == 'group_p19':
        report = GroupP19(user,request)
        report.create()
    elif type_fun == 'group_p20':
        report = GroupP20(user,request)
        report.create()
    elif type_fun == 'group_p21':
        report = GroupP21(user,request)
        report.create()
    elif type_fun == 'group_p22':
        report = GroupP22(user,request)
        report.create()
    elif type_fun == 'group_p23':
        report = GroupP23(user,request)
        report.create()
    elif type_fun == 'group_p24':
        report = GroupP24(user,request)
        report.create()
    elif type_fun == 'group_p25':
        report = GroupP25(user,request)
        report.create()

    if type_fun == 'vault_otd_rep':
        report = VaultOtd(user, request)
        report.create()
    # elif type_fun == 'vault_otd_rep_nnn':
    #     report = VaultOtdNNN(user,request)
    #     report.create()
    # elif type_fun == 'vault_otd_rep_hh':
    #     report = VaultOtdHH(user,request)
    #     report.create()


    type_fun = request.get('type_report')
    if type_fun == 'a_oth_1':
        report = AOth1(user,request)
        report.create()
    elif type_fun == 'a_oth_2':
        report = AOth2(user,request)
        report.create()
    elif type_fun == 'a_oth_3':
        report = AOth3(user,request)
        report.create()
    elif type_fun == 'a_oth_4':
        report = AOth4(user,request)
        report.create()
    elif type_fun == 'a_oth_5':
        report = AOth5(user,request)
        report.create()
    elif type_fun == 'a_oth_6':
        report = AOth6(user,request)
        report.create()
    elif type_fun == 'a_oth_7':
        report = AOth7(user,request)
        report.create()
    elif type_fun == 'a_oth_8':
        report = AOth8(user,request)
        report.create()
    elif type_fun == 'a_oth_9':
        report = AOth9(user,request)
        report.create()
    elif type_fun == 'a_oth_10':
        report = AOth10(user,request)
        report.create()
    elif type_fun == 'a_oth_11':
        report = AOth11(user,request)
        report.create()
    elif type_fun == 'a_oth_12':
        report = AOth12(user,request)
        report.create()
    elif type_fun == 'a_oth_13':
        report = AOth13(user,request)
        report.create()
    elif type_fun == 'a_oth_14':
        report = AOth14(user,request)
        report.create()
    elif type_fun == 'a_oth_15':
        report = AOth15(user,request)
        report.create()
    elif type_fun == 'a_oth_16':
        report = AOth16(user,request)
        report.create()
    elif type_fun == 'a_oth_17':
        report = AOth17(user,request)
        report.create()
    elif type_fun == 'a_oth_18':
        report = AOth18(user,request)
        report.create()
    elif type_fun == 'a_oth_19':
        report = AOth19(user,request)
        report.create()
    elif type_fun == 'a_oth_20':
        report = AOth20(user,request)
        report.create()
    elif type_fun == 'a_oth_21':
        report = AOth21(user,request)
        report.create()
    elif type_fun == 'a_oth_22':
        report = AOth22(user,request)
        report.create()
    elif type_fun == 'a_oth_23':
        report = AOth23(user,request)
        report.create()
    elif type_fun == 'a_oth_24':
        report = AOth24(user,request)
        report.create()
    elif type_fun == 'a_oth_25':
        report = AOth25(user,request)
        report.create()
    elif type_fun == 'a_oth_26':
        report = AOth26(user,request)
        report.create()
    elif type_fun == 'a_oth_27':
        report = AOth27(user,request)
        report.create()
    elif type_fun == 'a_oth_28':
        report = AOth28(user,request)
        report.create()
    elif type_fun == 'a_oth_29':
        report = AOth29(user,request)
        report.create()
    elif type_fun == 'a_oth_30':
        report = AOth30(user,request)
        report.create()
    elif type_fun == 'a_oth_31':
        report = AOth31(user,request)
        report.create()
    elif type_fun == 'a_oth_32':
        report = AOth32(user,request)
        report.create()
    elif type_fun == 'a_oth_33':
        report = AOth33(user,request)
        report.create()
    elif type_fun == 'a_oth_36':
        report = AOth36(user,request)
        report.create()
