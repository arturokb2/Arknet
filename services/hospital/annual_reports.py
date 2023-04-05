from genericpath import isfile
from itertools import count
from operator import le
from re import T
from services.hospital.reports import Reports
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from openpyxl.styles import numbers
from openpyxl import Workbook
from  openpyxl.styles import Font,Alignment,Border,Side
from openpyxl import load_workbook,styles
import json
import copy
import os
import shutil
from django.conf import  settings
from okb2.models import MyUser,Ds,otde
from celery.contrib import rdb
from docxtpl import DocxTemplate

from abc import  ABC
from datetime import datetime
from services.hospital.patient import PatientsData
from services.hospital.reports import *
from collections import OrderedDict
import services.hospital.reference_reports as rr
import numpy
# from services.hospital.reference_reports import get_list_otd_prof 
# from services.hospital.reference_reports import get_list_otd_prof

from okb2.models import V001,Ds,Ab_Obsh
from hospital.models import Manpy,Oper
from django.db.models import Q



border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin', color='000000'))

ymkd = 0

not_oper_count_otdl = ['ЭНДОКРИНОЛОГИЧЕСКОЕ','МЕД.РЕАБИЛИТАЦИЯ','ПУЛЬМОНОЛОГИЯ',
                       'ТОКСИКОЛОГИЯ','НЕВРОЛОГИЯ N1','НЕВРОЛОГИЯ N3','КАРДИОЛОГИЧЕСКОЕ']
class Create(Reports):
    def __init__(self,user,request):
        super().__init__(user,request)
        self._user_group_name = 'hospital_annual_reports_%s' % user
        # self.list_data = json.loads(self.request['list_data']) if self.request.get('list_data') != None else []
        self._path_shoblons = 'shoblons/hospital/forms'
        self._path_shoblons_oth = 'shoblons/hospital/oth'
        self._temp_dir = 'temp'
        self._border = Border(left=Side(border_style='thin',color='000000'),
                            right=Side(border_style='thin', color='000000'),
                            top=Side(border_style='thin',color='000000'),
                            bottom=Side(border_style='thin', color='000000'))


    def create(self):
        self._sluchays = self.get_sluchays()
        self.old_sluchays = copy.deepcopy(self._sluchays)
        type_fun = self._request.get('type_report')
        
        #Допилить с фильтрами
        if self._request.get('filters',None) != None:
            filters = json.loads(self._request.get('filters'))
            if filter:
                self._sluchays = self.filter(filters, self._sluchays)
        
        self._filename = self._request.get('filename', None)
        self._filename = f'{self._filename}.xlsx' if self._filename != None and len(self._filename) > 0 else f'{type_fun}_{self._user.user.id}.xlsx'

        if type_fun == 'annual_13_1_1':
            self._annual_13_1_1()
        elif type_fun == 'annual_13_1_2':
            self._annual_13_1_2()
        elif type_fun == 'annual_13_1_3':
            self._annual_13_1_3()
        elif type_fun == 'annual_13_1_4':
            self._annual_13_1_4()
        elif type_fun == 'annual_13_1_5':
            self._annual_13_1_5()
        elif type_fun == 'annual_14_1_1':
            self._annual_14_1_1()
        elif type_fun == 'annual_14_1_2':
            self._annual_14_1_2()
        elif type_fun == 'annual_14_1_3':
            self._annual_14_1_3()
        elif type_fun == 'annual_14_1_4':
            self._annual_14_1_4()
        elif type_fun == 'annual_14_1_5':
            self._annual_14_1_5()
        elif type_fun == 'annual_14_2_1':
            self._annual_14_2_1()
        elif type_fun == 'annual_14_3_1':
            self._annual_14_3_1()
        elif type_fun == 'annual_14_3_2':
            self._annual_14_3_2()
        elif type_fun == 'annual_14_3_3':
            self._annual_14_3_3()
        elif type_fun == 'annual_14_3_4':
            self._annual_14_3_4()
        elif type_fun == 'annual_14_3_5':
            self._annual_14_3_5()
        elif type_fun == 'annual_14_3_6':
            self._annual_14_3_6()
        elif type_fun == 'annual_14_3_7':
            self._annual_14_3_7()
        elif type_fun == 'annual_14_3_8':
            self._annual_14_3_8()
        elif type_fun == 'annual_30_1_1':
            self._annual_30_1_1()
        elif type_fun == 'annual_30_2_1':
            self._annual_30_2_1()
        elif type_fun == 'annual_30_2_2':
            self._annual_30_2_2()
        elif type_fun == 'annual_30_2_3':
            self._annual_30_2_3()
        elif type_fun == 'annual_30_3_1':
            self._annual_30_3_1()
        elif type_fun == 'annual_16_1_1':
            self._annual_16_1_1()
        elif type_fun == 'annual_57_1_1':
            self._annual_57_1_1()
        elif type_fun == 'annual_57_1_2':
            self._annual_57_1_2()
        elif type_fun == 'annual_57_1_3':
            self._annual_57_1_3()
        elif type_fun == 'annual_57_1_4':
            self._annual_57_1_4()
        elif type_fun == 'annual_pr_1':
            self._annual_pr_1()
        elif type_fun == 'annual_pr_2':
            self._annual_pr_2()
        elif type_fun == 'annual_pr_3':
            self._annual_pr_3()
        elif type_fun == 'annual_pr_4':
            self._annual_pr_4()
        elif type_fun == 'annual_pr_5':
            self._annual_pr_5()
        elif type_fun == 'annual_pr_6':
            self._annual_pr_6()
        elif type_fun == 'annual_pr_7':
            self._annual_pr_7()
        elif type_fun == 'annual_pr_8':
            self._annual_pr_8()
        elif type_fun == 'annual_pr_9':
            self._annual_pr_9()
        elif type_fun == 'annual_pr_a':
            self._annual_pr_a()
        elif type_fun == 'annual_pr_b':
            self._annual_pr_b()
        elif type_fun == 'annual_pr_v':
            self._annual_pr_v()
        elif type_fun == 'annual_pr_g':
            self._annual_pr_g()
        elif type_fun == 'annual_pr_d':
            self._annual_pr_d()
        


    def _annual_13_1_1(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
            doc = DocxTemplate(file_new)
            context = {'i_all':111,
                       'i_1':222,
                       'i_2':333}
            doc.render(context)
            doc.save(file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_13_1_2(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_13_1_3(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_13_1_4(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_13_1_5(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_1_1(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_1_2(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_1_3(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons, 'Ф13.docx'])
        if os.path.exists(file_shoblon):
            file_new = 'Ф13.docx'
            file_new = '/'.join([settings.MEDIA_ROOT, self._temp_dir, str(self._user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_1_4(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_1_5(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_2_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_2(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_3(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_4(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_5(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_6(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_7(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_14_3_8(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_30_1_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_30_2_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_30_2_2(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_30_2_3(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_30_3_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})


    def _annual_16_1_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_57_1_1(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_57_1_2(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_57_1_3(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_57_1_4(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

    def _annual_pr_1(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT,self._path_shoblons_oth,'annual_pr_1.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT,self._temp_dir,str(self._user.id),file_new])
            shutil.copy2(file_shoblon,file_new)
            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self._date_1.strftime("%d.%m.%Y")} по {self._date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5,column=1).value = str(self._user.statistics_type.name).capitalize()
            filetrs = ['otd','prof_k','count_sluchay','prof_k_n','goc_ek','rez_umer','rez_umer_goc_ek','rez_umer_goc_ek_sr']
            temp = self.filetr_sluchays(filetrs,self._sluchays)
            # res_set = [[r[0],r[1],0,0,0,0,0,0,0,0,0,0,0,0,0,0,0] for r in set(rez for rez in temp)]
            row = 9
            # res_set = [r for r in list(set(rez for rez in temp))]

            res_set = set()
            for t in temp:
                x = (t[0],t[1],None,None,None,None,None,None,None,None,None,None,None,None,None,None,None)
                res_set.add(x)
            
            res_set = [list(r) for r in res_set]
            for t in temp:
                for r in res_set:
                    if r[0] == t[0] and r[1] == t[1]:
                        r[2] = r[2] + t[2] if r[2] != None else t[2]
                        r[3] = r[3] + t[3] if r[3] != None else t[3]
                        r[5] = r[5] + t[4] if r[5] != None else t[4]
                        r[8] = r[8] + t[5] if r[8] != None else t[5]
                        r[11] = r[11] + t[6] if r[11] != None else t[6]
                        r[13] = r[13] + t[7] if r[13]!= None else t[7]
            
            for r in res_set:
                try:
                    r[4] = float("{0:.2f}".format(r[3] / r[2] ))
                except ZeroDivisionError:
                    r[4] = r[3]

            max_otd = max(len(r[0]) for r in res_set)
            max_prof_k = max(len(r[1]) for r in res_set)
            col_otd = str(sheet.cell(column=1, row=6)).split('.')[1][:-1]
            col_max_prof_k = str(sheet.cell(column=2, row=6)).split('.')[1][:-1]
            sheet.column_dimensions[col_otd[0]].width = max_otd + 10
            sheet.column_dimensions[col_max_prof_k[0]].width = max_prof_k + 10

            for i, res in enumerate(res_set):
                for l, r in enumerate(res):
                    sheet.cell(row=i + row,column=1 + l).value = r
                    sheet.cell(row=i + row, column=1 + l).alignment = Alignment(horizontal="center", vertical="center")
                    sheet.cell(row=i + row, column=1 + l).border = self._border
            # wb.save(file_new)
            self._path()
            wb.save(self._dir + self._filename)
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'download', 'text': self._dir + self._filename,'name':'Отчет о работе отделений (по ИБ)'})
    def _annual_pr_2(self):
        file_shoblon = '/'.join([settings.MEDIA_ROOT, self._path_shoblons_oth, 'annual_pr_2_1.xlsx'])
        if os.path.exists(file_shoblon):
            file_new = self._filename
            file_new = '/'.join([settings.MEDIA_ROOT,self._temp_dir,str(self._user.id),file_new])
            shutil.copy2(file_shoblon,file_new)
            wb = load_workbook(file_new)
            sheet = wb.active
            sheet.cell(row=4,column=1).value = f'За период с {self._date_1.strftime("%d.%m.%Y")} по {self._date_2.strftime("%d.%m.%Y")} г.'
            sheet.cell(row=5,column=1).value = str(self._user.statistics_type.name).capitalize()
            row = 10
            filetrs = ['otd', 'prof_k','count_oper','goc_ek','count_oper_all','goc_ek_oper']
            temp = self.filetr_sluchays(filetrs, self._sluchays)
            res_set = set()
            for t in temp:
                x = (
                t[0], t[1], None, None, None, None, None, None, None, None, None, None, None, None)
                res_set.add(x)

            res_set = [list(r) for r in res_set]
            for t in temp:
                for r in res_set:
                    if r[0] == t[0] and r[1] == t[1]:
                        r[2] = r[2] + t[2] if r[2] != None else t[2]
                        r[3] = r[3] + t[3] if r[3] != None else t[3]
                        r[4] = r[4] + t[4] if r[4] != None else t[4]
                        r[5] = r[5] + t[5] if r[5] != None else t[5]

            max_otd = max(len(r[0]) for r in res_set)
            max_prof_k = max(len(r[1]) for r in res_set)
            col_otd = str(sheet.cell(column=1, row=6)).split('.')[1][:-1]
            col_max_prof_k = str(sheet.cell(column=2, row=6)).split('.')[1][:-1]
            sheet.column_dimensions[col_otd[0]].width = max_otd + 10
            sheet.column_dimensions[col_max_prof_k[0]].width = max_prof_k + 10

            for i, res in enumerate(res_set):
                for l, r in enumerate(res):
                    sheet.cell(row=i + row, column=1 + l).value = r
                    sheet.cell(row=i + row, column=1 + l).alignment = Alignment(horizontal="center", vertical="center")
                    sheet.cell(row=i + row, column=1 + l).border = self._border

            self._path()
            wb.save(self._dir + self._filename)

        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'download', 'text': self._dir + self._filename,'name':'Отчет о хирур.работе отделений'})
    def _annual_pr_3(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_4(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_5(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_6(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_7(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_8(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_9(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_a(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_b(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_v(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_g(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})
    def _annual_pr_d(self):
        async_to_sync(get_channel_layer().group_send)(self._user_group_name,
                                                      {'type': 'report_data', 'text': 'Отчет cфромирован'})

class AnnualReportABC(ABC):
    def __init__(self,user,request):
        self.user = MyUser.objects.get(user=user)
        self.user_group_name = 'hospital_annual_reports_%s' % user
        self.request = request
        self.date_1 = datetime.strptime(self.request.get('date_1'),'%Y-%m-%d').date()
        self.date_2 = datetime.strptime(self.request.get('date_2'), '%Y-%m-%d').date()
        self.path_shoblons = 'shoblons/hospital/forms'
        self.path_shoblons_oth = 'shoblons/hospital/oth'
        self.temp_dir = 'temp'
        self.border = Border(left=Side(border_style='thin', color='000000'),
                              right=Side(border_style='thin', color='000000'),
                              top=Side(border_style='thin', color='000000'),
                              bottom=Side(border_style='thin', color='000000'))
    def create(self):
        pass
    def is_file(self,file,forms=False):
        if forms:
            file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons, file])
        else:
            file_shoblon = '/'.join([settings.MEDIA_ROOT, self.path_shoblons_oth, file])
        if os.path.exists(file_shoblon):
            filename = self.request.get('filename', None)
            type_fun = self.request.get('type_report')
            if forms:
                filename = f'{filename}.docx' if filename != None and len(filename) > 0 else f'{type_fun}_{self.user.user.id}.docx'
            else:
                filename = f'{filename}.xlsx' if filename != None and len(filename) > 0 else f'{type_fun}_{self.user.user.id}.xlsx'
            file_new = filename
            path = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self.user.user.id),''])
            if not os.path.isdir(path):
                os.mkdir(path)

            file_new = '/'.join([settings.MEDIA_ROOT, self.temp_dir, str(self.user.user.id), file_new])
            shutil.copy2(file_shoblon, file_new)
            return file_new
        return None
    def path(self):
        if not os.path.isdir(settings.MEDIA_ROOT + '/temp/' + f'{str(self.user.user.id)}/'):
            os.mkdir(settings.MEDIA_ROOT + '/temp/' + f'{str(self.user.user.id)}/')
        return settings.MEDIA_ROOT + '/temp/' + f'{str(self.user.user.id)}/'
    
    def filters_list(self):
        filters = json.loads(self.request.get('filters',None))
        filter_list = []
        if filters:
            keys = filters['filter'].keys()
            for k in keys:
                if k == 'datv':
                    filter_list.append({'filter': 'Период выбытия', 'value': filters['filter'].get(k)})
                elif k == 'datp':
                    filter_list.append({'filter': 'Период поступления', 'value': filters['filter'].get(k)})
                elif k == 'otd':
                    filter_list.append({'filter':'Отделение','value':filters['filter'].get(k)[k]})
                elif k == 'prof':
                    filter_list.append({'filter':'Профиль койки','value':filters['filter'].get(k)[k]})
                elif k == 'fam':
                    filter_list.append({'filter':'Фимилия (начало фамилии)','value':filters['filter'].get(k)[k]})
                elif k == 'im':
                    filter_list.append({'filter':'Имя (начало имени)','value':filters['filter'].get(k)[k]})
                elif k == 'ot':
                    filter_list.append({'filter':'Отчество (начало отчества)','value':filters['filter'].get(k)[k]})
                elif k == 'pol':
                    filter_list.append({'filter':'Пол','value':filters['filter'].get(k)[k]})
                elif k == 'type_lgots':
                    filter_list.append({'filter':'Тип льготы','value':filters['filter'].get(k)[k]})
                elif k == 'in_t':
                    filter_list.append({'filter':'Льгота','value':filters['filter'].get(k)[k]})
                elif k == 'r_n':
                    filter_list.append({'filter':'Социальный статус','value':filters['filter'].get(k)[k]})
                elif k == 'age_group':
                    filter_list.append({'filter':'Возрастная группа','value':filters['filter'].get(k)[k]})
                elif k == 'goc':
                    filter_list.append({'filter':'Форма госпитализации','value':filters['filter'].get(k)[k]})
                elif k == 'prpg':
                    filter_list.append({'filter':'Вид госпитализации','value':filters['filter'].get(k)[k]})
                elif k == 'vrez':
                    filter_list.append({'filter':'Давность заболевания','value':filters['filter'].get(k)[k]})
                elif k == 'dskz':
                    filter_list.append({'filter': 'Ds основной ', 'value': filters['filter'].get(k)})
                elif k == 'dsc':
                    filter_list.append({'filter': 'Ds сопутствующий ', 'value': filters['filter'].get(k)})
                elif k == 'dspat':
                    filter_list.append({'filter': 'Ds патологоанатомический ', 'value': filters['filter'].get(k)})
                elif k == 'ds_osl':
                    filter_list.append({'filter': 'Ds осложнения ', 'value': filters['filter'].get(k)})
                elif k == 'ds_onk':
                    filter_list.append({'filter': 'Ds онкология ', 'value': filters['filter'].get(k)})
                elif k == 'ksg_osn':
                    filter_list.append({'filter':'КСГ основного Ds','value':filters['filter'].get(k)[k]})
                elif k == 'c_oksm':
                    filter_list.append({'filter':'Гражданство','value':filters['filter'].get(k)[k]})
                elif k == 'terr':
                    filter_list.append({'filter':'Территория проживания','value':filters['filter'].get(k)[k]})
                elif k == 'reg':
                    filter_list.append({'filter':'Регион (обл.,р-н)','value':filters['filter'].get(k)[k]})
                elif k == 'rai_in':
                    filter_list.append({'filter':'АО г.Тюмень','value':filters['filter'].get(k)[k]})
                elif k == 'cj':
                    filter_list.append({'filter':'Категория проживания','value':filters['filter'].get(k)[k]})
                elif k == 'lpy':
                    filter_list.append({'filter':'Направившее учреждение','value':filters['filter'].get(k)[k]})
                elif k == 'ctkom':
                    filter_list.append({'filter':'Страховая организация','value':filters['filter'].get(k)[k]})
                elif k == 'vds':
                    filter_list.append({'filter':'Источник покрытия затрат','value':filters['filter'].get(k)[k]})
                elif k == 'icx':
                    filter_list.append({'filter':'Исход лечения','value':filters['filter'].get(k)[k]})
                elif k == 'otdel_let':
                    filter_list.append({'filter':'Отд-е летального исхода','value':filters['filter'].get(k)[k]})
                elif k == 'kod_vra':
                    filter_list.append({'filter':'Лечащий врач','value':filters['filter'].get(k)})
                elif k == 'kod_op':
                    filter_list.append({'filter':'Код операции','value':filters['filter'].get(k)[k]})
                elif k == 'pr_osob':
                    filter_list.append({'filter':'Особенность выполнения операции','value':filters['filter'].get(k)[k]})

                elif k == 't_trv':
                    filter_list.append({'filter':'Тип травмы','value':filters['filter'].get(k)[k]})
                elif k == 'trav_ns':
                    filter_list.append({'filter':'Тип телесных повреждений','value':filters['filter'].get(k)[k]})
                elif k == 'disability':
                    filter_list.append({'filter':'Наличие закрытого листа нетрудос','value':filters['filter'].get(k)})
                elif k == 'srber':
                    filter_list.append({'filter':'Срок беременности с ..по..нед','value':filters['filter'].get(k)})
                elif k == 'potd':
                    filter_list.append({'filter':'Внутрибольничный перевод','value':filters['filter'].get(k)[k]})
                elif k == 'kod_y':
                    filter_list.append({'filter':'Перевод в др. ЛПУ','value':filters['filter'].get(k)[k]})
                elif k == 'dskz_prich':
                    filter_list.append({'filter': 'Причина травмы ', 'value': filters['filter'].get(k)})
                elif k == 'pr_per':
                    filter_list.append({'filter': 'Причина перевода в др.ЛПУ', 'value': filters['filter'].get(k)[k]})
                elif k == 'time_minuts_po':
                    filter_list.append({'filter': 'Длительность пребывания в ПО (мин)', 'value': filters['filter'].get(k)[k]})
                elif k == 'stay_in_mo':
                    filter_list.append({'filter': 'Пребывание в МО (к-дн)', 'value': filters['filter'].get(k)[k]})
                elif k == 'metod_hmp':
                    filter_list.append({'filter': 'Вид ВТМП', 'value': filters['filter'].get(k)[k]})
                elif k == 'vid_hmp':
                    filter_list.append({'filter': 'Методы ВТМП', 'value': filters['filter'].get(k)[k]})
                elif k == 'ymer_ymer':
                    filter_list.append({'filter': 'Умершие', 'value': filters['filter'].get(k)[k]})
            for f in range(len(filter_list)):
                if isinstance(filter_list[f]['value'],dict):
                    keys = list(filter_list[f]['value'].keys())
                    if len(keys) == 1:
                        filter_list[f]['value'] = filter_list[f]['value'][keys[0]]
                    elif len(keys) == 2:
                        val1 = filter_list[f]['value'][keys[0]]
                        val2 = filter_list[f]['value'][keys[1]]
                        if len(val1) == 10 and (val1[4] == '-' and val1[7] == '-'):
                            val = val1.split('-')
                            val1 = f'{val[2]}-{val[1]}-{val[0]}'
                            val = val2.split('-')
                            val2 = f'{val[2]}-{val[1]}-{val[0]}'
                            filter_list[f]['value'] = f"С {val1} по {val2}"
                        else:
                            filter_list[f]['value'] = f"С {filter_list[f]['value'][keys[0]]} по {filter_list[f]['value'][keys[1]]}"
            return filter_list
        return None



    



def insert_sheet_AN_13_1_4(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = kwargs['data']


    nzO2_6 = list(Ds.objects.values('kod').filter(kod__range=('O02', 'O06.9')))
    nzO2_6 = [k['kod'] for k in nzO2_6]

    # ds_O04 = []
    nzO04 = list(Ds.objects.values('kod').filter(kod__range=('O04', 'O04.9')))
    nzO04 = [k['kod'] for k in nzO04]

    nzO03 = list(Ds.objects.values('kod').filter(kod__range=('O03', 'O03.9')))
    nzO03 = [k['kod'] for k in nzO03]
    #
    # ds_O02_009 = []
    # nzO02_009 = list(Ds.objects.values('kod').filter(kod__range=('O02', 'O09.9')))
    # nzO02_009 = [k['kod'] for k in nzO02_009]
    #
    #
    # for d in data:
    #     if d.sluchay.dskz and d.sluchay.dskz.kod in nzO04:
    #         ds_O04.append(d)
    #
    #
    # for d in data:
    #     if d.sluchay.dskz and d.sluchay.dskz.kod in nzO02_009:
    #         ds_O02_009.append(d)


    n1 = dict(all_=0,up_to_12=0,from_12_to_22=0,pregnancy_n1=0,pregnancy_n2=0,pregnancy_n3=0)
    n1_1 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n1_2 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n1_3 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n2 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n3 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n3_1 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n3_2_1 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)
    n3_2_2 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0,pregnancy_n3=0)


    n3_2_4 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0, pregnancy_n3=0)
    n3_2_5 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0, pregnancy_n3=0)
    n3_2_6 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0, pregnancy_n3=0)
    n3_2_7 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0, pregnancy_n3=0)
    n3_2_8 = dict(all_=0, up_to_12=0, from_12_to_22=0, pregnancy_n1=0, pregnancy_n2=0, pregnancy_n3=0)


    for v in data:
        if v.vb_a:
            if v.sluchay.dskz.kod in nzO04:
                n1['all_'] += 1
                n1['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
                n1['from_12_to_22'] += 1 if v.vb_a.srber and  12 <= v.vb_a.srber <= 22 else 0
                n1['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
                n1['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
                n1['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0
            if v.sluchay.dskz.kod in nzO04:
                if v.vb_a.pria and v.vb_a.pria.kod == 1:
                    n1_1['all_'] += 1
                    n1_1['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
                    n1_1['from_12_to_22'] += 1 if v.vb_a.srber and 12 <= v.vb_a.srber <= 22 else 0
                    n1_1['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
                    n1_1['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
                    n1_1['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0
            if v.sluchay.dskz.kod in nzO04:
                if v.vb_a.pria and v.vb_a.pria.kod == 2:
                    n1_2['all_'] += 1
                    n1_2['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
                    n1_2['from_12_to_22'] += 1 if v.vb_a.srber and 12 <= v.vb_a.srber <= 22 else 0
                    n1_2['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
                    n1_2['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
                    n1_2['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0
            if v.sluchay.dskz:
                if v.sluchay.dskz.kod in nzO2_6:
                    n3['all_'] +=1
                    n3['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
                    n3['from_12_to_22'] += 1 if v.vb_a.srber and 12 <= v.vb_a.srber <= 22 else 0
                    n3['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
                    n3['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
                    n3['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0

                    if v.vb_a.pria and v.vb_a.pria.kod == 1:
                        n3_1['all_'] += 1
                        n3_1['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
                        n3_1['from_12_to_22'] += 1 if v.vb_a.srber and 12 <= v.vb_a.srber <= 22 else 0
                        n3_1['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
                        n3_1['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
                        n3_1['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0

                    if v.sluchay.dskz.kod in nzO03:
                        n3_2_1['all_'] += 1
                        n3_2_1['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
                        n3_2_1['from_12_to_22'] += 1 if v.vb_a.srber and 12 <= v.vb_a.srber <= 22 else 0
                        n3_2_1['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
                        n3_2_1['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
                        n3_2_1['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0

                    if v.sluchay.dskz.kod in nzO04:
                        n3_2_2['all_'] += 1
                        n3_2_2['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
                        n3_2_2['from_12_to_22'] += 1 if v.vb_a.srber and 12 <= v.vb_a.srber <= 22 else 0
                        n3_2_2['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
                        n3_2_2['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
                        n3_2_2['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0

                        if 18 >= v.patient_year:
                            n3_2_4['all_'] += 1
                            n3_2_4['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
                            n3_2_4['from_12_to_22'] += 1 if v.vb_a.srber and 12 <= v.vb_a.srber <= 22 else 0
                            n3_2_4['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
                            n3_2_4['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
                            n3_2_4['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0
                        if 18 <= v.patient_year <= 29:
                            n3_2_5['all_'] += 1
                            n3_2_5['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
                            n3_2_5['from_12_to_22'] += 1 if v.vb_a.srber and 12 <= v.vb_a.srber <= 22 else 0
                            n3_2_5['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
                            n3_2_5['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
                            n3_2_5['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0
                        if 30 <= v.patient_year <= 39:
                            n3_2_6['all_'] += 1
                            n3_2_6['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
                            n3_2_6['from_12_to_22'] += 1 if v.vb_a.srber and 12 <= v.vb_a.srber <= 22 else 0
                            n3_2_6['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
                            n3_2_6['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
                            n3_2_6['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0
                        if 40 <= v.patient_year <= 49:
                            n3_2_7['all_'] += 1
                            n3_2_7['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
                            n3_2_7['from_12_to_22'] += 1 if v.vb_a.srber and 12 <= v.vb_a.srber <= 22 else 0
                            n3_2_7['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
                            n3_2_7['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
                            n3_2_7['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0
                        if 50 <= v.patient_year :
                            n3_2_8['all_'] += 1
                            n3_2_8['up_to_12'] += 1 if v.vb_a.srber and v.vb_a.srber < 12 else 0
                            n3_2_8['from_12_to_22'] += 1 if v.vb_a.srber and 12 <= v.vb_a.srber <= 22 else 0
                            n3_2_8['pregnancy_n1'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 1 else 0
                            n3_2_8['pregnancy_n2'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber == 2 else 0
                            n3_2_8['pregnancy_n3'] += 1 if v.vb_a.n_ber and v.vb_a.n_ber >= 3 else 0




    row = 9
    sheet.cell(row=row, column=3).value = n1['all_']
    sheet.cell(row=row, column=4).value = n1['up_to_12']
    sheet.cell(row=row, column=5).value = n1['from_12_to_22']
    sheet.cell(row=row, column=6).value = n1['pregnancy_n1']
    sheet.cell(row=row, column=7).value = n1['pregnancy_n2']
    sheet.cell(row=row, column=8).value = n1['pregnancy_n3']
    row = 10
    sheet.cell(row=row, column=3).value = n1_1['all_']
    sheet.cell(row=row, column=4).value = n1_1['up_to_12']
    sheet.cell(row=row, column=5).value = n1_1['from_12_to_22']
    sheet.cell(row=row, column=6).value = n1_1['pregnancy_n1']
    sheet.cell(row=row, column=7).value = n1_1['pregnancy_n2']
    sheet.cell(row=row, column=8).value = n1_1['pregnancy_n3']
    row = 11
    sheet.cell(row=row, column=3).value = n1_2['all_']
    sheet.cell(row=row, column=4).value = n1_2['up_to_12']
    sheet.cell(row=row, column=5).value = n1_2['from_12_to_22']
    sheet.cell(row=row, column=6).value = n1_2['pregnancy_n1']
    sheet.cell(row=row, column=7).value = n1_2['pregnancy_n2']
    sheet.cell(row=row, column=8).value = n1_2['pregnancy_n3']
    row = 14
    sheet.cell(row=row, column=3).value = n3['all_']
    sheet.cell(row=row, column=4).value = n3['up_to_12']
    sheet.cell(row=row, column=5).value = n3['from_12_to_22']
    sheet.cell(row=row, column=6).value = n3['pregnancy_n1']
    sheet.cell(row=row, column=7).value = n3['pregnancy_n2']
    sheet.cell(row=row, column=8).value = n3['pregnancy_n3']
    row = 15
    sheet.cell(row=row, column=3).value = n3_1['all_']
    sheet.cell(row=row, column=4).value = n3_1['up_to_12']
    sheet.cell(row=row, column=5).value = n3_1['from_12_to_22']
    sheet.cell(row=row, column=6).value = n3_1['pregnancy_n1']
    sheet.cell(row=row, column=7).value = n3_1['pregnancy_n2']
    sheet.cell(row=row, column=8).value = n3_1['pregnancy_n3']
    row = 16
    sheet.cell(row=row, column=3).value = n3_2_1['all_']
    sheet.cell(row=row, column=4).value = n3_2_1['up_to_12']
    sheet.cell(row=row, column=5).value = n3_2_1['from_12_to_22']
    sheet.cell(row=row, column=6).value = n3_2_1['pregnancy_n1']
    sheet.cell(row=row, column=7).value = n3_2_1['pregnancy_n2']
    sheet.cell(row=row, column=8).value = n3_2_1['pregnancy_n3']

    row = 17
    sheet.cell(row=row, column=3).value = n3_2_2['all_']
    sheet.cell(row=row, column=4).value = n3_2_2['up_to_12']
    sheet.cell(row=row, column=5).value = n3_2_2['from_12_to_22']
    sheet.cell(row=row, column=6).value = n3_2_2['pregnancy_n1']
    sheet.cell(row=row, column=7).value = n3_2_2['pregnancy_n2']
    sheet.cell(row=row, column=8).value = n3_2_2['pregnancy_n3']

    row = 18
    sheet.cell(row=row, column=3).value = n3_2_4['all_']
    sheet.cell(row=row, column=4).value = n3_2_4['up_to_12']
    sheet.cell(row=row, column=5).value = n3_2_4['from_12_to_22']
    sheet.cell(row=row, column=6).value = n3_2_4['pregnancy_n1']
    sheet.cell(row=row, column=7).value = n3_2_4['pregnancy_n2']
    sheet.cell(row=row, column=8).value = n3_2_4['pregnancy_n3']

    row = 19
    sheet.cell(row=row, column=3).value = n3_2_5['all_']
    sheet.cell(row=row, column=4).value = n3_2_5['up_to_12']
    sheet.cell(row=row, column=5).value = n3_2_5['from_12_to_22']
    sheet.cell(row=row, column=6).value = n3_2_5['pregnancy_n1']
    sheet.cell(row=row, column=7).value = n3_2_5['pregnancy_n2']
    sheet.cell(row=row, column=8).value = n3_2_5['pregnancy_n3']

    row = 20
    sheet.cell(row=row, column=3).value = n3_2_6['all_']
    sheet.cell(row=row, column=4).value = n3_2_6['up_to_12']
    sheet.cell(row=row, column=5).value = n3_2_6['from_12_to_22']
    sheet.cell(row=row, column=6).value = n3_2_6['pregnancy_n1']
    sheet.cell(row=row, column=7).value = n3_2_6['pregnancy_n2']
    sheet.cell(row=row, column=8).value = n3_2_6['pregnancy_n3']

    row = 21
    sheet.cell(row=row, column=3).value = n3_2_7['all_']
    sheet.cell(row=row, column=4).value = n3_2_7['up_to_12']
    sheet.cell(row=row, column=5).value = n3_2_7['from_12_to_22']
    sheet.cell(row=row, column=6).value = n3_2_7['pregnancy_n1']
    sheet.cell(row=row, column=7).value = n3_2_7['pregnancy_n2']
    sheet.cell(row=row, column=8).value = n3_2_7['pregnancy_n3']

    row = 22
    sheet.cell(row=row, column=3).value = n3_2_8['all_']
    sheet.cell(row=row, column=4).value = n3_2_8['up_to_12']
    sheet.cell(row=row, column=5).value = n3_2_8['from_12_to_22']
    sheet.cell(row=row, column=6).value = n3_2_8['pregnancy_n1']
    sheet.cell(row=row, column=7).value = n3_2_8['pregnancy_n2']
    sheet.cell(row=row, column=8).value = n3_2_8['pregnancy_n3']

    sheet.cell(row=3, column=1).value = str(name).capitalize()
    sheet.cell(row=4, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'

def get_rez_a_oth_30_1_1(data,date):
    _=[0,0,0,0,0,0,0,0,0,0,0]
    for d in data:
        year = d.patient_year
        _[0]+=1
        if d.patient.cj:
            kod = d.patient.cj.kod
            if kod == 2:
                _[1]+=1
        if 0 < year <= 17:
            _[2]+=1
        if d.patient.pol and d.patient.pol.id_pol == 1:
            if year >= settings.OLD_M:
                _[3]+=1
        if d.patient.pol and d.patient.pol.id_pol == 2:
            if year >= settings.OLD_G:
                _[3]+=1
        if d.sluchay.datv >= date:
            _[4]+=1
            if d.patient.pol and d.patient.pol.id_pol == 1:
                if year >= settings.OLD_M:
                    _[5] += 1
                    if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                        _[7]+1
            if d.patient.pol and d.patient.pol.id_pol == 2:
                if year >= settings.OLD_G:
                    _[5] += 1
                    if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                        _[7]+1
            if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                _[6]+=1

            if d.le_vr and d.le_vr.kd != None and d.le_vr.kd != '':
                _[8] += d.le_vr.kd
                if d.patient.pol and d.patient.pol.id_pol == 1:
                    if year >= settings.OLD_M:
                        _[9] += 1
                if d.patient.pol and d.patient.pol.id_pol == 2:
                    if year >= settings.OLD_G:
                        _[9] += 1


    return _
def insert_sheet_AN_30_1_1(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    sheet.cell(row=4, column=1).value = str(name).capitalize()
    sheet.cell(row=5, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    data =  rr.get_list_otd_prof(kwargs['data'])

    row=9
    nn=1
    rez_all = []
    for d in data:
        otd = d[0][0]
        otd = otde.objects.filter(naim=otd)[0]
        for prof in d[1]:
            nn+=1
            row+=1
            sheet.cell(row, column=1).value = nn
            sheet.cell(row, column=2).value = prof[0]
            sheet.cell(row, column=3).value = otd.number_beds if otd else None
            sheet.cell(row, column=4).value = otd.number_beds if otd else None
            for t in range(1,5):
                if t !=2:
                    sheet.cell(row, column=t).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row, column=t).border = border
            rez = get_rez_a_oth_30_1_1(prof[1],date_2)
            rez.append(otd.number_beds if otd and otd.number_beds else 0)
            rez_all.append(rez)
            for n,r in enumerate(rez):
                if n != 11:
                    sheet.cell(row, column=5+n).value = r if r !=0 else None
                    sheet.cell(row, column=5 + n).alignment = styles.Alignment(horizontal="center", vertical="center")
                    sheet.cell(row, column=5 + n).border = border
    r = None
    for o in range(len(rez_all)):
        if o == 0:
            r = numpy.array(rez_all[o])
        else:
            r += numpy.array(rez_all[o])
    rez_all = r.tolist()
    nn = rez_all.pop()
    row=9
    sheet.cell(row, column=3).value = nn
    sheet.cell(row, column=4).value = nn
    for n,r in enumerate(rez_all):
        sheet.cell(row, column=5 + n).value = r if r !=0 else None
def get_rez_a_oth_30_2_1(data,ds1,ds2=None,t=1):
    if t == 1:
        rez_all = [0,0,0,0,0,0]
        rez_24 = [0,0,0,0,0,0]
    if t == 2:
        rez_all = [0,0,0,0,0,0,0,0,0,0]
    if t == 3:
        rez_all = [0,0,0,0,0,0,0]

    for d in data:
        if d.sluchay.goc and d.sluchay.goc.tip_name == 'Экстренная':
            ds = d.sluchay.dskz.kod if d.sluchay.dskz else ''
            if t == 1:
                if ds2 is None :
                    if ds in ds1:
                        if d.sluchay.vrez:
                            rez_all[0]+=1
                            if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105,106]:
                                rez_all[1]+=1
                            if d.sluchay.oper.count() == 0:
                                rez_all[2] += 1
                                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                                    rez_all[3] += 1
                            if d.sluchay.oper.count() > 0:
                                rez_all[4] += 1
                                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                                    rez_all[5] += 1
                        if d.sluchay.vrez and d.sluchay.vrez.kod not in [1,2,3,4,5,6,7,8,9,10]:
                            rez_24[0]+=1
                            if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105,106]:
                                rez_24[1]+=1
                            if d.sluchay.oper.count() == 0:
                                rez_24[2] += 1
                                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                                    rez_24[3] += 1
                            if d.sluchay.oper.count() > 0:
                                rez_24[4] += 1
                                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                                    rez_24[5] += 1
                if ds2 is not None:
                    if ds == ds1 or ds == ds2:
                        if d.sluchay.vrez:
                            rez_all[0]+=1
                            if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105,106]:
                                rez_all[1]+=1
                            if d.sluchay.oper.count() == 0:
                                rez_all[2] += 1
                                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                                    rez_all[3] += 1
                            if d.sluchay.oper.count() > 0:
                                rez_all[4] += 1
                                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                                    rez_all[5] += 1
                        if d.sluchay.vrez and d.sluchay.vrez.kod not in [1,2,3,4,5,6,7,8,9,10]:
                            rez_24[0]+=1
                            if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105,106]:
                                rez_24[1]+=1
                            if d.sluchay.oper.count() == 0:
                                rez_24[2] += 1
                                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                                    rez_24[3] += 1
                            if d.sluchay.oper.count() > 0:
                                rez_24[4] += 1
                                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                                    rez_24[5] += 1
            elif t == 2:
                if ds2 is None:
                    if ds in ds1:
                        rez_all[0]+=1
                        if d.sluchay.lpy and d.sluchay.lpy.kod == 97:
                            rez_all[1] += 1
                        if d.sluchay.lpy and d.sluchay.lpy.kod in [7,6]:
                            rez_all[2] += 1
                        if d.sluchay.lpy is None:
                            rez_all[3] += 1
                        if d.sluchay.vrez and d.sluchay.vrez.kod not in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
                            rez_all[4] += 1
                            if d.le_vr and d.le_vr.kd != None and d.le_vr.kd != '':
                                rez_all[9] += d.le_vr.kd
                        if d.sluchay.oper.count() == 0:
                            rez_all[5] += 1
                            if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                                rez_all[6] += 1
                        if d.sluchay.oper.count() > 0:
                            rez_all[7] += 1
                            if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                                rez_all[8] += 1

                if ds2 is not None:
                    if ds == ds1 or ds == ds2:
                        rez_all[0]+=1
                        if d.sluchay.lpy and d.sluchay.lpy.kod == 97:
                            rez_all[1] += 1
                        if d.sluchay.lpy and d.sluchay.lpy.kod in [7,6]:
                            rez_all[2] += 1
                        if d.sluchay.lpy is None:
                            rez_all[3] += 1
                        if d.sluchay.vrez and d.sluchay.vrez.kod not in [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]:
                            rez_all[4] += 1
                            if d.le_vr and d.le_vr.kd != None and d.le_vr.kd != '':
                                rez_all[9] += d.le_vr.kd
                        if d.sluchay.oper.count() == 0:
                            rez_all[5] += 1
                            if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                                rez_all[6] += 1
                        if d.sluchay.oper.count() > 0:
                            rez_all[7] += 1
                            if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                                rez_all[8] += 1
            elif t == 3:
                year = d.patient_year
                if ds2 is None:
                    if ds in ds1:
                        if 3 > year :
                            rez_all[0]+=1
                        if 3 <= year < 7:
                            rez_all[1] += 1
                        if 7 <= year < 15:
                            rez_all[2] += 1
                        if 15 <= year < 40:
                            rez_all[3] += 1
                        if 40 <= year <= 49:
                            rez_all[4] += 1
                        if 50 <= year <= 59:
                            rez_all[5] += 1
                        if year >= settings.OLD_G:
                            rez_all[6] += 1
                if ds2 is not None:
                    if ds == ds1 or ds == ds2:
                        if 3 > year:
                            rez_all[0] += 1
                        if 3 <= year < 7:
                            rez_all[1] += 1
                        if 7 <= year < 15:
                            rez_all[2] += 1
                        if 15 <= year < 40:
                            rez_all[3] += 1
                        if 40 <= year <= 49:
                            rez_all[4] += 1
                        if 50 <= year <= 59:
                            rez_all[5] += 1
                        if year >= settings.OLD_G:
                            rez_all[6] += 1
    if t == 1:
        return [rez_all,rez_24]
    if t == 2 or t == 3:
        return rez_all

def insert_sheet_AN_30_2_1(**kwargs):
    sheet1 = kwargs['sheet'][0]
    sheet2 = kwargs['sheet'][1]
    sheet3 = kwargs['sheet'][2]
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']

    nzK56 = list(Ds.objects.values('kod').filter(kod__range=('K56', 'K56.9')))
    nzK56 = [k['kod'] for k in nzK56]
    nzK35 = list(Ds.objects.values('kod').filter(kod__range=('K35', 'K35.9')))
    nzK35 = [k['kod'] for k in nzK35]
    nzK42_46 = list(Ds.objects.values('kod').filter(kod__range=('K42', 'K46.9')))
    nzK42_46 = [k['kod'] for k in nzK42_46]
    nzK85 = list(Ds.objects.values('kod').filter(kod__range=('K85', 'K85.9')))
    nzK85 = [k['kod'] for k in nzK85]
    nzO00 = list(Ds.objects.values('kod').filter(kod__range=('K85', 'K85.9')))
    nzO00 = [k['kod'] for k in nzO00]
    ##Лист1
    dsK56_rez_all,dsK56_rez_24 = get_rez_a_oth_30_2_1(kwargs['data'],ds1=nzK56)
    dsK35_rez_all,dsK35_rez_24 = get_rez_a_oth_30_2_1(kwargs['data'], ds1=nzK35)
    dsK25_K26_rez_all,dsK25_K26_rez_24 = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K25.0', ds2='K26.0')
    dsK251_K261_rez_all,dsK251_K261_rez_24 = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K25.1', ds2='K26.1')
    dsK252_K262_rez_all,dsK252_K262_rez_24 = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K25.2', ds2='K26.2')
    dsK92_rez_all,dsK92_rez_24 = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K92.2',ds2='K92.2')

    r40_rez_all,r40_rez_24 = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K40.0', ds2='K41.0')
    r42_rez_all,r42_rez_24 = get_rez_a_oth_30_2_1(kwargs['data'], ds1=nzK42_46)

    r40_42_all = [r40_rez_all,r42_rez_all]
    r40_42_24 = [r40_rez_24,r42_rez_24]
    r = None
    for o in range(len(r40_42_all)):
        if o == 0:
            r = numpy.array(r40_42_all[o])
        else:
            r += numpy.array(r40_42_all[o])
    rez_40_42_all = r.tolist()
    r = None
    for o in range(len(r40_42_24)):
        if o == 0:
            r = numpy.array(r40_42_24[o])
        else:
            r += numpy.array(r40_42_24[o])
    rez_40_42_24 = r.tolist()

    dsK80_K81_rez_all,dsK80_K81_rez_24 = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K80.0', ds2='K81.0')

    dsK85_rez_all,dsK85_rez_24 = get_rez_a_oth_30_2_1(kwargs['data'], ds1=nzK85)
    dsO00_rez_all,dsO00_rez_24 = get_rez_a_oth_30_2_1(kwargs['data'], ds1=nzO00)

    _all = [dsK56_rez_all,dsK35_rez_all,dsK25_K26_rez_all,dsK251_K261_rez_all,dsK252_K262_rez_all,
            dsK92_rez_all,rez_40_42_all,dsK80_K81_rez_all,dsK85_rez_all,dsO00_rez_all]
    _24 = [dsK56_rez_24,dsK35_rez_24,dsK25_K26_rez_24,dsK251_K261_rez_24,dsK252_K262_rez_24,
           dsK92_rez_24,rez_40_42_24,dsK80_K81_rez_24,dsK85_rez_24,dsO00_rez_24]
    r = None
    for o in range(len(_all)):
        if o == 0:
            r = numpy.array(_all[o])
        else:
            r += numpy.array(_all[o])
    rez_all = r.tolist()
    r = None
    for o in range(len(_24)):
        if o == 0:
            r = numpy.array(_24[o])
        else:
            r += numpy.array(_24[o])
    rez_all_24 = r.tolist()

    _all.append(rez_all)
    _24.append(rez_all_24)
    sheet1.cell(row=4, column=1).value = str(name).capitalize()
    sheet1.cell(row=5, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    row = 9
    for n,v in enumerate(_all):
        row+=1
        for t,a in enumerate(v):
            sheet1.cell(row=row, column=5+t).value = a if a != 0 else None
        row+=1
        for t,a in enumerate(_24[n]):
            sheet1.cell(row=row, column=5+t).value = a if a != 0 else None
    ##Лист2
    dsK56_rez_all = get_rez_a_oth_30_2_1(kwargs['data'],ds1=nzK56,t=2)
    dsK35_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1=nzK35,t=2)
    dsK25_K26_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K25.0', ds2='K26.0',t=2)
    dsK251_K261_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K25.1', ds2='K26.1',t=2)
    dsK252_K262_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K25.2', ds2='K26.2',t=2)
    dsK92_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K92.2',ds2='K92.2',t=2)

    r40_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K40.0', ds2='K41.0',t=2)
    r42_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1=nzK42_46,t=2)

    r40_42_all = [r40_rez_all,r42_rez_all]
    r = None
    for o in range(len(r40_42_all)):
        if o == 0:
            r = numpy.array(r40_42_all[o])
        else:
            r += numpy.array(r40_42_all[o])
    rez_40_42_all = r.tolist()
    dsK80_K81_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K80.0', ds2='K81.0',t=2)
    dsK85_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1=nzK85,t=2)
    dsO00_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1=nzO00,t=2)

    _all = [dsK56_rez_all,dsK35_rez_all,dsK25_K26_rez_all,dsK251_K261_rez_all,dsK252_K262_rez_all,
            dsK92_rez_all,rez_40_42_all,dsK80_K81_rez_all,dsK85_rez_all,dsO00_rez_all]
    r = None
    for o in range(len(_all)):
        if o == 0:
            r = numpy.array(_all[o])
        else:
            r += numpy.array(_all[o])
    rez_all = r.tolist()
    _all.append(rez_all)
    sheet2.cell(row=4, column=1).value = str(name).capitalize()
    sheet2.cell(row=5, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    row = 8
    for n,v in enumerate(_all):
        row+=1
        for t,a in enumerate(v):
            if t !=9:
                sheet2.cell(row=row, column=4+t).value = a if a != 0 else None
            else:
                try:
                    k = float('{0:.2f}'.format(v[9] /v[0]))
                except ZeroDivisionError:
                    k = 0
                sheet2.cell(row=row, column=4 + t).value = k if k != 0 else None
        row += 1
    ##Лист3
    dsK56_rez_all = get_rez_a_oth_30_2_1(kwargs['data'],ds1=nzK56,t=3)
    dsK35_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1=nzK35,t=3)
    dsK25_K26_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K25.0', ds2='K26.0',t=3)
    dsK251_K261_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K25.1', ds2='K26.1',t=3)
    dsK252_K262_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K25.2', ds2='K26.2',t=3)
    dsK92_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K92.2',ds2='K92.2',t=3)

    r40_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K40.0', ds2='K41.0',t=3)
    r42_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1=nzK42_46,t=3)

    r40_42_all = [r40_rez_all,r42_rez_all]
    r = None
    for o in range(len(r40_42_all)):
        if o == 0:
            r = numpy.array(r40_42_all[o])
        else:
            r += numpy.array(r40_42_all[o])
    rez_40_42_all = r.tolist()
    dsK80_K81_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1='K80.0', ds2='K81.0',t=3)
    dsK85_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1=nzK85,t=3)
    dsO00_rez_all = get_rez_a_oth_30_2_1(kwargs['data'], ds1=nzO00,t=3)

    _all = [dsK56_rez_all,dsK35_rez_all,dsK25_K26_rez_all,dsK251_K261_rez_all,dsK252_K262_rez_all,
            dsK92_rez_all,rez_40_42_all,dsK80_K81_rez_all,dsK85_rez_all,dsO00_rez_all]
    r = None
    for o in range(len(_all)):
        if o == 0:
            r = numpy.array(_all[o])
        else:
            r += numpy.array(_all[o])
    rez_all = r.tolist()
    _all.append(rez_all)
    sheet3.cell(row=4, column=1).value = str(name).capitalize()
    sheet3.cell(row=5, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    row = 6
    for n,v in enumerate(_all):
        row+=1
        for t,a in enumerate(v):
            sheet3.cell(row=row, column=4+t).value = a if a!=0 else None
        row += 1
def insert_sheet_APR_1(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = rr.get_list_otd_prof(kwargs['data'])
    sheet.cell(row=4,column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()
    row = 8
    font = styles.Font(size=14, name='Arial')
    #Всего пациентов
    sl_all = []
    for d in data:
        for b in range(1,18):
            sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        row+=1
        sheet.row_dimensions[row].height = 45
        sheet.cell(row=row, column=1).value = d[0][0]
        sheet.cell(row=row, column=1).font = font
        sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        #Список итог всех пациентов в отделении
        otd_prof_sl_all = []
        for prof in d[1]:
            sheet.row_dimensions[row].height = 45
            sheet.cell(row=row, column=2).value = prof[0]
            sheet.cell(row=row, column=2).font = font
            sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
            rez = get_rez_apr_1(prof[1])
            otd_prof_sl_all.append(rez)
            sl_all.append(rez)
            for v in range(len(rez)):
                try:
                    if v != 15:
                        sheet.cell(row=row, column=3+v).value = rez[v] if rez[v] != 0 else None
                except:
                    print(rez)
                sheet.cell(row=row, column=3+v).font = font
                sheet.cell(row=row, column=3+v).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
            else:
                for b in range(2,18):
                    sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))

            if len(d[1]) > 1:
                row+=1
        else:
            if len(d[1]) > 1:
                sheet.row_dimensions[row].height = 20
                sheet.cell(row=row, column=1).value = 'итого по отделению'
                sheet.cell(row=row, column=1).font = font
                r = None
                for o in range(len(otd_prof_sl_all)):
                    if o == 0:
                        r = numpy.array(otd_prof_sl_all[o])
                    else:
                        r+= numpy.array(otd_prof_sl_all[o])
                rez = r.tolist()
                rez = get_rez_apr_1(rez,False)
                for r in range(len(rez)):
                    if r != 15:
                        sheet.cell(row=row, column=3+r).value = rez[r] if rez[r] != 0 else None
                        sheet.cell(row=row, column=3+r).font = font
                        sheet.cell(row=row, column=3+r).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
    else:
        row+=1
        sheet.row_dimensions[row].height = 20
        for b in range(1,18):
            sheet.cell(row=row, column=b).border = styles.Border(top=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row, column=1).value = 'ВСЕГО ПО СТАЦИОНАРУ'
        sheet.cell(row=row, column=1).font = font
        r = None
        for o in range(len(sl_all)):
            if o == 0:
                r = numpy.array(sl_all[o])
            else:
                r+= numpy.array(sl_all[o])
        rez = r.tolist()
        rez = get_rez_apr_1(rez,False)
        for r in range(len(rez)):
            if r != 15:
                sheet.cell(row=row, column=3+r).value = rez[r] if rez[r] != 0 else None
                sheet.cell(row=row, column=3+r).font = font
                sheet.cell(row=row, column=3+r).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)

def insert_sheet_APR_2(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = rr.get_list_otd_prof(kwargs['data'])
    sheet.cell(row=4,column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = str(name).capitalize()
    row = 7
    font = styles.Font(size=12, name='Arial')
    #Всего пациентов
    sl_all = []
    count_all_otd = 0
    for d in data:
        for b in range(1,15):
            sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))
        row+=1
        count_otd = 0
        for prof in d[1]:
            count_otd += len(prof[1])
        if d[0][0] not in not_oper_count_otdl:
            count_all_otd +=count_otd

        sheet.row_dimensions[row].height = 45
        sheet.cell(row=row, column=1).value = d[0][0]
        sheet.cell(row=row, column=1).font = font
        sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
        #Список итог всех пациентов в отделении
        otd_prof_sl_all = []
        for prof in d[1]:
            sheet.row_dimensions[row].height = 45
            sheet.cell(row=row, column=2).value = prof[0]
            sheet.cell(row=row, column=2).font = font
            sheet.cell(row=row, column=2).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
            rez = get_rez_apr_2(prof[1],count=count_otd)
            otd_prof_sl_all.append(rez)
            sl_all.append(rez)
            for v in range(len(rez)):
                if v != 12:
                    sheet.cell(row=row, column=3+v).value = rez[v] if rez[v] != 0 else None
                    sheet.cell(row=row, column=3+v).font = font
                    sheet.cell(row=row, column=3+v).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
            else:
                for b in range(2,15):
                    sheet.cell(row=row, column=b).border = styles.Border(bottom=styles.Side(border_style='thin', color='000000'))

            if len(d[1]) > 1:
                row+=1
        else:
            if len(d[1]) > 1:
                sheet.row_dimensions[row].height = 20
                sheet.cell(row=row, column=1).value = 'итого по отделению'
                sheet.cell(row=row, column=1).font = font
                r = None
                for o in range(len(otd_prof_sl_all)):
                    if o == 0:
                        r = numpy.array(otd_prof_sl_all[o])
                    else:
                        r+= numpy.array(otd_prof_sl_all[o])
                rez = r.tolist()
                rez = get_rez_apr_2(rez,False,count=count_otd)
                for r in range(len(rez)):
                    if r != 12:
                        sheet.cell(row=row, column=3+r).value = rez[r] if rez[r] != 0 else None
                        sheet.cell(row=row, column=3+r).font = font
                        sheet.cell(row=row, column=3+r).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
    else:
        row+=1
        sheet.row_dimensions[row].height = 20
        for b in range(1,15):
            sheet.cell(row=row, column=b).border = styles.Border(top=styles.Side(border_style='thin', color='000000'))
        sheet.cell(row=row, column=1).value = 'ВСЕГО ПО СТАЦИОНАРУ'
        sheet.cell(row=row, column=1).font = font
        r = None
        for o in range(len(sl_all)):
            if o == 0:
                r = numpy.array(sl_all[o])
            else:
                r+= numpy.array(sl_all[o])
        rez = r.tolist()

        rez = get_rez_apr_2(rez,False,count_all_otd)
        for r in range(len(rez)):
            if r != 12:
                sheet.cell(row=row, column=3+r).value = rez[r] if rez[r] != 0 else None
                sheet.cell(row=row, column=3+r).font = font
                sheet.cell(row=row, column=3+r).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
        row+=1
        sheet.merge_cells(f"A{sheet.cell(row=row, column=1).row}:N{sheet.cell(row=row, column=1).row}")
        sheet.cell(row=row, column=1).value = f'Всего предоперац.к\дн пл.б-х: {int(rez[11])}'
        sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center",wrap_text=True)
        sheet.cell(row=row, column=1).font = font
        sheet.row_dimensions[row].height = 30

def insert_sheet_VRA(**kwargs):
    sheet = kwargs['sheet'][0]
    sheet1 = kwargs['sheet'][1]
    sheet2 = kwargs['sheet'][2]
    sheet3 = kwargs['sheet'][3]

    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    vra = kwargs['vra']
    font = styles.Font(size=12, name='Arial')
     ### Операции
    sheet.cell(row=4,column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet.cell(row=5, column=1).value = f'Хирург - {vra.naim} {vra.ini} ({vra.kod})'
    oper_set = set()
    data = []
    for d in kwargs['data']:
        opers = rr.get_opers(d)
        if opers:
            for o in opers:
                if o.kodx and o.kodx.kod == vra.kod:
                    if o.kod_op:
                        oper_set.add(o.kod_op.kod)
    for o in oper_set:
        data.append([[o],[0,0,0,0,0,0]])

    for op in range(len(data)):
        for d in kwargs['data']:
            opers = rr.get_opers(d)
            if opers:
                for o in opers:
                    if o.kodx and o.kodx.kod == vra.kod:
                        if o.kod_op and o.kod_op.kod == data[op][0][0]:
                            data[op][1][0] += 1
                            if d.sluchay.goc and d.sluchay.goc.id_tip == 1:
                                data[op][1][2] += 1
                            if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105,106]:
                                data[op][1][3] += 1
                            if o.kodan:
                                data[op][1][4] += 1
                            # if o.oslo.count()>0:
                            #     data[op][1][5] += 1
    count = sum([d[1][0] for d in data])
    row = 8
    all_temp = [x[1] for x in data]
    all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(6)]
    all_temp[1] = 100

    for d in data:
        row+=1
        v001 = V001.objects.filter(kod=d[0][0])
        if v001.count() > 0:
            sheet.row_dimensions[row].height = 35
            sheet.cell(row=row, column=1).value = f'{v001[0].kod } - {v001[0].naim}'
            sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)
            sheet.cell(row=row, column=1).font = font
            sheet.cell(row=row, column=1).border = border
            for v in range(len(d[1])):
                if v == 0:
                    sheet.cell(row=row, column=v+2).value = d[1][v]
                    try:
                        d[1][1] = float('{0:.2f}'.format(d[1][v]*100/count))
                    except ZeroDivisionError:
                        d[1][1] = 0
                else:
                    sheet.cell(row=row, column=v+2).value = d[1][v] if d[1][v] != 0 else None
                sheet.cell(row=row, column=v+2).alignment = styles.Alignment(horizontal="center", vertical="center")
                sheet.cell(row=row, column=v+2).font = font
                sheet.cell(row=row, column=v+2).border = border
    else:
        row+=1
        sheet.row_dimensions[row].height = 35
        sheet.cell(row=row, column=1).value = 'ИТОГО'
        sheet.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet.cell(row=row, column=1).font = font
        sheet.cell(row=row, column=1).border = border
        for a in range(len(all_temp)):
            sheet.cell(row=row, column=a+2).value = all_temp[a] if all_temp[a] != 0 else None
            sheet.cell(row=row, column=a+2).alignment = styles.Alignment(horizontal="center", vertical="center")
            sheet.cell(row=row, column=a+2).font = font
            sheet.cell(row=row, column=a+2).border = border
    ### Ноз.группы
    data = []
    ds_set = set()
    for d in kwargs['data']:
        if d.le_vr.kod and d.le_vr.kod.kod == vra.kod:
            if d.sluchay.dskz:
                ds_set.add(d.sluchay.dskz.kod[:3])

    for o in ds_set:
        data.append([[o],[0,0,0,0,0,0,0,0,0,0,0,0]])

    for ds in range(len(data)):
        for d in kwargs['data']:
            if d.le_vr.kod and d.le_vr.kod.kod == vra.kod:
                if d.sluchay.dskz and d.sluchay.dskz.kod[:3] == data[ds][0][0]:
                    data[ds][1][0] += 1
                    try:
                        data[ds][1][2] += d.le_vr.otd
                    except TypeError:
                        pass
                    if d.sluchay.goc and d.sluchay.goc.id_tip == 1: 
                        data[ds][1][3] += 1
                    if d.sluchay.goc and d.sluchay.goc.id_tip == 3:
                        data[ds][1][5] += 1
                    if d.sluchay.icx and d.sluchay.icx.id_iz == 101:
                        data[ds][1][7] += 1
                    if d.sluchay.icx and d.sluchay.icx.id_iz == 102:
                        data[ds][1][8] += 1
                    if d.sluchay.icx and d.sluchay.icx.id_iz == 103:
                        data[ds][1][9] += 1
                    if d.sluchay.icx and d.sluchay.icx.id_iz == 104:
                        data[ds][1][10] += 1
                    if d.sluchay.icx and d.sluchay.icx.id_iz in [105,106]:
                        data[ds][1][11] += 1
    row = 9
    all_temp = [x[1] for x in data]
    all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(12)]
    sheet1.cell(row=4,column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet1.cell(row=5, column=1).value = f'{vra.naim} {vra.ini} ({vra.kod})'
    for d in data:
        row+=1
        ds = Ds.objects.filter(kod=d[0][0])[0]
        sheet1.row_dimensions[row].height = 35
        sheet1.cell(row=row, column=1).value = f'{ds.kod} {ds.naim}'
        sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="left", vertical="center",wrap_text=True)

        try:
            d[1][1] = float('{0:.2f}'.format(d[1][0]*100/all_temp[0]))
        except ZeroDivisionError:
            pass
        try:
            d[1][2] = float('{0:.2f}'.format(d[1][2]/d[1][0]))
        except ZeroDivisionError:
            pass
        try:
            d[1][4] = float('{0:.2f}'.format(d[1][3]*100/d[1][0]))
        except ZeroDivisionError:
            pass
        try:
            d[1][6] = float('{0:.2f}'.format(d[1][5]*100/d[1][0]))
        except ZeroDivisionError:
            pass

        for v in range(len(d[1])):
            sheet1.cell(row=row, column=2+v).value = d[1][v] if d[1][v] != 0 else None
            sheet1.cell(row=row, column=2+v).alignment = styles.Alignment(horizontal="center", vertical="center")
    else:
        row+=1
        sheet1.row_dimensions[row].height = 20
        sheet1.cell(row=row, column=1).value = 'ИТОГО'
        sheet1.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        all_temp[1] = 100

        try:
            all_temp[2] = float('{0:.2f}'.format(all_temp[2]/all_temp[0]))
        except ZeroDivisionError:
            pass
        try:
            all_temp[4] = float('{0:.2f}'.format(all_temp[3]*100/all_temp[0]))
        except ZeroDivisionError:
            pass
        try:
            all_temp[6] = float('{0:.2f}'.format(all_temp[5]*100/all_temp[0]))
        except ZeroDivisionError:
            pass

        for a in range(len(all_temp)):
            sheet1.cell(row=row, column=2+a).value = all_temp[a] if all_temp[a] != 0 else None
            sheet1.cell(row=row, column=2+a).alignment = styles.Alignment(horizontal="center", vertical="center")
    ###ДС
    ds_set = set()
    temp_data = []
    for d in kwargs['data']:
        if d.sluchay.manpy.count() > 0:
            mans = [Manpy.objects.get(id=m['id']) for m in d.sluchay.manpy.values('id')]
            for m in mans:
                if m.tnvr and m.tnvr.kod == vra.kod:
                    if m.kodmn and m.kodmn.kod == '01013':
                        temp_data.append(d)
                        if d.sluchay.dskz:
                            ds_set.add(d.sluchay.dskz.kod)
    data = []
    for o in ds_set:
        data.append([o,0])

    for ds in range(len(data)):
        for d in kwargs['data']:
            if d.sluchay.manpy.count() > 0:
                mans = [Manpy.objects.get(id=m['id']) for m in d.sluchay.manpy.values('id')]
                for m in mans:
                    if m.tnvr and m.tnvr.kod == vra.kod:
                        if m.kodmn and m.kodmn.kod == '01013':
                            if d.sluchay.dskz and d.sluchay.dskz.kod == data[ds][0]:
                                data[ds][1]+= 1
    row = 6

    sheet2.cell(row=2,column=1).value = f'Отчет за период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    sheet2.cell(row=4, column=1).value = f'{vra.naim} {vra.ini} ({vra.kod})'

    all_count = sum([d[1] for d in data])
    for d in data:
        row+=1
        ds = Ds.objects.filter(kod=d[0])[0]
        sheet2.cell(row=row, column=1).value = f'{ds.kod} {ds.naim}'
        sheet2.cell(row=row, column=2).value = d[1]
        sheet2.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")
    else:
        row+=1
        sheet2.cell(row=row, column=1).value = 'ИТОГО'
        sheet2.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet2.cell(row=row, column=2).value = all_count
        sheet2.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")

    ### man
    manpys = []
    kod_set = set()
    data = []
    row = 4
    for d in kwargs['data']:
        if d.sluchay.manpy.count() > 0:
            mans = [Manpy.objects.get(id=m['id']) for m in d.sluchay.manpy.values('id')]
            for m in mans:
                if m.tnvr and m.tnvr.kod == vra.kod:
                    manpys.append(m)
                    if m.kodmn:
                        kod_set.add(m.kodmn.kod)
    for o in kod_set:
        data.append([o,0])
    for d in range(len(data)):
        for m in manpys:
            if m.kodmn and data[d][0] == m.kodmn.kod:
                data[d][1]+= 1
    all_count = sum([d[1] for d in data])
    for d in data:
        row+=1
        am = Ab_Obsh.objects.filter(kod=d[0])[0]
        sheet3.cell(row=row, column=1).value = f'{am.kod} {am.ima}'
        sheet3.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")    
        sheet3.cell(row=row, column=2).value = d[1]
        sheet3.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")    
    else:
        row+=1
        sheet3.cell(row=row, column=1).value = 'ИТОГО'
        sheet3.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet3.cell(row=row, column=2).value = all_count
        sheet3.cell(row=row, column=1).alignment = styles.Alignment(horizontal="center", vertical="center")
        sheet3.cell(row=row, column=2).alignment = styles.Alignment(horizontal="center", vertical="center")  


def ds_list_filter(data,ds1,ds2):
    ds = Ds.objects.values('kod').filter(kod__range=(ds1,ds2)).all()
    if ds.count() > 0:
        ds = [d['kod'] for d in ds]
    else:
        ds = None
    result = []

    for d in data:
        # if d.sluchay.dskz:
        #     if d.sluchay.dskz.kod in ds:
        #         result.append(d)
        if d.sluchay.dspat:
            if d.sluchay.dspat.kod in ds:
                result.append(d)
        else:
            if d.sluchay.dskz:
                if d.sluchay.dskz.kod in ds:
                    result.append(d)
    return result
def get_rez_N_14_1(data):
    global ymkd
    _ = [0,0,0,0,0,0,0,0,0]
    for d in data:
        if d.sluchay.rslt is not None and d.sluchay.rslt.id_tip not in [105,106]:
            _[0]+=1
            if d.sluchay.goc  and d.sluchay.goc.tip_name == 'Экстренная':
                _[1]+= 1
            if d.sluchay.lpy and d.sluchay.lpy.naim == 'СКОРАЯ ПОМОЩЬ':
                _[2] += 1
            if d.le_vr  and d.le_vr.kd and d.le_vr.kd != '':
                _[3] += int(d.le_vr.kd)
        if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105,106]:
            _[4] +=1
            if d.le_vr and d.le_vr.kd and d.le_vr.kd != '':
                ymkd+= int(d.le_vr.kd)
        # if d.sluchay.dspat:

            # _[5] += 1
            # if d.sluchay.rasx and  d.sluchay.rasx == '1':
            #     _[6] += 1
            # if d.sluchay.wskr and d.sluchay.wskr == '3':
            #     _[7] += 1
            if d.sluchay.wskr and d.sluchay.wskr == '2':
                _[5] += 1
                if d.sluchay.rasx and  d.sluchay.rasx == '1':
                    _[6] += 1
            if d.sluchay.wskr and d.sluchay.wskr == '3':
                _[7] += 1
                if d.sluchay.rasx and d.sluchay.rasx == '1':
                    _[8] += 1
    return _



def insert_sheet_AN_14_1_A(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = kwargs['data']
    sheet.cell(row=4, column=1).value = str(name).capitalize()
    sheet.cell(row=5, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'

    result_all_sl = []
    # result = get_rez_N_14_1(ds_list_filter(data,'A00','T98'))
    # for r in range(len(result)):
    #     sheet.cell(row=13, column=4+r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'A00', 'B99.9'))
    for r in range(len(result)):
        sheet.cell(row=14, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'A00', 'A09.9'))
    for r in range(len(result)):
        sheet.cell(row=15, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'A15', 'A16.9'))
    for r in range(len(result)):
        sheet.cell(row=16, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'A39', 'A39.9'))
    for r in range(len(result)):
        sheet.cell(row=17, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'A40', 'A41.9'))
    for r in range(len(result)):
        sheet.cell(row=18, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'A50', 'A64.9'))
    for r in range(len(result)):
        sheet.cell(row=19, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'A80', 'A80.9'))
    for r in range(len(result)):
        sheet.cell(row=20, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'B15', 'B19.9'))
    for r in range(len(result)):
        sheet.cell(row=21, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'B20', 'B24.9'))
    for r in range(len(result)):
        sheet.cell(row=22, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C00', 'D48.9'))
    for r in range(len(result)):
        sheet.cell(row=23, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'C00', 'C97.9'))
    for r in range(len(result)):
        sheet.cell(row=24, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C81', 'C96.9'))
    for r in range(len(result)):
        sheet.cell(row=25, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C82', 'C82.9'))
    for r in range(len(result)):
        sheet.cell(row=26, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C83.0', 'C83.0'))
    for r in range(len(result)):
        sheet.cell(row=27, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C83.1', 'C83.1'))
    for r in range(len(result)):
        sheet.cell(row=28, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C83.3', 'C83.3'))
    for r in range(len(result)):
        sheet.cell(row=29, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C83.8', 'C83.8'))
    for r in range(len(result)):
        sheet.cell(row=30, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C83.9', 'C83.9'))
    for r in range(len(result)):
        sheet.cell(row=31, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C84', 'C84.9'))
    for r in range(len(result)):
        sheet.cell(row=32, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C84.5', 'C84.5'))
    for r in range(len(result)):
        sheet.cell(row=33, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C85', 'C85.9'))
    for r in range(len(result)):
        sheet.cell(row=34, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C88.0', 'C88.0'))
    for r in range(len(result)):
        sheet.cell(row=35, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C91.1', 'C91.1'))
    for r in range(len(result)):
        sheet.cell(row=36, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C92.1', 'C92.1'))
    for r in range(len(result)):
        sheet.cell(row=37, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'C69.2', 'C69.2'))
    for r in range(len(result)):
        sheet.cell(row=38, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'D10', 'D36.9'))
    for r in range(len(result)):
        sheet.cell(row=39, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'D25', 'D25.9'))
    for r in range(len(result)):
        sheet.cell(row=40, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'D27', 'D27.9'))
    for r in range(len(result)):
        sheet.cell(row=41, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'D50', 'D89.9'))
    for r in range(len(result)):
        sheet.cell(row=42, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)
    result = get_rez_N_14_1(ds_list_filter(data, 'D50', 'D64.9'))
    for r in range(len(result)):
        sheet.cell(row=43, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'D60', 'D61.9'))
    for r in range(len(result)):
        sheet.cell(row=44, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'D65', 'D69.9'))
    for r in range(len(result)):
        sheet.cell(row=45, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'D66', 'D68.9'))
    for r in range(len(result)):
        sheet.cell(row=46, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'D80', 'D89.9'))
    for r in range(len(result)):
        sheet.cell(row=47, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E00', 'E89.9'))
    for r in range(len(result)):
        sheet.cell(row=48, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'E01', 'E03.9'))
    for r in range(len(result)):
        sheet.cell(row=49, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E05', 'E05.9'))
    for r in range(len(result)):
        sheet.cell(row=50, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E06', 'E06.9'))
    for r in range(len(result)):
        sheet.cell(row=51, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E10', 'E14.9'))
    for r in range(len(result)):
        sheet.cell(row=52, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E10', 'E10.9'))
    for r in range(len(result)):
        sheet.cell(row=53, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E11', 'E11.9'))
    for r in range(len(result)):
        sheet.cell(row=54, column=4 + r).value = result[r] if result[r] != 0 else None

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'E10.2', 'E10.2')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'E11.2', 'E11.2')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'E12.2', 'E12.2')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'E13.2', 'E13.2')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'E14.2', 'E14.2')))
    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=55, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'E10.3', 'E10.3')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'E11.3', 'E11.3')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'E12.3', 'E12.3')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'E13.3', 'E13.3')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'E14.3', 'E14.3')))
    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=56, column=4 + r).value = rez[r] if rez[r] != 0 else None


    result = get_rez_N_14_1(ds_list_filter(data, 'E22', 'E22.9'))
    for r in range(len(result)):
        sheet.cell(row=57, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E23.0', 'E23.0'))
    for r in range(len(result)):
        sheet.cell(row=58, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E23.2', 'E23.2'))
    for r in range(len(result)):
        sheet.cell(row=59, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E25', 'E25.9'))
    for r in range(len(result)):
        sheet.cell(row=60, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E28', 'E28.9'))
    for r in range(len(result)):
        sheet.cell(row=61, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E29', 'E29.9'))
    for r in range(len(result)):
        sheet.cell(row=62, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E66', 'E66.9'))
    for r in range(len(result)):
        sheet.cell(row=63, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E70.0', 'E70.1'))
    for r in range(len(result)):
        sheet.cell(row=64, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E74.2', 'E74.2'))
    for r in range(len(result)):
        sheet.cell(row=65, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E75.2', 'E75.2'))
    for r in range(len(result)):
        sheet.cell(row=66, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E76.0', 'E76.3'))
    for r in range(len(result)):
        sheet.cell(row=67, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'E84', 'E84.9'))
    for r in range(len(result)):
        sheet.cell(row=68, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'F01', 'F99.9'))
    for r in range(len(result)):
        sheet.cell(row=69, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'F10', 'F19.9'))
    for r in range(len(result)):
        sheet.cell(row=70, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G00', 'G98.9'))
    for r in range(len(result)):
        sheet.cell(row=71, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'G00', 'G09.9'))
    for r in range(len(result)):
        sheet.cell(row=72, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G00', 'G00.9'))
    for r in range(len(result)):
        sheet.cell(row=73, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G04', 'G04.9'))
    for r in range(len(result)):
        sheet.cell(row=74, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G10', 'G12.9'))
    for r in range(len(result)):
        sheet.cell(row=75, column=4 + r).value = result[r] if result[r] != 0 else None

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'G20', 'G20.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'G21', 'G21.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'G23', 'G25.9')))
    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=76, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G20', 'G20.9'))
    for r in range(len(result)):
        sheet.cell(row=77, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G25', 'G25.9'))
    for r in range(len(result)):
        sheet.cell(row=78, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G30', 'G31.9'))
    for r in range(len(result)):
        sheet.cell(row=79, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G30', 'G30.9'))
    for r in range(len(result)):
        sheet.cell(row=80, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G35', 'G37.9'))
    for r in range(len(result)):
        sheet.cell(row=81, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G35', 'G35.9'))
    for r in range(len(result)):
        sheet.cell(row=82, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G40', 'G47.9'))
    for r in range(len(result)):
        sheet.cell(row=83, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G40', 'G41.9'))
    for r in range(len(result)):
        sheet.cell(row=84, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G45', 'G45.9'))
    for r in range(len(result)):
        sheet.cell(row=85, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G50', 'G64.9'))
    for r in range(len(result)):
        sheet.cell(row=86, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G61.0', 'G61.0'))
    for r in range(len(result)):
        sheet.cell(row=87, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G70', 'G73.9'))
    for r in range(len(result)):
        sheet.cell(row=88, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G70.0', 'G70.2'))
    for r in range(len(result)):
        sheet.cell(row=89, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G71.0', 'G71.0'))
    for r in range(len(result)):
        sheet.cell(row=90, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G80', 'G83.9'))
    for r in range(len(result)):
        sheet.cell(row=91, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G80', 'G80.9'))
    for r in range(len(result)):
        sheet.cell(row=92, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G90', 'G90.9'))
    for r in range(len(result)):
        sheet.cell(row=93, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'G95.1', 'G95.1'))
    for r in range(len(result)):
        sheet.cell(row=94, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H00', 'H59.9'))
    for r in range(len(result)):
        sheet.cell(row=95, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'H16.0', 'H16.0'))
    for r in range(len(result)):
        sheet.cell(row=96, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H25', 'H26.9'))
    for r in range(len(result)):
        sheet.cell(row=97, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H30', 'H30.9'))
    for r in range(len(result)):
        sheet.cell(row=98, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H33.0', 'H33.0'))
    for r in range(len(result)):
        sheet.cell(row=99, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H35.3', 'H35.3'))
    for r in range(len(result)):
        sheet.cell(row=100, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H40', 'H40.9'))
    for r in range(len(result)):
        sheet.cell(row=101, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H44.2', 'H44.2'))
    for r in range(len(result)):
        sheet.cell(row=102, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H46', 'H48.9'))
    for r in range(len(result)):
        sheet.cell(row=103, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H47.2', 'H47.2'))
    for r in range(len(result)):
        sheet.cell(row=104, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H54', 'H54.9'))
    for r in range(len(result)):
        sheet.cell(row=105, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H54.0', 'H54.0'))
    for r in range(len(result)):
        sheet.cell(row=106, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H60', 'H95.9'))
    for r in range(len(result)):
        sheet.cell(row=107, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)
    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'H65', 'H66.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'H68', 'H74.9')))

    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=108, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'H65.0', 'H65.1')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'H66.0', 'H66.0')))

    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=109, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'H65.2', 'H65.4')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'H66.1', 'H66.3')))

    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=110, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H68', 'H69.9'))
    for r in range(len(result)):
        sheet.cell(row=111, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H72', 'H72.9'))
    for r in range(len(result)):
        sheet.cell(row=112, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H74', 'H74.9'))
    for r in range(len(result)):
        sheet.cell(row=113, column=4 + r).value = result[r] if result[r] != 0 else None

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'H80', 'H80.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'H81', 'H81.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'H83', 'H83.9')))
    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=114, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H80', 'H80.9'))
    for r in range(len(result)):
        sheet.cell(row=115, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H81.0', 'H81.0'))
    for r in range(len(result)):
        sheet.cell(row=116, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H90', 'H90.9'))
    for r in range(len(result)):
        sheet.cell(row=117, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H90.0', 'H90.0'))
    for r in range(len(result)):
        sheet.cell(row=118, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'H90.3', 'H90.3'))
    for r in range(len(result)):
        sheet.cell(row=119, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I00', 'I99.9'))
    for r in range(len(result)):
        sheet.cell(row=120, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'I00', 'I02.9'))
    for r in range(len(result)):
        sheet.cell(row=121, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I05', 'I09.9'))
    for r in range(len(result)):
        sheet.cell(row=122, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I05', 'I08.9'))
    for r in range(len(result)):
        sheet.cell(row=123, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I10', 'I13.9'))
    for r in range(len(result)):
        sheet.cell(row=124, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I10', 'I10.9'))
    for r in range(len(result)):
        sheet.cell(row=125, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I11', 'I11.9'))
    for r in range(len(result)):
        sheet.cell(row=126, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I12', 'I12.9'))
    for r in range(len(result)):
        sheet.cell(row=127, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I13', 'I13.9'))
    for r in range(len(result)):
        sheet.cell(row=128, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I20', 'I25.9'))
    for r in range(len(result)):
        sheet.cell(row=129, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I20', 'I20.9'))
    for r in range(len(result)):
        sheet.cell(row=130, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I20.0', 'I20.0'))
    for r in range(len(result)):
        sheet.cell(row=131, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I21', 'I21.9'))
    for r in range(len(result)):
        sheet.cell(row=132, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I22', 'I22.9'))
    for r in range(len(result)):
        sheet.cell(row=133, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I24', 'I24.9'))
    for r in range(len(result)):
        sheet.cell(row=134, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I25', 'I25.9'))
    for r in range(len(result)):
        sheet.cell(row=135, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I25.8', 'I25.8'))
    for r in range(len(result)):
        sheet.cell(row=136, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I27', 'I27.9'))
    for r in range(len(result)):
        sheet.cell(row=137, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I30', 'I51.9'))
    for r in range(len(result)):
        sheet.cell(row=138, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I30', 'I30.9'))
    for r in range(len(result)):
        sheet.cell(row=139, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I33', 'I33.9'))
    for r in range(len(result)):
        sheet.cell(row=140, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I34', 'I37.9'))
    for r in range(len(result)):
        sheet.cell(row=141, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I40', 'I40.9'))
    for r in range(len(result)):
        sheet.cell(row=142, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I42', 'I42.9'))
    for r in range(len(result)):
        sheet.cell(row=143, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I44.0', 'I44.3'))
    for r in range(len(result)):
        sheet.cell(row=144, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I47.2', 'I47.2'))
    for r in range(len(result)):
        sheet.cell(row=145, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I48', 'I48.9'))
    for r in range(len(result)):
        sheet.cell(row=146, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I49.5', 'I49.5'))
    for r in range(len(result)):
        sheet.cell(row=147, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I60', 'I69.9'))
    for r in range(len(result)):
        sheet.cell(row=148, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I60', 'I60.9'))
    for r in range(len(result)):
        sheet.cell(row=149, column=4 + r).value = result[r]  if result[r] != 0 else None

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'I61', 'I61.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'I62', 'I62.9')))

    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=150, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I63', 'I63.9'))
    for r in range(len(result)):
        sheet.cell(row=151, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I64', 'I64.9'))
    for r in range(len(result)):
        sheet.cell(row=152, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I65', 'I66.9'))
    for r in range(len(result)):
        sheet.cell(row=153, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I67', 'I67.9'))
    for r in range(len(result)):
        sheet.cell(row=154, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I67.2', 'I67.2'))
    for r in range(len(result)):
        sheet.cell(row=155, column=4 + r).value = result[r] if result[r] != 0 else None

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'I70.2', 'I70.2')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'I73.1', 'I73.1')))

    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=156, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I80', 'I89.9'))
    for r in range(len(result)):
        sheet.cell(row=157, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I80', 'I80.9'))
    for r in range(len(result)):
        sheet.cell(row=158, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I81', 'I81.9'))
    for r in range(len(result)):
        sheet.cell(row=159, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'I83', 'I83.9'))
    for r in range(len(result)):
        sheet.cell(row=160, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'J00', 'J98.9'))
    for r in range(len(result)):
        sheet.cell(row=161, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'J00', 'J06.9'))
    for r in range(len(result)):
        sheet.cell(row=162, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'J04', 'J04.9'))
    for r in range(len(result)):
        sheet.cell(row=163, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'J05', 'J05.9'))
    for r in range(len(result)):
        sheet.cell(row=164, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'J09', 'J11.9'))
    for r in range(len(result)):
        sheet.cell(row=165, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'J12', 'J18.9'))
    for r in range(len(result)):
        sheet.cell(row=166, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'J20', 'J22.9'))
    for r in range(len(result)):
        sheet.cell(row=167, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'J30.1', 'J30.1'))
    for r in range(len(result)):
        sheet.cell(row=168, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'J35', 'J36.9'))
    for r in range(len(result)):
        sheet.cell(row=169, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'J40', 'J43.9'))
    for r in range(len(result)):
        sheet.cell(row=170, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'J44', 'J44.9'))
    for r in range(len(result)):
        sheet.cell(row=171, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'J47', 'J47.9'))
    for r in range(len(result)):
        sheet.cell(row=172, column=4 + r).value = result[r] if result[r] != 0 else None

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'J45', 'J45.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'J46', 'J46.9')))

    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=173, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'J84', 'J94.9'))
    for r in range(len(result)):
        sheet.cell(row=174, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K00', 'K92.9'))
    for r in range(len(result)):
        sheet.cell(row=175, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'K25', 'K26.9'))
    for r in range(len(result)):
        sheet.cell(row=176, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K29', 'K29.9'))
    for r in range(len(result)):
        sheet.cell(row=177, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K40', 'K46.9'))
    for r in range(len(result)):
        sheet.cell(row=178, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K50', 'K52.9'))
    for r in range(len(result)):
        sheet.cell(row=179, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K50', 'K50.9'))
    for r in range(len(result)):
        sheet.cell(row=180, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K51', 'K51.9'))
    for r in range(len(result)):
        sheet.cell(row=181, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K55', 'K63.9'))
    for r in range(len(result)):
        sheet.cell(row=182, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K56', 'K56.9'))
    for r in range(len(result)):
        sheet.cell(row=183, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K57', 'K57.9'))
    for r in range(len(result)):
        sheet.cell(row=184, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K58', 'K58.9'))
    for r in range(len(result)):
        sheet.cell(row=185, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K60', 'K60.9'))
    for r in range(len(result)):
        sheet.cell(row=186, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K61', 'K61.9'))
    for r in range(len(result)):
        sheet.cell(row=187, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K64', 'K64.9'))
    for r in range(len(result)):
        sheet.cell(row=188, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K70', 'K76.9'))
    for r in range(len(result)):
        sheet.cell(row=189, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K74', 'K74.9'))
    for r in range(len(result)):
        sheet.cell(row=190, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K80', 'K83.9'))
    for r in range(len(result)):
        sheet.cell(row=191, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K85', 'K86.9'))
    for r in range(len(result)):
        sheet.cell(row=192, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'K85', 'K85.9'))
    for r in range(len(result)):
        sheet.cell(row=193, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'L00', 'L98.9'))
    for r in range(len(result)):
        sheet.cell(row=194, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'L10', 'L10.9'))
    for r in range(len(result)):
        sheet.cell(row=195, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'L12', 'L12.9'))
    for r in range(len(result)):
        sheet.cell(row=196, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'L13.0', 'L13.0'))
    for r in range(len(result)):
        sheet.cell(row=197, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'L40', 'L40.9'))
    for r in range(len(result)):
        sheet.cell(row=198, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'L40.5', 'L40.5'))
    for r in range(len(result)):
        sheet.cell(row=199, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'L93.0', 'L93.0'))
    for r in range(len(result)):
        sheet.cell(row=200, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'L94.0', 'L94.0'))
    for r in range(len(result)):
        sheet.cell(row=201, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M00', 'M99.9'))
    for r in range(len(result)):
        sheet.cell(row=202, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'M00', 'M25.9'))
    for r in range(len(result)):
        sheet.cell(row=203, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M02', 'M02.9'))
    for r in range(len(result)):
        sheet.cell(row=204, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M05', 'M06.9'))
    for r in range(len(result)):
        sheet.cell(row=205, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M08', 'M08.9'))
    for r in range(len(result)):
        sheet.cell(row=206, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M15', 'M19.9'))
    for r in range(len(result)):
        sheet.cell(row=207, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M30', 'M35.9'))
    for r in range(len(result)):
        sheet.cell(row=208, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M32', 'M32.9'))
    for r in range(len(result)):
        sheet.cell(row=209, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M40', 'M43.9'))
    for r in range(len(result)):
        sheet.cell(row=210, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M45', 'M49.9'))
    for r in range(len(result)):
        sheet.cell(row=211, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M45', 'M45.9'))
    for r in range(len(result)):
        sheet.cell(row=212, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M50', 'M54.9'))
    for r in range(len(result)):
        sheet.cell(row=213, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M65', 'M67.9'))
    for r in range(len(result)):
        sheet.cell(row=214, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M80', 'M94.9'))
    for r in range(len(result)):
        sheet.cell(row=215, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'M80', 'M81.9'))
    for r in range(len(result)):
        sheet.cell(row=216, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'N00', 'N99.9'))
    for r in range(len(result)):
        sheet.cell(row=217, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'N00', 'N15.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'N25', 'N28.9')))

    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=218, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'N17', 'N19.9'))
    for r in range(len(result)):
        sheet.cell(row=219, column=4 + r).value = result[r] if result[r] != 0 else None

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'N20', 'N21.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'N23', 'N23.9')))

    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=220, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'N30', 'N32.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'N34', 'N36.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'N39', 'N39.9')))
    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=221, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'N40', 'N42.9'))
    for r in range(len(result)):
        sheet.cell(row=222, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'N60', 'N60.9'))
    for r in range(len(result)):
        sheet.cell(row=223, column=4 + r).value = result[r] if result[r] != 0 else None

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'N70', 'N73.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'N75', 'N76.9')))
    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=224, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'N70', 'N70.9'))
    for r in range(len(result)):
        sheet.cell(row=225, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'N80', 'N80.9'))
    for r in range(len(result)):
        sheet.cell(row=226, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'N86', 'N86.9'))
    for r in range(len(result)):
        sheet.cell(row=227, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'N91', 'N94.9'))
    for r in range(len(result)):
        sheet.cell(row=228, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'N97', 'N97.9'))
    for r in range(len(result)):
        sheet.cell(row=229, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'O00', 'O99.9'))
    for r in range(len(result)):
        sheet.cell(row=230, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'P00', 'P04.9'))
    for r in range(len(result)):
        sheet.cell(row=231, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'Q00', 'Q99.9'))
    for r in range(len(result)):
        sheet.cell(row=232, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'Q00', 'Q07.9'))
    for r in range(len(result)):
        sheet.cell(row=233, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'Q10', 'Q15.9'))
    for r in range(len(result)):
        sheet.cell(row=234, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'Q20', 'Q28.9'))
    for r in range(len(result)):
        sheet.cell(row=235, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'Q38', 'Q45.9'))
    for r in range(len(result)):
        sheet.cell(row=236, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'Q43', 'Q43.9'))
    for r in range(len(result)):
        sheet.cell(row=237, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'Q50', 'Q52.9'))
    for r in range(len(result)):
        sheet.cell(row=238, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'Q56', 'Q56.9'))
    for r in range(len(result)):
        sheet.cell(row=239, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'Q80', 'Q80.9'))
    for r in range(len(result)):
        sheet.cell(row=240, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'Q85.0', 'Q85.0'))
    for r in range(len(result)):
        sheet.cell(row=241, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'Q90', 'Q90.9'))
    for r in range(len(result)):
        sheet.cell(row=242, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'R00', 'R99.9'))
    for r in range(len(result)):
        sheet.cell(row=243, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'S00', 'T98.9'))
    for r in range(len(result)):
        sheet.cell(row=244, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result_all = []
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'S02', 'S02.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'S12', 'S12.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'S22', 'S22.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'S32', 'S32.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'S42', 'S42.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'S52', 'S52.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'S62', 'S62.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'S72', 'S72.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'S82', 'S82.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'S92', 'S92.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'T02', 'T02.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'T08', 'T08.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'T10', 'T10.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'T12', 'T12.9')))
    result_all.append(get_rez_N_14_1(ds_list_filter(data, 'T14.2', 'T14.2')))

    r = None
    for o in range(len(result_all)):
        if o == 0:
            r = numpy.array(result_all[o])
        else:
            r += numpy.array(result_all[o])
    rez = r.tolist()
    for r in range(len(rez)):
        sheet.cell(row=245, column=4 + r).value = rez[r] if rez[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'S02', 'S02.9'))
    for r in range(len(result)):
        sheet.cell(row=246, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'S05', 'S05.9'))
    for r in range(len(result)):
        sheet.cell(row=247, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'S06', 'S06.9'))
    for r in range(len(result)):
        sheet.cell(row=248, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'T20', 'T30.9'))
    for r in range(len(result)):
        sheet.cell(row=249, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'T36', 'T50.9'))
    for r in range(len(result)):
        sheet.cell(row=250, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'T40.0', 'T40.6'))
    for r in range(len(result)):
        sheet.cell(row=251, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'T51', 'T65.9'))
    for r in range(len(result)):
        sheet.cell(row=252, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'T51', 'T51.9'))
    for r in range(len(result)):
        sheet.cell(row=253, column=4 + r).value = result[r] if result[r] != 0 else None

    result = get_rez_N_14_1(ds_list_filter(data, 'U07.1', 'U07.2'))
    for r in range(len(result)):
        sheet.cell(row=254, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    result = get_rez_N_14_1(ds_list_filter(data, 'Z00', 'Z99.9'))
    for r in range(len(result)):
        sheet.cell(row=255, column=4 + r).value = result[r] if result[r] != 0 else None
    result_all_sl.append(result)

    r = None
    for o in range(len(result_all_sl)):
        if o == 0:
            r = numpy.array(result_all_sl[o])
        else:
            r += numpy.array(result_all_sl[o])
    result = r.tolist()

    for r in range(len(result)):
        sheet.cell(row=13, column=4+r).value = result[r] if result[r] != 0 else None
def insert_sheet_AN_14_1_4(**kwargs):
    data = kwargs['data']
    doc = kwargs['doc']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']

    v1 = 0
    v2 = 0
    v3 = 0
    v4 = 0
    v5 = 0
    v6 = 0
    v7 = 0
    v8 = 0
    v9 = 0
    v10 = 0
    v11 = 0

    v12 = 0
    v13 = 0
    v14 = 0
    v15 = 0
    v16 = 0
    v17 = 0
    v18 = 0
    v19 = 0
    v20 = 0
    v21 = 0
    v22 = 0
    v23 = 0
    v24 = 0
    v25 = 0
    v26 = 0
    v27 = 0
    v28 = 0
    v29 = 0
    v30 = 0
    v31 = 0
    v32 = 0
    v33 = 0
    v34 = 0
    v35 = 0
    v36 = 0
    v37 = 0
    v38 = 0
    v39 = 0
    v40 = 0
    v41 = 0
    v42 = 0
    v43 = 0
    v44 = 0
    v45 = 0
    v46 = 0
    v47 = 0
    v48 = 0
    opers_list = set()

    for d in data:
        opers =  rr.get_opers(d)
        year = d.patient_year
        ## v47,48
        if d.sluchay.rslt and d.sluchay.rslt.id_tip == 102:
            v47 += 1
            if d.le_vr.kd != '' and d.le_vr.kd != None:
                v48 += int(d.le_vr.kd)
        ##
        ## v1 - v11
        if d.le_trv is not None:
            if d.le_trv.t_trv and d.le_trv.t_trv.kod == 7:
                v1 += 1
                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                    v2 += 1
        ## v12 - v21
        if d.sluchay.dskz and d.sluchay.dskz.kod[:3] in ['I21','I22','I23']:
            if d.sluchay.vrez:
                if d.sluchay.vrez.kod == 10:
                    v12 += 1
                if d.sluchay.vrez.kod == 7:
                    v13 += 1
                if d.sluchay.vrez.kod in [1,2]:
                    v14 += 1
            if opers is not None:
                opers_list.clear()
                for oper in opers:
                    if oper.kod_op:
                        if oper.kod_op.kod == "A25.30.036.001":
                            v15 += 1
                        if oper.kod_op.kod in ["A16.12.004.009","A16.12.004.012","A16.12.004.013","A16.12.028"]:
                            v16 += 1
                        opers_list.add(oper.kod_op.kod)
                if "A25.30.036.001" in opers_list:
                    for o in ["A16.12.004.009","A16.12.004.012","A16.12.004.013","A16.12.028"]:
                        if o in opers_list:
                            v17 += 1

            if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                if (d.le_vr.kd !="" and d.le_vr.kd != None) and int(d.le_vr.kd) == 1:
                    v18 += 1
                    if year <= settings.OLD_M:
                        v19 += 1
                    if opers is not None:
                        for oper in opers:
                            if oper.kod_op:
                                if oper.kod_op.kod == "A25.30.036.001":
                                    v20 += 1
                                if oper.kod_op.kod in ["A16.12.004.009", "A16.12.004.012", "A16.12.004.013", "A16.12.028"]:
                                    v21 += 1
        ## V22 - V24
        if d.sluchay.dskz and d.sluchay.dskz.kod[:3] in ['I60','I61','I62','I63','I64',
                                                         'I65','I66','I67','I68','I69']:
            if d.sluchay.vrez:
                if d.sluchay.vrez.kod == 10:
                    v22 += 1
                if d.sluchay.vrez.kod == 6:
                    v23 += 1
            if opers is not None:
                for oper in opers:
                    if oper.kod_op:
                        if oper.kod_op.kod == "A25.30.036.001":
                            v24 += 1

        if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
            ## V25 - V26
            if d.vb_a:
                print('++')
            ## V27 - V29
            if 0 <= year <= 17:
                v27 += 1
            if 18 <= year <= settings.OLD_M:
                v28 += 1
            if d.sluchay.dskz and d.sluchay.dskz.kod[:3] in ['I60', 'I61', 'I62', 'I63', 'I64',
                                                             'I65', 'I66', 'I67', 'I68', 'I69']:
                v29 += 1
            ## V30 - V34
            if d.sluchay.p_per:
                if d.sluchay.p_per.kod == 3:
                    v30 += 1
                    if 0 <= year <= 17:
                        v31 += 1
                if d.sluchay.p_per.kod == 9:
                    v32 += 1
                    if 0 <= year <= 17:
                        v33 += 1
                if d.sluchay.p_per.kod == 1:
                    v34 += 1
            if d.le_vr.aro_n != "" and d.le_vr.aro_n != None:
                ## V39 - V43
                if int(d.le_vr.aro_n) <= 24 :
                    v39 += 1
                if int(d.le_vr.aro_n) <= 72:
                    v40 += 1
                if int(d.le_vr.aro_n) >= 720:
                    v41 += 1
                if d.le_vr.aro_let == '1':
                    v42 += 1
                if d.le_vr.aro_let == '2':
                    v43 += 1
            year = d.patient_year
            pol = d.patient.pol.id_pol if d.patient.pol else None
            old = False
            if pol == 1 and year >= settings.OLD_M:
                old = True
            elif pol == 2 and year > settings.OLD_G:
                old = True
            if old:
                if d.sluchay.dskz and d.sluchay.dskz.kod in ['S72.0','S72.1','S72.2']:
                    v44 += 1
                    if opers is not None:
                        v45 += 1
                        if d.sluchay.med_dev.count() > 1:
                            v46 += 1

    context = {
        'n': str(name).capitalize(),
        'd': f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.',
        'v1': v1 if v1 != 0 else '',
        'v2': v1 if v2 != 0 else '',
        'v3': v3 if v3 != 0 else '',
        'v4': v4 if v4 !=0 else '',
        'v5': v5 if v5 !=0 else '',
        'v6': v6 if v6 !=0 else '',
        'v7': v7 if v7 !=0 else '',
        'v8': v8 if v8 !=0 else '',
        'v9': v9 if v9 !=0 else '',
        'v10': v10 if v10 !=0 else '',
        'v11': v11 if v11 !=0 else '',
        'v12': v12 if v12 !=0 else '',
        'v13': v13 if v13 !=0 else '',
        'v14': v14 if v14 !=0 else '',
        'v15': v15 if v15 !=0 else '',
        'v16': v16 if v16 !=0 else '',
        'v17': v17 if v17 !=0 else '',
        'v18': v18 if v18 !=0 else '',
        'v19': v19 if v19 !=0 else '',
        'v20': v20 if v20 !=0 else '',
        'v21': v21 if v21 !=0 else '',
        'v22': v22 if v22 !=0 else '',
        'v23': v23 if v23 !=0 else '',
        'v24': v24 if v24 !=0 else '',
        'v25': v25 if v25 !=0 else '',
        'v26': v26 if v26 !=0 else '',
        'v27': v27 if v27 !=0 else '',
        'v28': v28 if v28 !=0 else '',
        'v29': v29 if v29 !=0 else '',
        'v30': v30 if v30 !=0 else '',
        'v31': v31 if v31 !=0 else '',
        'v32': v32 if v32 !=0 else '',
        'v33': v33 if v33 !=0 else '',
        'v34': v34 if v34 !=0 else '',
        'v35': v35 if v35 !=0 else '',
        'v36': v36 if v36 !=0 else '',
        'v37': v37 if v37 !=0 else '',
        'v38': v38 if v38 !=0 else '',
        'v39': v39 if v39 !=0 else '',
        'v40': v40 if v40 !=0 else '',
        'v41': v41 if v41 !=0 else '',
        'v42': v42 if v42 !=0 else '',
        'v43': v43 if v43 !=0 else '',
        'v44': v44 if v44 !=0 else '',
        'v45': v45 if v45 !=0 else '',
        'v46': v46 if v46 !=0 else '',
        'v47': v47 if v47 != 0 else '',
        'v48': v48 if v48 != 0 else ''
    }
    doc.render(context)

def write_sheet_AN_14_3_1(sheet,data,n,itr=None,typ=None):
    result = get_rez_N_14_3_1(data,iter,typ)

def get_rez_N_14_3_1(data,itr=None,typ=None):
    if typ == 'det':
        year_11_1_d = 0
        year_11_2_d = 14
        year_11_1_e = 0
        year_11_1_f = 15
        year_11_2_f = 17

    _ = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
    #24,#25 не заполнено
    if itr is None:
        year = data.patient_year
        if data.sluchay.oper.count() > 0:
            _[0] += 1
        if year_11_1_d <= year <= year_11_2_d:
            _[1] += 1
        if year == year_11_1_e:
            _[2] += 1
        if year_11_1_f <= year <= year_11_2_f:
            _[3] += 1
        if data.sluchay.tip_oms == '2':
            _[4] += 1
            if year_11_1_d <= year <= year_11_2_d:
                _[5] += 1
            if year == year_11_1_e:
                _[6] += 1
            if year_11_1_f <= year <= year_11_2_f:
                _[7] += 1
        if data.sluchay.oslo.count() > 0:
            _[8] += 1
            if year_11_1_d <= year <= year_11_2_d:
                _[9] += 1
            if year == year_11_1_e:
                _[10] += 1
            if year_11_1_f <= year <= year_11_2_f:
                _[11] += 1
            if data.sluchay.tip_oms == '2':
                _[12] += 1
                if year_11_1_d <= year <= year_11_2_d:
                    _[13] += 1
                if year == year_11_1_e:
                    _[14] += 1
                if year_11_1_f <= year <= year_11_2_f:
                    _[15] += 1
        if data.sluchay.oper.count() > 0:
            if data.sluchay.rslt and data.sluchay.rslt.id_tip in [105, 106]:
                _[16] += 1
                if year_11_1_d <= year <= year_11_2_d:
                    _[17] += 1
                if year == year_11_1_e:
                    _[18] += 1
                if year_11_1_f <= year <= year_11_2_f:
                    _[19] += 1
                if data.sluchay.tip_oms == '2':
                    _[20] += 1
                    if year_11_1_d <= year <= year_11_2_d:
                        _[21] += 1
                    if year == year_11_1_e:
                        _[22] += 1
                    if year_11_1_f <= year <= year_11_2_f:
                        _[23] += 1
    else:
        for d in data:
            year = d.patient_year
            if d.sluchay.oper.count() > 0:
                _[0] += 1
            if year_11_1_d <= year <= year_11_2_d:
                _[1] += 1
            if year == year_11_1_e:
                _[2] += 1
            if year_11_1_f <= year <= year_11_2_f:
                _[3] += 1
            if d.sluchay.tip_oms == '2':
                _[4] += 1
                if year_11_1_d <= year <= year_11_2_d:
                    _[5] += 1
                if year == year_11_1_e:
                    _[6] += 1
                if year_11_1_f <= year <= year_11_2_f:
                    _[7] += 1
            if d.sluchay.oslo.count() > 0:
                _[8] += 1
                if year_11_1_d <= year <= year_11_2_d:
                    _[9] += 1
                if year == year_11_1_e:
                    _[10] += 1
                if year_11_1_f <= year <= year_11_2_f:
                    _[11] += 1
                if d.sluchay.tip_oms == '2':
                    _[12] += 1
                    if year_11_1_d <= year <= year_11_2_d:
                        _[13] += 1
                    if year == year_11_1_e:
                        _[14] += 1
                    if year_11_1_f <= year <= year_11_2_f:
                        _[15] += 1
            if d.sluchay.oper.count() > 0:
                if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                    _[16] += 1
                    if year_11_1_d <= year <= year_11_2_d:
                        _[17] += 1
                    if year == year_11_1_e:
                        _[18] += 1
                    if year_11_1_f <= year <= year_11_2_f:
                        _[19] += 1
                    if d.sluchay.tip_oms == '2':
                        _[20] += 1
                        if year_11_1_d <= year <= year_11_2_d:
                            _[21] += 1
                        if year == year_11_1_e:
                            _[22] += 1
                        if year_11_1_f <= year <= year_11_2_f:
                            _[23] += 1
    return _
def insert_sheet_AN_14_3_1(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = kwargs['data']
    typ = kwargs['typ']

    # for n in range(127):
    #     for k in range(12):
    #         sheet.cell(row=11+n, column=3+k).value = 0
    # all_data_oper = []
    # for d in data:
    #     if d.sluchay.oper.count() > 0:
    #         oper = d.sluchay.oper.filter(pop=True).first()
    #         if oper and oper.kod_op and d.sluchay.dskz:
    #             all_data_oper.append(d)
    #             kod = oper.kod_op.kod
    #             ### 2
    #             if kod == 'A16.12.052':
    #                 write_sheet_AN_14_3_1(sheet,d,12,itr=False,typ=typ)
    #             ###
    #             ### 2.1
    #             if d.sluchay.dskz.kod[1]=='S' or d.sluchay.dskz.kod[1]=='T':
    #                 if kod[:7] in ['059', '016', '004', '013', '053', '005', '084', '007', '008', '010', '017', '037', '052',
    #                                '026', '009', '003', '002', '045', '042', '006', '022', '060', '018', '025',
    #                                '076', '092', '011', '015', '028', '027', '054', '012', '001', '021', '046', '051', '014']:
    #                     pass
    #             ###
    #             ### 2.2
    #             if d.sluchay.dskz.kod[1] == 'S' or d.sluchay.dskz.kod[1] == 'T':
    #                 pass
    #             ###
    # if len(all_data_oper) > 0:
    #     write_sheet_AN_14_3_1(sheet,all_data_oper,11)
    result_all = []
    sheet.cell(row=4, column=1).value = str(name).capitalize()
    sheet.cell(row=5, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    con = 0
    for d in data:
        opers = rr.get_opers(d)
        if opers is not None:
            con+=len(opers)
        row = get_row_god_14(d, row=1)
        if len(row) > 0:
            result = get_rez_N_14_3_1(data=d,itr=None,typ='det')
            for r in row:
                for res in range(len(result)):
                    if sheet.cell(row=r, column=3 + res).value != None:
                        val = int(sheet.cell(row=r, column=3 + res).value) + result[res]
                        sheet.cell(row=r, column=3 + res).value = val if val != 0 else None
                    else:
                        sheet.cell(row=r, column=3 + res).value = result[res] if result[res] != 0 else None
    row = [12,44,46,57,60,64,77,87,96,97,99,100,106,116,130,131,133,135,136,137]
    oper_count = 0
    manip_count = 0
    oslo_count = 0
    oslo_count_vmp = 0

    ymer_count = 0
    ymer_count_vmp = 0

    for r in row:
        oper = int(sheet.cell(row=r, column=3).value) if sheet.cell(row=r, column=3).value != None else 0
        manip = int(sheet.cell(row=r, column=7).value) if sheet.cell(row=r, column=7).value != None else 0
        oslo = int(sheet.cell(row=r, column=11).value) if sheet.cell(row=r, column=11).value != None else 0
        oslovmp = int(sheet.cell(row=r, column=15).value) if sheet.cell(row=r, column=15).value != None else 0

        ymer =  int(sheet.cell(row=r, column=19).value) if sheet.cell(row=r, column=19).value != None else 0
        ymer_vmp = int(sheet.cell(row=r, column=23).value) if sheet.cell(row=r, column=23).value != None else 0

        oper_count+=oper
        manip_count+=manip
        oslo_count+=oslo
        oslo_count_vmp+=oslovmp
        ymer_count+=ymer
        ymer_count_vmp+=ymer_vmp
    # result_all = get_rez_N_14_3_1(data=data, itr=True, typ='det')
    # for res in range(len(result_all)):
    #     sheet.cell(row=11, column=3 + res).value = result_all[res] if result_all[res] != 0 else None
    sheet.cell(row=11, column=3).value = oper_count if oper_count != 0 else None
    sheet.cell(row=11, column=7).value = manip_count if manip_count != 0 else None
    sheet.cell(row=11, column=11).value = oslo_count if oslo_count != 0 else None

    sheet.cell(row=11, column=15).value = oslo_count_vmp if oslo_count_vmp != 0 else None
    sheet.cell(row=11, column=19).value = ymer_count if ymer_count != 0 else None
    sheet.cell(row=11, column=23).value = ymer_count_vmp if ymer_count_vmp != 0 else None


def insert_sheet_AN_14_3_2(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = kwargs['data']
    sheet.cell(row=4, column=1).value = str(name).capitalize()
    sheet.cell(row=5, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    list_row = [12, 44, 46, 57, 60, 64, 77, 87, 96, 97, 99, 100, 106, 116, 130, 131, 133, 135, 136, 137]
    list_result = []
    oper_str =  [[[],[],[],[]] for r in range(137)]
    i = 0
    for row in range(12,138):
        oper_str[i][0].append(row)
        oper_str[i][1].append(sheet.cell(row=row, column=1).value)
        i+=1
    opers_data = rr.get_list_oper(data)
    for oper in opers_data:
        result_row = get_row_god_14(data=oper[0][0],row=1,itr=True)
        result = get_rez_N_14_3_1(data=oper[1], itr=True, typ='det')
        v = V001.objects.get(kod=oper[0][0])
        for row in result_row:
            for r in range(len(oper_str)):
                if row in oper_str[r][0]:
                    oper_str[r][2].append(f'{oper[0][0]}-{v.naim[:10]}')
                    oper_str[r][3].append(result)

    for row in oper_str:
        if len(row[0])>0 and len(row[2]) >0:
            height = 20
            for oper in row[2]:
                height += 12
            sheet.row_dimensions[row[0][0]].height = height
            opers_str = '\n'.join(row[2])
            text = row[1]
            sheet.cell(row=row[0][0], column=1).value = opers_str + '\n' + text[0]
            data_oper = row[3]
            r = None
            for o in range(len(data_oper)):
                if o == 0:
                    r = numpy.array(data_oper[o])
                else:
                    r += numpy.array(data_oper[o])
            rez_data = r.tolist()
            if row[0][0] in list_row:
                list_result.append(rez_data)
            for result in data_oper:
                for res in range(len(result)):
                    r = str(result[res]) if result[res] != 0 else ''
                    val = str(sheet.cell(row=row[0][0], column=3 + res).value)
                    if val != 'None':
                        sheet.cell(row=row[0][0], column=3 + res).value = val+'\n'+r
                    else:
                        sheet.cell(row=row[0][0], column=3 + res).value = r
            else:

                for res in range(len(rez_data)):
                    r = str(rez_data[res]) if rez_data[res] != 0 else ''
                    val = str(sheet.cell(row=row[0][0], column=3 + res).value)
                    if val != 'None':
                        sheet.cell(row=row[0][0], column=3 + res).value = val+'\n'+r
                    else:
                        sheet.cell(row=row[0][0], column=3 + res).value = r

    r = None
    for o in range(len(list_result)):
        if o == 0:
            r = numpy.array(list_result[o])
        else:
            r += numpy.array(list_result[o])
    rez_data = r.tolist()
    for r in range(len(rez_data)):
        sheet.cell(row=11, column=3 + r).value = rez_data[r] if rez_data[r] != 0 else None



def write_sheet_AN_14_3_3_1(data,itr=None):
    _ = [0,0,0,0,0,0]
    if itr == None:
        _[0] = 1
        if data.sluchay.tip_oms == '2':
            _[1] += 1
        if data.sluchay.oslo.count() > 0:
            _[2] += 1
            if data.sluchay.tip_oms == '2':
                _[3] += 1
        if data.sluchay.rslt and data.sluchay.rslt.id_tip in [105, 106]:
            _[4] += 1
            if data.sluchay.tip_oms == '2':
                _[5] += 1
    else:
        for d in data:
            _[0] += d.sluchay.oper.count()
            if d.sluchay.tip_oms == '2':
                _[1] += 1
            if d.sluchay.oslo.count() > 0:
                _[2] += 1
                if d.sluchay.tip_oms == '2':
                    _[3] += 1
            if d.sluchay.rslt and d.sluchay.rslt.id_tip in [105, 106]:
                _[4] += 1
                if d.sluchay.tip_oms == '2':
                    _[5] += 1
    # for i in range(len(_)):
    #     sheet.cell(row=n, column=3+i).value = int(sheet.cell(row=n, column=3+i).value) + int(_[i])
    return _

def write_sheet_AN_14_3_8(data):
    _ = [0, 0, 0]
    # _[0] += data.sluchay.oper.count()
    _[0] += 1
    _[1] += 1
    if data.sluchay.le_vr.kd != None and data.sluchay.le_vr.kd != '':
        _[2] += int(data.sluchay.le_vr.kd)
    return _

def insert_sheet_AN_14_3_3_1(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = kwargs['data']

    # for n in range(127):
    #     for k in range(6):
    #         sheet.cell(row=10+n, column=3+k).value = 0
    all_data_oper = []
    opp = 0
    sheet.cell(row=4, column=1).value = str(name).capitalize()
    sheet.cell(row=5, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    # for d in data:
    #     opers =  rr.get_opers(d)
    #     if d.sluchay.otd and opers and d.sluchay.otd.naim == 'КАРДИОЛОГИЧЕСКОЕ':
    #         opp += len(opers)
    #     if opers:
    #         all_data_oper.append(d)
    #         oper_pop = d.sluchay.oper.filter(pop=True).first()
    #         if oper_pop and oper_pop.kod_op and d.sluchay.dskz:
    #             kod = oper_pop.kod_op.kod
    #             ### 2;2.2.2;2.4.1;2.4.1.1;2.5;2.7;2.8
    #             if kod == 'A16.12.052':
    #                 write_sheet_AN_14_3_3_1(sheet,d,11,itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 31, itr=True)
    #                 continue
    #             if kod == 'A16.12.053':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 11, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 16, itr=True)
    #                 continue
    #             if kod in ['A16.12.008.001','A16.12.008.002']:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 11, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 25, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 26, itr=True)
    #                 continue
    #             if d.sluchay.dskz.kod[1] == 'S' and (kod[:7] in ['035','051']):
    #                 write_sheet_AN_14_3_3_1(sheet, d, 11, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 38, itr=True)
    #                 continue
    #             ###
    #             ### 2.1
    #             if d.sluchay.dskz.kod[1] == 'S' or d.sluchay.dskz.kod[1] == 'T':
    #                 if kod[:7] in ['059', '016', '004', '013', '053', '005', '084', '007', '008', '010', '017', '037',
    #                                '052',
    #                                '026', '009', '003', '002', '045', '042', '006', '022', '060', '018', '025',
    #                                '076', '092', '011', '015', '028', '027', '054', '012', '001', '021', '046', '051',
    #                                '014']:
    #                     write_sheet_AN_14_3_3_1(sheet,d,12,itr=True)
    #                     write_sheet_AN_14_3_3_1(sheet, d, 38, itr=True)
    #                     write_sheet_AN_14_3_3_1(sheet, d, 42, itr=True)
    #                     continue
    #             if   d.sluchay.dskz.kod[1] == 'M' and kod[:7] in ['035','051']:
    #                 write_sheet_AN_14_3_3_1(sheet,d,12,itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 39, itr=True)
    #                 continue
    #             ###
    #             ### 2.2;2.2.1
    #             # if d.sluchay.dskz.kod[1] == 'S' or d.sluchay.dskz.kod[1] == 'T':
    #             #     pass
    #             if '034' <= kod[:7] <= '036':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 13, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 14, itr=True)
    #                 continue
    #             if kod in ['A16.23.017.002','A16.23.017.003','A16.23.017.004','A16.23.017.005','A16.23.017.006']:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 13, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 16, itr=True)
    #                 continue
    #             ###
    #             ### 2.2.1.1
    #             if kod == 'A16.23.036.004':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 15, itr=True)
    #                 continue
    #             ###
    #
    #             ### 2.3;'2.3.1;2.3.1.1;2.3.2;
    #             if 'I60' <= d.sluchay.dskz.kod <= 'I63.9':
    #                 if kod[:7] in ['003','001','007','010','013','023']:
    #                     write_sheet_AN_14_3_3_1(sheet, d, 18, itr=True)
    #                     continue
    #                 if '015' <= kod[:7] <= '017':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 18, itr=True)
    #                     write_sheet_AN_14_3_3_1(sheet, d, 19, itr=True)
    #                     continue
    #                 if kod in ['A16.23.028','A16.23.051','A16.23.026','A16.23.037']:
    #                     write_sheet_AN_14_3_3_1(sheet, d, 18, itr=True)
    #                     write_sheet_AN_14_3_3_1(sheet, d, 19, itr=True)
    #                     continue
    #                 if kod[:7] == '016':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 18, itr=True)
    #                     write_sheet_AN_14_3_3_1(sheet, d, 21, itr=True)
    #                     continue
    #                 if kod[:7] == '017':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 20, itr=True)
    #                     continue
    #             ###
    #             ### 2.5
    #             if kod[:7] in ['023','024','032','033','038']:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 31, itr=True)
    #                 continue
    #             if d.sluchay.dskz.kod[1] == 'C':
    #                 if kod[:7] in ['055','061','062']:
    #                     write_sheet_AN_14_3_3_1(sheet, d, 31, itr=True)
    #                     continue
    #             if d.sluchay.dskz.kod[1] in ['C','D']:
    #                 if kod[:7] >= '067' and kod[:7] <= '069' or kod[:7] == '071' and kod[:7] == '073':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 31, itr=True)
    #                     continue
    #                 if kod[:7] == '075' or kod[:7] == '082' or kod[:7] == '083' or kod[:7] == '089':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 31, itr=True)
    #                     continue
    #                 if  kod[:7] == '001' or kod[:7] == '002':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 31, itr=True)
    #                     continue
    #                 if kod[:7] == '007' or kod[:7] == '024':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 31, itr=True)
    #                     continue
    #                 if kod[:7] == '032' or kod[1:7] == '023.001':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 31, itr=True)
    #                     continue
    #             ###
    #             ### 2.6;2.6.1;2.6.1.1;2.6.2;2.6.2.1;2.7;2.11
    #             if kod[:7] in ['014','030','031','040','041','072','077','019','074']:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 32, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 33, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 35, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 36, itr=True)
    #                 continue
    #             if kod[1:7] == '074.001':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 32, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 33, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 34, itr=True)
    #                 continue
    #             if d.sluchay.dskz.kod[1] == 'G' and '055' <= kod[:7] <= '058':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 32, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 35, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 37, itr=True)
    #                 continue
    #             if d.sluchay.dskz.kod[1] and kod[1:7] in ['039','042','078','079']:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 32, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 35, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 37, itr=True)
    #                 continue
    #             if kod[:7] in ['006','018','020','045','005','010','012','013','017','023',
    #                            '046','048','049','050','059','060','068','088','089',
    #                            '076','028','052'] or '063' <= kod[:7] <= '066' or kod[1:7] in ['042.001','042.002']:
    #                 if d.sluchay.dskz.kod[1] == 'G' or d.sluchay.dskz.kod[1] == 'H':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 32, itr=True)
    #                     write_sheet_AN_14_3_3_1(sheet, d, 38, itr=True)
    #                     write_sheet_AN_14_3_3_1(sheet, d, 42, itr=True)
    #                     continue
    #
    #             if '009' <= kod[:7] <= '013' or kod[:7] in ['005','015']:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 32, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 35, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 36, itr=True)
    #                 continue
    #             if kod[1:7] in ['003.001','017.001']:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 32, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 33, itr=True)
    #                 continue
    #
    #             ###
    #             ### 2.7
    #             if d.sluchay.dskz.kod[1] == 'S' or d.sluchay.dskz.kod[1] == 'T':
    #                 if kod[:7] == '085':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 38, itr=True)
    #                     write_sheet_AN_14_3_3_1(sheet, d, 39, itr=True)
    #                     continue
    #             ###
    #             ### 2.8
    #             if d.sluchay.dskz.kod[1] == 'M' and kod[:3] == 'A22' or \
    #                 kod[:7] in ['029','008','010','032','025','032']:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 39, itr=True)
    #                 continue
    #             ###
    #             ### 2.9
    #             if kod[:7] == '027':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 40, itr=True)
    #                 continue
    #             ###
    #             ### 2.10
    #             if '007' <= kod[:7] <= '011' or kod[:7] in ['043','044','047','053','054']:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 41, itr=True)
    #                 continue
    #             ###
    #             ### 4;
    #             if kod[4:6] == '26':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 45, itr=True)
    #                 continue
    #             ###
    #             ###4.1
    #             if 'H40' <= d.sluchay.dskz.kod[:3] <= 'H42':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 46, itr=True)
    #                 continue
    #             ###
    #             ###9
    #             if 'A16.30.001' <= kod <= 'A16.30.005':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 86, itr=True)
    #                 continue
    #             if 'A16.30.021' <= kod <= 'A16.30.028':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 86, itr=True)
    #                 continue
    #             if kod[4:6] in ['05','14','15','17','18','19']:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 86, itr=True)
    #                 continue
    #             if kod[4:6] == '16':
    #                 if kod[:7] in ['037','050','051','054','055','059']:
    #                     write_sheet_AN_14_3_3_1(sheet, d, 86, itr=True)
    #                     continue
    #             ###
    #             ###9.2
    #             if 'K36' <= d.sluchay.dskz.kod[:3] <= 'K38':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 88, itr=True)
    #                 continue
    #             ###
    #             ###13.2
    #             if 'A16.20.038' == kod:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 101, itr=True)
    #                 continue
    #             ###
    #             ###14.0
    #             if 'O00' == d.sluchay.dskz.kod[:3]:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 105, itr=True)
    #                 write_sheet_AN_14_3_3_1(sheet, d, 106, itr=True)
    #                 continue
    #             if 'O02' <= d.sluchay.dskz.kod[:3] <= 'O08':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 105, itr=True)
    #                 continue
    #             ###
    #             ###14.6
    #             if 'A16.20.072' == kod:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 111, itr=True)
    #                 continue
    #             ###
    #             ###15.0:
    #             if kod[4:6] == '30':
    #                 if kod[:7] == '058':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                     continue
    #             if kod[4:6] == '03':
    #                 if 'Q65' == d.sluchay.dskz.kod[:3]:
    #                     write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                     continue
    #                 if '001' <= kod[:7] <= '012':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                     continue
    #                 if '015' <= kod[:7] <= '024':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                     continue
    #                 if '068' <= kod[:7] <= '071':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                     continue
    #                 if '080' == kod[:7] or kod[:7] == '081':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                     continue
    #                 if kod[:7] in ['063','082']:
    #                     write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                     continue
    #                 if kod[:7] == '024' or kod[:7] == '025':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                     continue
    #             if kod[4:6] == '02':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                 continue
    #             if kod[4:6] == '03':
    #                 if '058' == kod[:7]:
    #                     write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                     continue
    #             if '017' <= kod[:7] <= '020':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                 continue
    #             if '029' <= kod[:7] <= '033':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                 continue
    #             if '048' <= kod[:7] <= '050':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                 continue
    #             if '034' <= kod[:7] <= '044':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                 continue
    #             if kod[:7] in ['053','028']:
    #                 write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                 continue
    #             if '021' <= kod[:7] <= '028':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 115, itr=True)
    #                 continue
    #             ###
    #             ###15.2
    #             if kod[4:6] == '30':
    #                 if '012' < kod[:7] < '007':
    #                     write_sheet_AN_14_3_3_1(sheet, d, 117, itr=True)
    #                     continue
    #             ###
    #             ###16
    #             if kod[4:6] == '20':
    #                 if '20.031' == kod[4:]:
    #                     write_sheet_AN_14_3_3_1(sheet, d, 129, itr=True)
    #                     continue
    #             ###
    #             ###17
    #             if kod[4:6] == '01':
    #                 write_sheet_AN_14_3_3_1(sheet, d, 130, itr=True)
    #                 continue
    #             ###
    #         write_sheet_AN_14_3_3_1(sheet, d, 136, itr=True)
    # else:
    #     write_sheet_AN_14_3_3_1(sheet,all_data_oper,10)
    #
    # for n in range(127):
    #     for k in range(6):
    #         sheet.cell(row=10+n, column=3+k).value = sheet.cell(row=10+n, column=3+k).value if int(sheet.cell(row=10+n, column=3+k).value) != 0 else None
    result_all = []

    for d in data:
        row = get_row_god_14(d,row=0)
        if len(row) > 0:
            result = write_sheet_AN_14_3_3_1(d)
            result_all.append(result)
            for r in set(row):
                for res in range(len(result)):
                    if sheet.cell(row=r, column=3 + res).value != None:
                        val = int(sheet.cell(row=r, column=3 + res).value) + result[res]
                        sheet.cell(row=r, column=3 + res).value = val if val != 0 else None
                    else:
                        sheet.cell(row=r, column=3 + res).value = result[res] if result[res] != 0 else None
    # result = write_sheet_AN_14_3_3_1(data,itr=True)
    # for res in range(len(result)):
    #     sheet.cell(row=10, column=3 + res).value = result[res] if result[res] != 0 else None
    row = [11,43,45,56,59,63,76,86,95,96,98,99,105,115,129,130,132,134,135,136]
    oper_count = 0
    oper_count_vmt = 0
    oslo_count = 0
    oslo_count_vmt = 0
    ymer_count = 0
    ymer_count_vmt = 0
    for r in row:
        oper_count += int(sheet.cell(row=r, column=3).value) if sheet.cell(row=r, column=3).value != None else 0
        oper_count_vmt += int(sheet.cell(row=r, column=4).value) if sheet.cell(row=r, column=4).value != None else 0
        oslo_count += int(sheet.cell(row=r, column=5).value) if sheet.cell(row=r, column=5).value != None else 0
        oslo_count_vmt += int(sheet.cell(row=r, column=6).value) if sheet.cell(row=r, column=6).value != None else 0
        ymer_count += int(sheet.cell(row=r, column=7).value) if sheet.cell(row=r, column=7).value != None else 0
        ymer_count_vmt += int(sheet.cell(row=r, column=8).value) if sheet.cell(row=r, column=8).value != None else 0
    sheet.cell(row=10, column=3).value = oper_count if oper_count != 0 else None
    sheet.cell(row=10, column=4).value = oper_count_vmt if oper_count_vmt != 0 else None
    sheet.cell(row=10, column=5).value = oslo_count if oslo_count != 0 else None
    sheet.cell(row=10, column=6).value = oslo_count_vmt if oslo_count_vmt != 0 else None
    sheet.cell(row=10, column=7).value = ymer_count if ymer_count != 0 else None
    sheet.cell(row=10, column=8).value = ymer_count_vmt if ymer_count_vmt != 0 else None


def get_rez_apr_1(data,d=None):
    if d == None:
        all_temp = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        _ = [0,0,0,0,0,0,0,0,0,0]
        for patient in data:
            _[0] += 1
            _[1] += int(patient.sluchay.le_vr.kd) if patient.sluchay.le_vr.kd != ''and patient.sluchay.le_vr.kd != None else 0
            if patient.sluchay.goc is not None and patient.sluchay.goc.tip_name == 'Экстренная':
                _[2] += 1
                _[3] += int(patient.sluchay.le_vr.kd) if patient.sluchay.le_vr.kd != '' and patient.sluchay.le_vr.kd != None else 0
            if patient.sluchay.rslt and patient.sluchay.rslt.id_tip in [105, 106]:
                _[4] += 1
                _[5] += int(patient.sluchay.le_vr.kd) if patient.sluchay.le_vr.kd != '' and patient.sluchay.le_vr.kd != None else 0
                if patient.sluchay.goc is not None and patient.sluchay.goc.tip_name == 'Экстренная':
                    _[6] += 1
                    _[7] += int(patient.sluchay.le_vr.kd) if patient.sluchay.le_vr.kd != '' and patient.sluchay.le_vr.kd != None else 0
                if patient.patient_year in [0, 1]:
                    _[8] += 1
                    _[9] += int(patient.sluchay.le_vr.kd) if patient.sluchay.le_vr.kd != '' and patient.sluchay.le_vr.kd != None else 0

        all_temp[0] = _[0]
        all_temp[1] = _[1]

        try:
            all_temp[2] = float('{0:.2f}'.format(_[1]/_[0]))
        except ZeroDivisionError:
            all_temp[2] = 0

        all_temp[3] = _[2]
        try:
            all_temp[4] = float('{0:.2f}'.format(_[2]*100/_[0]))
        except ZeroDivisionError:
            all_temp[4] = 0
        try:
            all_temp[5]= float('{0:.2f}'.format(_[3]/_[2]))
        except ZeroDivisionError:
            all_temp[5] = 0

        all_temp[6] = _[4]

        try:
            all_temp[7] = float('{0:.2f}'.format(_[4]*100/_[0]))
        except ZeroDivisionError:
            all_temp[7] = 0

        try:
            all_temp[8]= float('{0:.2f}'.format(_[5]/_[4]))
        except ZeroDivisionError:
            all_temp[8] = 0

        all_temp[9] = _[6]

        try:
            all_temp[10] = float('{0:.2f}'.format(_[6]*100/_[0]))
        except ZeroDivisionError:
            all_temp[10] = 0

        try:
            all_temp[11]= float('{0:.2f}'.format(_[7]/_[6]))
        except ZeroDivisionError:
            all_temp[11] = 0

        all_temp[12] = _[8]

        try:
            all_temp[13] = float('{0:.2f}'.format(_[8]*100/_[0]))
        except ZeroDivisionError:
            all_temp[13] = 0

        try:
            all_temp[14]= float('{0:.2f}'.format(_[9]/_[8]))
        except ZeroDivisionError:
            all_temp[14] = 0
        all_temp[15] = _[3]

        # bf = BetterFilter()
        # sp = CountSluchaySpecification() ^ ProfKNSpecification() ^\
        #     GocEkSpecification() ^ GocEkNSpecification() ^\
        #     RezUmerSpecification() ^ RezUmerKdSpecification() ^ \
        #     RezUmerGocEkSpecification() ^ RezUmerGocEkSrSpecification() ^\
        #     RezUmerDetSpecification() ^ RezUmerKdDetSpecification()
        #
        #
        # all_temp = []
        # age_one = 0
        # kd = 0
        # for patient in data:
        #     for p in bf.filter(patient,sp):
        #         temp = bf.format_list(p)
        #         for t in range(len(temp)):
        #             if temp[t] == 'None':
        #                 temp[t] = 0
        #         all_temp.append([int(i) for i in temp])
        #     if patient.patient_year in [0,1]:
        #         if patient.sluchay.rslt and patient.sluchay.rslt.id_tip in [105, 106]:
        #             age_one+=1
        #             kd += int(patient.sluchay.le_vr.kd) if patient.sluchay.le_vr.kd != ''and patient.sluchay.le_vr.kd != None else 0
        #
        #
        # all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(10)]
        # try:
        #     all_temp.insert(2,float('{0:.2f}'.format(all_temp[1]/all_temp[0])))
        # except ZeroDivisionError:
        #     all_temp.insert(2,0)
        #
        # try:
        #     all_temp.insert(4,float('{0:.2f}'.format(all_temp[3]*100/all_temp[0])))
        # except ZeroDivisionError:
        #     all_temp.insert(4,0)
        # try:
        #     all_temp[5]= float('{0:.2f}'.format(all_temp[5]/all_temp[3]))
        # except ZeroDivisionError:
        #     all_temp[5] = 0
        #
        #
        # try:
        #     all_temp.insert(7,float('{0:.2f}'.format(all_temp[6]*100/all_temp[0])))
        # except ZeroDivisionError:
        #     all_temp.insert(7,0)
        # try:
        #     all_temp[8]= float('{0:.2f}'.format(all_temp[8]/all_temp[6]))
        # except ZeroDivisionError:
        #     all_temp[8] = 0
        #
        #
        # try:
        #     all_temp.insert(10,float('{0:.2f}'.format(all_temp[9]*100/all_temp[0])))
        # except ZeroDivisionError:
        #     all_temp.insert(10,0)
        # try:
        #     all_temp[11]= float('{0:.2f}'.format(all_temp[11]/all_temp[9]))
        # except ZeroDivisionError:
        #     all_temp[11] = 0
        #
        #
        # try:
        #     all_temp.insert(13,float('{0:.2f}'.format(all_temp[12]*100/all_temp[0])))
        # except ZeroDivisionError:
        #     all_temp.insert(13,0)
        #
        # try:
        #     all_temp[14]= float('{0:.2f}'.format(all_temp[14]/all_temp[12]))
        # except ZeroDivisionError:
        #     all_temp[14] = 0

        # all_temp[15] = age_one
        # all_temp.insert(16,0)
        # all_temp.insert(17,0)


        return all_temp

    try:
        data[2] = float('{0:.2f}'.format(data[1]/data[0]))
    except ZeroDivisionError:
        data[2] = 0

    try:
        data[4] = float('{0:.2f}'.format(data[3]*100/data[0]))
    except ZeroDivisionError:
        data[4] = 0
    try:
        data[5]= float('{0:.2f}'.format(data[15]/data[3]))
    except ZeroDivisionError:
        data[5] = 0

    try:
        data[7] = float('{0:.2f}'.format(data[6]*100/data[0]))
    except ZeroDivisionError:
        data[7] = 0

    try:
        data[8]= float('{0:.2f}'.format(data[8]/data[6]))
    except ZeroDivisionError:
        data[8] = 0

    try:
        data[10] = float('{0:.2f}'.format(data[9]*100/data[0]))
    except ZeroDivisionError:
        data[10] = 0
    try:
        data[11]= float('{0:.2f}'.format(data[11]/data[9]))
    except ZeroDivisionError:
        data[11] = 0

    try:
        data[13] = float('{0:.2f}'.format(data[12]*100/data[0]))
    except ZeroDivisionError:
        data[13] = 0
    try:
        data[14]= float('{0:.2f}'.format(data[14]/data[12]))
    except ZeroDivisionError:
        data[14] = 0

    return data
def get_rez_apr_2(data,d=None,count=0):
    if d == None:
        bf = BetterFilter()
        sp = OperCountSpecification() ^ OperCountGocEkSpecification() ^ OperAllCountSpecification() ^ OperAllCountGocEkSpecification() ^\
            OperAllKdSpecification() ^ PredOperPlkdSpecification() ^ OsloCountAllSpecification()^ RezUmerOperSpecification() ^ PredOperPlkdSpecification()
        all_temp = []
        for patient in data:
            for p in bf.filter(patient,sp):
                temp = bf.format_list(p)
                for t in range(len(temp)):
                    if temp[t] == 'None':
                        temp[t] = 0
                all_temp.append([int(i) for i in temp])
        all_temp = [sum([all_temp[i][x] for i in range(len(all_temp))]) for x in range(9)]
        data_count = len(data)


        # all_temp.insert(5,0)
        all_temp.insert(6,0)
        all_temp.insert(8,0)
        all_temp.insert(10,0)
        all_temp.insert(12,copy.deepcopy(all_temp[4]))

        try:
            all_temp[4]= float('{0:.2f}'.format(all_temp[4]/all_temp[0]))
        except ZeroDivisionError:
            all_temp[4] = 0
        try:
            all_temp[5]= float('{0:.2f}'.format(all_temp[11]/(all_temp[0]-all_temp[1])))
        except ZeroDivisionError:
            all_temp[5] = 0

        # try:
        #     all_temp[6]= float('{0:.2f}'.format(all_temp[0]*100/all_temp[2]))
        # except ZeroDivisionError:
        #     all_temp[6] = 0
        if count != 0:
            try:
                # all_temp[6]= float('{0:.2f}'.format(count*100/count))
                all_temp[6]= float('{0:.2f}'.format(all_temp[0]*100/data_count))
            except ZeroDivisionError:
                all_temp[6] = 0
        else:
            all_temp[6] = 0
        try:
            all_temp[8]= float('{0:.2f}'.format(all_temp[7]*100/all_temp[0]))
        except ZeroDivisionError:
            all_temp[8] = 0
        try:
            all_temp[10]= float('{0:.2f}'.format(all_temp[9]*100/all_temp[0]))
        except ZeroDivisionError:
            all_temp[10] = 0
        return all_temp
    try:
        data[4]= float('{0:.2f}'.format(data[12]/data[0]))
    except ZeroDivisionError:
        data[4] = 0
    try:
        data[5]= float('{0:.2f}'.format(data[11]/(data[0]-data[1])))
    except ZeroDivisionError:
        data[5] = 0
    # try:
    #     data[6]= float('{0:.2f}'.format(data[0]*100/data[2]))
    # except ZeroDivisionError:
    #     data[6] = 0
    if count != 0:
        try:
            data[6]= float('{0:.2f}'.format(data[0]*100/count))
        except ZeroDivisionError:
            data[6] = 0
    else:
        data[6] = 0
    try:
        data[8]= float('{0:.2f}'.format(data[7]*100/data[0]))
    except ZeroDivisionError:
        data[8] = 0
    try:
        data[10]= float('{0:.2f}'.format(data[9]*100/data[0]))
    except ZeroDivisionError:
        data[10] = 0

    # print(count)
    # print(data)
    return data

def get_row_god_14(data,row,itr=None):
    rows = []
    if itr is None:
        opers =  rr.get_opers(data)
        if opers is not None:
            for oper in opers:
                kod = oper.kod_op.kod if oper.kod_op else None
                if kod is not None:
                    if kod == 'A16.01.012':
                        rows.append(130+row)
                    elif kod == 'A06.10.006':
                        rows.append(76 + row)
                        rows.append(77 + row)
                    elif kod == 'A16.10.021.001':
                        rows.append(136 + row)
                        pass
                    elif kod == 'A16.08.006.001':
                        rows.append(56 + row)
                    elif kod == 'A16.08.003':
                        rows.append(59 + row)
                        rows.append(60 + row)
                    elif kod == 'A16.30.079':
                        rows.append(136 + row)
                        pass
                    elif kod == 'A06.09.005.002':
                        rows.append(59 + row)
                    elif kod == 'A16.12.004.009':
                        rows.append(63 + row)
                        rows.append(74 + row)
                        rows.append(75 + row)
                    elif kod == 'A16.12.028.017':
                        rows.append(76 + row)
                        rows.append(77 + row)
                    elif kod == 'A06.12.049':
                        rows.append(76 + row)
                        rows.append(77 + row)
                    elif kod == 'A16.30.006.002':
                        rows.append(136 + row)
                        pass
                    elif kod == 'A16.01.004':
                        rows.append(130 + row)
                    elif kod == 'A18.05.002.001':
                        rows.append(136 + row)
                        pass
                    elif kod == 'A16.12.028':
                        rows.append(76 + row)
                        rows.append(77 + row)
                    elif kod == 'A16.12.003.001':
                        rows.append(76 + row)
                        rows.append(77 + row)
                    elif kod == 'A16.28.025':
                        rows.append(95 + row)
                    elif kod == 'A16.08.006':
                        rows.append(76 + row)
                        rows.append(85 + row)
                    elif kod == 'A16.12.028.007':
                        rows.append(76 + row)
                        rows.append(77 + row)
                        rows.append(82 + row)
                    elif kod == 'A16.12.004':
                        rows.append(76 + row)
                        rows.append(77 + row)
                    elif kod == 'A16.12.003':
                        rows.append(76 + row)
                        rows.append(77 + row)
                    elif kod == 'A06.12.006':
                        rows.append(76 + row)
                        rows.append(77 + row)
                    elif kod == 'A16.12.026.012':
                        rows.append(64 + row)
                        rows.append(74 + row)
                        rows.append(75 + row)
                    elif kod == 'A06.12.056':
                        rows.append(76 + row)
                        rows.append(77 + row)
                    elif kod == 'A16.03.028.007':
                        rows.append(115 + row)
                    elif kod == 'A11.12.001.005':
                        rows.append(76 + row)
                        rows.append(85 + row)
                    elif kod == 'A18.05.002.005':
                        rows.append(86 + row)
                    elif kod == 'A16.09.004':
                        rows.append(59 + row)
                    elif kod == 'A16.30.017.003':
                        rows.append(115 + row)
                        rows.append(122 + row)
                    elif kod == 'A16.30.076':
                        rows.append(130 + row)
                    elif kod == 'A18.05.011':
                        rows.append(136 + row)
                        pass
                    elif kod == 'A18.05.011.002':
                        rows.append(86 + row)
                    elif kod == 'A16.06.009.002':
                        rows.append(86 + row)
                    elif kod == 'A11.12.001':
                        rows.append(76 + row)
                        rows.append(85 + row)
                    elif kod == 'A18.05.001':
                        rows.append(86 + row)
                    elif kod == 'A16.03.022.006':
                        rows.append(115 + row)
                        rows.append(116 + row)
                    elif kod == 'A16.26.093':
                        rows.append(45 + row)
                        rows.append(53 + row)
                        rows.append(54 + row)
                    elif kod == 'A16.26.094':
                        rows.append(45 + row)
                        rows.append(53 + row)
                        rows.append(54 + row)
                    else:
                        rows.append(136 + row)
    else:
        if data == 'A16.01.012':
            rows.append(130 + row)
        elif data == 'A06.10.006':
            rows.append(76 + row)
            rows.append(77 + row)
        elif data == 'A16.10.021.001':
            rows.append(136 + row)
            pass
        elif data == 'A16.08.006.001':
            rows.append(56 + row)
        elif data == 'A16.08.003':
            rows.append(59 + row)
            rows.append(60 + row)
        elif data == 'A16.30.079':
            rows.append(136 + row)
            pass
        elif data == 'A06.09.005.002':
            rows.append(59 + row)
        elif data == 'A16.12.004.009':
            rows.append(63 + row)
            rows.append(74 + row)
            rows.append(75 + row)
        elif data == 'A16.12.028.017':
            rows.append(76 + row)
            rows.append(77 + row)
        elif data == 'A06.12.049':
            rows.append(76 + row)
            rows.append(77 + row)
        elif data == 'A16.30.006.002':
            rows.append(136 + row)
            pass
        elif data == 'A16.01.004':
            rows.append(130 + row)
        elif data == 'A18.05.002.001':
            rows.append(136 + row)
            pass
        elif data == 'A16.12.028':
            rows.append(76 + row)
            rows.append(77 + row)
        elif data == 'A16.12.003.001':
            rows.append(76 + row)
            rows.append(77 + row)
        elif data == 'A16.28.025':
            rows.append(95 + row)
        elif data == 'A16.08.006':
            rows.append(76 + row)
            rows.append(85 + row)
        elif data == 'A16.12.028.007':
            rows.append(76 + row)
            rows.append(77 + row)
            rows.append(82 + row)
        elif data == 'A16.12.004':
            rows.append(76 + row)
            rows.append(77 + row)
        elif data == 'A16.12.003':
            rows.append(76 + row)
            rows.append(77 + row)
        elif data == 'A06.12.006':
            rows.append(76 + row)
            rows.append(77 + row)
        elif data == 'A16.12.026.012':
            rows.append(64 + row)
            rows.append(74 + row)
            rows.append(75 + row)
        elif data == 'A06.12.056':
            rows.append(76 + row)
            rows.append(77 + row)
        elif data == 'A16.03.028.007':
            rows.append(115 + row)
        elif data == 'A11.12.001.005':
            rows.append(76 + row)
            rows.append(85 + row)
        elif data == 'A18.05.002.005':
            rows.append(86 + row)
        elif data == 'A16.09.004':
            rows.append(59 + row)
        elif data == 'A16.30.017.003':
            rows.append(115 + row)
            rows.append(122 + row)
        elif data == 'A16.30.076':
            rows.append(130 + row)
        elif data == 'A18.05.011':
            rows.append(136 + row)
            pass
        elif data == 'A18.05.011.002':
            rows.append(86 + row)
        elif data == 'A16.06.009.002':
            rows.append(86 + row)
        elif data == 'A11.12.001':
            rows.append(76 + row)
            rows.append(85 + row)
        elif data == 'A18.05.001':
            rows.append(86 + row)
        elif data == 'A16.03.022.006':
            rows.append(115 + row)
            rows.append(116 + row)
        elif data == 'A16.26.093':
            rows.append(45 + row)
            rows.append(53 + row)
            rows.append(54 + row)
        elif data == 'A16.26.094':
            rows.append(45 + row)
            rows.append(53 + row)
            rows.append(54 + row)
        else:
            rows.append(136 + row)
    return rows
def insert_sheet_AN_14_3_8(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = kwargs['data']
    result_all = []
    sheet.cell(row=4, column=1).value = str(name).capitalize()
    sheet.cell(row=5, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'
    for d in data:
        row = get_row_god_14(d, row=-1)
        if len(row) > 0:
            result = write_sheet_AN_14_3_8(d)
            for r in row:
                for res in range(len(result)):
                    if sheet.cell(row=r, column=3 + res).value != None:
                        val = int(sheet.cell(row=r, column=3 + res).value) + result[res]
                        sheet.cell(row=r, column=3 + res).value = val if val != 0 else None
                    else:
                        sheet.cell(row=r, column=3 + res).value = result[res] if result[res] != 0 else None
    row = [10,42,44,55,58,62,75,85,94,95,97,98,104,114,128,129,131,133,134,135]
    oper_count = 0
    oper_sl_count = 0
    kd_oper_count = 0
    for r in row:
        oper_count += int(sheet.cell(row=r, column=3).value) if sheet.cell(row=r, column=3).value != None else 0
        oper_sl_count += int(sheet.cell(row=r, column=4).value) if sheet.cell(row=r, column=4).value != None else 0
        kd_oper_count += int(sheet.cell(row=r, column=5).value) if sheet.cell(row=r, column=5).value != None else 0
    sheet.cell(row=9, column=3).value = oper_count if oper_count != 0 else None
    sheet.cell(row=9, column=4).value = oper_sl_count if oper_sl_count != 0 else None
    sheet.cell(row=9, column=5).value = kd_oper_count if kd_oper_count != 0 else None
    sl = int(sheet.cell(row=9, column=4).value) if sheet.cell(row=9, column=4).value != None else None
    kd = int(sheet.cell(row=9, column=5).value) if sheet.cell(row=9, column=5).value != None else None
    if sl is not None and kd is not None:
        try:
            v = float('{0:.2f}'.format(kd / sl))
        except ZeroDivisionError:
            v = None
        sheet.cell(row=9, column=6).value = v

def insert_sheet_AN_14_4(**kwargs):
    sheet = kwargs['sheet']
    name = kwargs['name']
    date_1 = kwargs['date_1']
    date_2 = kwargs['date_2']
    data = kwargs['data']

    sheet.cell(row=4, column=1).value = str(name).capitalize()
    sheet.cell(row=5, column=1).value = f'За период с {date_1.strftime("%d.%m.%Y")} по {date_2.strftime("%d.%m.%Y")} г.'

    nzE10_E14 = list(Ds.objects.values('kod').filter(Q(kod__range=('E10', 'E11.9'))|Q(kod__range=('E13','E14.9'))))
    nzE10_E14 = [k['kod'] for k in nzE10_E14]

    nzI10_I13 = list(Ds.objects.values('kod').filter(kod__range=('I10', 'I13.9')))
    nzI10_I13 = [k['kod'] for k in nzI10_I13]


    nzI25 = list(Ds.objects.values('kod').filter(kod__range=('I25', 'I25.9')))
    nzI25 = [k['kod'] for k in nzI25]


    nzJ40_J43 = list(Ds.objects.values('kod').filter(kod__range=('J40', 'J43.9')))
    nzJ40_J43 = [k['kod'] for k in nzJ40_J43]

    nzJ44 = list(Ds.objects.values('kod').filter(kod__range=('J44', 'J44.9')))
    nzJ44 = [k['kod'] for k in nzJ44]

    nzJ47 = list(Ds.objects.values('kod').filter(kod__range=('J47', 'J47.9')))
    nzJ47 = [k['kod'] for k in nzJ47]

    nzJ45_J46 = list(Ds.objects.values('kod').filter(Q(kod__range=('J45', 'J45.9'))|Q(kod__range=('J46','J46.9'))))
    nzJ45_J46 = [k['kod'] for k in nzJ45_J46]

    nz_list = []

    nz_list.extend([nzE10_E14,nzI10_I13,nzI25,nzJ40_J43,nzJ44,nzJ47,nzJ45_J46])
    row = 10
    for nz in nz_list:
        row += 1
        n = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        for d in data:
            if d.sluchay.dspat:
                ds_kod = d.sluchay.dspat.kod
            else:
                if d.sluchay.dskz:
                    ds_kod = d.sluchay.dskz.kod

            if ds_kod in nz:
                year = d.patient_year
                if 0 <= year <= 14:
                    n[0] +=1
                if 15 <= year <= 19:
                    n[1] +=1
                if 20 <= year <= 24:
                    n[2] += 1
                if 25 <= year <= 29:
                    n[3] += 1
                if 30 <= year <= 34:
                    n[4] += 1
                if 35 <= year <= 39:
                    n[5] += 1
                if 40 <= year <= 44:
                    n[6] += 1
                if 45 <= year <= 49:
                    n[7] += 1
                if 50 <= year <= 54:
                    n[8] += 1
                if 55 <= year <= 59:
                    n[9] += 1
                if 60 <= year <= 64:
                    n[10] += 1
                if 65 <= year <= 69:
                    n[11] += 1
                if 70 <= year <= 74:
                    n[12] += 1
                if 75 <= year <= 79:
                    n[13] += 1
                if 80 <= year <= 84:
                    n[14] += 1
                if 85 <= year:
                    n[15] += 1
        for r in range(len(n)):
            sheet.cell(row=row, column=4+r).value = n[r] if n[r] != 0 else None

def ds_list(ds1,ds2,rm,rw):
    ds = Ds.objects.values('kod').filter(kod__range=(ds1,ds2)).all()
    if ds.count() > 0:
        ds = [d['kod'] for d in ds]
        return [ds,[rm,rw]]
    return []

def get_rez_god_16_1_1(ds,data):
    pol_m = [0,0,0,0,0,0,0,0,0,0,0,0]
    pol_w = [0,0,0,0,0,0,0,0,0,0,0,0]
    for d in data:
        if d.sluchay.dskz:
            if d.sluchay.dskz.kod in ds:
                year = d.patient_year
                pol = d.patient.pol.id_pol if d.patient.pol else None
                dis = d.sluchay.disability
                if dis.dat_l1 != None and dis.dat_l2 != None:
                    date = (dis.dat_l2-dis.dat_l1).days
                    if pol == 1:
                        pol_m[0]+=date
                        pol_m[1]+=1
                        if 15 < year < 19:
                            pol_m[2]+=1
                        if 20 < year < 24:
                            pol_m[3] += 1
                        if 25 < year < 29:
                            pol_m[4] += 1
                        if 30 < year < 34:
                            pol_m[5] += 1
                        if 35 < year < 39:
                            pol_m[6] += 1
                        if 40 < year < 44:
                            pol_m[7] += 1
                        if 45 < year < 49:
                            pol_m[8] += 1
                        if 50 < year < 54:
                            pol_m[9] += 1
                        if 55 < year < 59:
                            pol_m[10] += 1
                        if settings.OLD_G <= year :
                            pol_m[11] +=1

                    else:
                        pol_w[0]+=date
                        pol_w[1]+=1
                        if 15 < year < 19:
                            pol_w[2] += 1
                        if 20 < year < 24:
                            pol_w[3] += 1
                        if 25 < year < 29:
                            pol_w[4] += 1
                        if 30 < year < 34:
                            pol_w[5] += 1
                        if 35 < year < 39:
                            pol_w[6] += 1
                        if 40 < year < 44:
                            pol_w[7] += 1
                        if 45 < year < 49:
                            pol_w[8] += 1
                        if 50 < year < 54:
                            pol_w[9] += 1
                        if 55 < year < 59:
                            pol_w[10] += 1
                        if settings.OLD_G <= year :
                            pol_w[11] += 1


    return [pol_m,pol_w]
def insert_sheet_AN_16_1_1(**kwargs):
    sheet = kwargs['sheet']
    data = kwargs['data']

    nzAB = ds_list('A00','B99',5,6)
    nzA15_A19 = ds_list('A15','A19',7,8)
    nzC00_D48 = ds_list('C00','D48',9,10)
    nzC00_C97 = ds_list('C00','C97',11,12)
    nzD50_D89 = ds_list('D50','D89',13,14)
    nzE00_E90 = ds_list('E00','E90',15,16)
    nzE10_E14 = ds_list('E10','E14',17,18)
    nzF00_F99 = ds_list('F00','F99',19,20)
    nzG00_G99 = ds_list('G00','G99',21,22)
    nzH00_H59 = ds_list('H00','H59',23,24)
    nzH60_H95 = ds_list('H60','H95',25,26)
    nzI00_I99 = ds_list('I00','I99',27,28)
    nzI20_I25 = ds_list('I20','I25',29,30)
    nzI60_I69 = ds_list('I60','I69',31,32)
    nzJ00_J99 = ds_list('J00','J99',33,34)

    nzJ00_J06 = ds_list('J04','J06',35,36)
    nzJ00_J06[0].append('J00')
    nzJ00_J06[0].append('J01')

    nzJ09_J11 = [['J09','J11'],[37,38]]
    nzJ12_J18 = ds_list('J12','J18',39,40)
    nz_K00_K93 = ds_list('K00','K93',41,42)
    nzL00_L99 = ds_list('L00','L99',43,44)
    nzM00_M99 = ds_list('M00','M99',45,46)
    nzN00_N99 = ds_list('N00','N99',47,48)
    nzO00_O99 = ds_list('O00','O99',49,49)
    nzQ00_Q99 = ds_list('Q00','Q99',50,51)
    nzS00_T98 = ds_list('S00','T98',52,53)
    nzU07 = [['U07.1','U07.2'],[54,55]]
    nzO03_O08 = ds_list('O03','O08',58,58)
    nz_list = []
    nz_list.extend([nzAB,nzA15_A19,nzC00_D48,nzC00_C97,nzD50_D89,nzE00_E90,
                   nzE10_E14,nzF00_F99,nzG00_G99,nzH00_H59,nzH60_H95,nzI00_I99,
                   nzI20_I25,nzI60_I69,nzJ00_J99,nzJ00_J06,nzJ09_J11,nzJ12_J18,
                   nz_K00_K93,nzL00_L99,nzM00_M99,nzN00_N99,nzO00_O99,nzQ00_Q99,
                   nzS00_T98,nzU07,nzO03_O08])

    for nz in nz_list:
        ds = nz[0]
        rm = nz[1][0]
        rw = nz[1][1]
        resultM, resultW = get_rez_god_16_1_1(ds, data)
        if rm != rw:
            for n,v in enumerate((resultM)):
                sheet.cell(row=rm, column=5 + n).value = v
            for n, v in enumerate((resultW)):
                sheet.cell(row=rw, column=5 + n).value = v
        else:
            for n, v in enumerate((resultW)):
                sheet.cell(row=rw, column=5 + n).value = v







class AnnualPr1(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user,request)
    def create(self):
        file = self.is_file('annual_pr_1.xlsx')
        if file is not None:
            wb = load_workbook(file)
            sheet = wb.active
            patients = PatientsData(self.date_1,self.date_2,self.user)
            patients.sluchays(cah=True)
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_APR_1(**dic)

            wb.save(self.path() + f'annual_pr_1_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'download',
                                                                                'text': self.path() + f'annual_pr_1_{self.user.user.id}.xlsx',
                                                                                'name': 'Отчет о работе отделений (по ИБ)'})
        else:
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'text': 'Отсутствует шаблон - annual_pr_1.xlsx'})
            raise FileNotFoundError('Отсутствует шаблон - annual_pr_1.xlsx')

class AnnualPr2(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user,request)
    def create(self):
        file = self.is_file('annual_pr_2.xlsx')
        if file is not None:
            wb = load_workbook(file)
            sheet = wb.active
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_APR_2(**dic)
                        
            wb.save(self.path() + f'annual_pr_2_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                            {'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name, {'type': 'download',
                                                                                    'text': self.path() + f'annual_pr_2_{self.user.user.id}.xlsx',
                                                                                    'name': 'Отчет о хирур.работе отделений'})

        else:
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'text': 'Отсутствует шаблон - annual_pr_2.xlsx'})
            raise FileNotFoundError('Отсутствует шаблон - annual_pr_2.xlsx')

class Annual_13_1_1(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user

    def create(self):
        file = self.is_file('Ф13.docx', forms=True)
        if file:
            doc = DocxTemplate(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            nzO02_006 = list(Ds.objects.values('kod').filter(kod__range=('O02', 'O06.9')))
            nzO02_006 = [k['kod'] for k in nzO02_006]

            nzO02 = list(Ds.objects.values('kod').filter(kod__range=('O02', 'O02.9')))
            nzO02 = [k['kod'] for k in nzO02]

            nzO03 = list(Ds.objects.values('kod').filter(kod__range=('O03', 'O03.9')))
            nzO03 = [k['kod'] for k in nzO03]

            nzO04 = list(Ds.objects.values('kod').filter(kod__range=('O04', 'O04.9')))
            nzO04 = [k['kod'] for k in nzO04]

            nzO05 = list(Ds.objects.values('kod').filter(kod__range=('O05', 'O05.9')))
            nzO05 = [k['kod'] for k in nzO05]

            nzO06 = list(Ds.objects.values('kod').filter(kod__range=('O06', 'O06.9')))
            nzO06 = [k['kod'] for k in nzO06]

            nzO00 = list(Ds.objects.values('kod').filter(kod__range=('O00', 'O00.9')))
            nzO00 = [k['kod'] for k in nzO00]

            nzO01 = list(Ds.objects.values('kod').filter(kod__range=('O01', 'O01.9')))
            nzO01 = [k['kod'] for k in nzO01]

            nzO07 = list(Ds.objects.values('kod').filter(kod__range=('O07', 'O07.9')))
            nzO07 = [k['kod'] for k in nzO07]

            data_12 = []
            data_22 = []

            for d in patients.patients:
                srber = 0
                if d.vb_a:
                    if d.vb_a.srber != None and d.vb_a.srber != '':
                        srber = d.vb_a.srber
                if srber < 12:
                    data_12.append(d)
                if 12 <= srber < 22:
                    data_22.append(d)
                # if d.vb_a and d.vb_a.srber != None and d.vb_a.srber != '' and d.vb_a.srber < 12:
                #     data_12.append(d)
                # if d.vb_a and d.vb_a.srber != None and d.vb_a.srber != ''and  12 <= d.vb_a.srber <= 22:
                #     data_22.append(d)

            ord12 = OrderedDict()
            ord12['ii1_4']=0
            ord12['ii1_5']=0
            ord12['ii1_6'] = 0
            ord12['ii1_7'] = 0
            ord12['ii1_8'] = 0
            ord12['ii1_9'] = 0
            ord12['ii1_10'] = 0
            ord12['ii1_11'] = 0

            ord12['ii2_4'] = 0
            ord12['ii2_5'] = 0
            ord12['ii2_6'] = 0
            ord12['ii2_7'] = 0
            ord12['ii2_8'] = 0
            ord12['ii2_9'] = 0
            ord12['ii2_10'] = 0
            ord12['ii2_11'] = 0

            ord12['ii3_4'] = 0
            ord12['ii3_5'] = 0
            ord12['ii3_6'] = 0
            ord12['ii3_7'] = 0
            ord12['ii3_8'] = 0
            ord12['ii3_9'] = 0
            ord12['ii3_10'] = 0
            ord12['ii3_11'] = 0

            ord12['ii4_4'] = 0
            ord12['ii4_5'] = 0
            ord12['ii4_6'] = 0
            ord12['ii4_7'] = 0
            ord12['ii4_8'] = 0
            ord12['ii4_9'] = 0
            ord12['ii4_10'] = 0
            ord12['ii4_11'] = 0

            ord12['ii5_4'] = 0
            ord12['ii5_5'] = 0
            ord12['ii5_6'] = 0
            ord12['ii5_7'] = 0
            ord12['ii5_8'] = 0
            ord12['ii5_9'] = 0
            ord12['ii5_10'] = 0
            ord12['ii5_11'] = 0

            ord12['ii6_4'] = 0
            ord12['ii6_5'] = 0
            ord12['ii6_6'] = 0
            ord12['ii6_7'] = 0
            ord12['ii6_8'] = 0
            ord12['ii6_9'] = 0
            ord12['ii6_10'] = 0
            ord12['ii6_11'] = 0

            ord12['ii7_4'] = 0
            ord12['ii7_5'] = 0
            ord12['ii7_6'] = 0
            ord12['ii7_7'] = 0
            ord12['ii7_8'] = 0
            ord12['ii7_9'] = 0
            ord12['ii7_10'] = 0
            ord12['ii7_11'] = 0

            ord12['ii8_4'] = 0
            ord12['ii8_5'] = 0
            ord12['ii8_6'] = 0
            ord12['ii8_7'] = 0
            ord12['ii8_8'] = 0
            ord12['ii8_9'] = 0
            ord12['ii8_10'] = 0
            ord12['ii8_11'] = 0

            ord12['ii9_4'] = 0
            ord12['ii9_5'] = 0
            ord12['ii9_6'] = 0
            ord12['ii9_7'] = 0
            ord12['ii9_8'] = 0
            ord12['ii9_9'] = 0
            ord12['ii9_10'] = 0
            ord12['ii9_11'] = 0

            ord12['iii1_4'] = 0
            ord12['iii1_5'] = 0
            ord12['iii1_6'] = 0
            ord12['iii1_7'] = 0
            ord12['iii1_8'] = 0
            ord12['iii1_9'] = 0
            ord12['iii1_10'] = 0
            ord12['iii1_11'] = 0

            ord12['iii2_4'] = 0
            ord12['iii2_5'] = 0
            ord12['iii2_6'] = 0
            ord12['iii2_7'] = 0
            ord12['iii2_8'] = 0
            ord12['iii2_9'] = 0
            ord12['iii2_10'] = 0
            ord12['iii2_11'] = 0

            ord12['iii3_4'] = 0
            ord12['iii3_5'] = 0
            ord12['iii3_6'] = 0
            ord12['iii3_7'] = 0
            ord12['iii3_8'] = 0
            ord12['iii3_9'] = 0
            ord12['iii3_10'] = 0
            ord12['iii3_11'] = 0

            ord12['iii4_4'] = 0
            ord12['iii4_5'] = 0
            ord12['iii4_6'] = 0
            ord12['iii4_7'] = 0
            ord12['iii4_8'] = 0
            ord12['iii4_9'] = 0
            ord12['iii4_10'] = 0
            ord12['iii4_11'] = 0

            ord12['iii5_4'] = 0
            ord12['iii5_5'] = 0
            ord12['iii5_6'] = 0
            ord12['iii5_7'] = 0
            ord12['iii5_8'] = 0
            ord12['iii5_9'] = 0
            ord12['iii5_10'] = 0
            ord12['iii5_11'] = 0

            ord12['iii6_4'] = 0
            ord12['iii6_5'] = 0
            ord12['iii6_6'] = 0
            ord12['iii6_7'] = 0
            ord12['iii6_8'] = 0
            ord12['iii6_9'] = 0
            ord12['iii6_10'] = 0
            ord12['iii6_11'] = 0

            ord12['iii7_4'] = 0
            ord12['iii7_5'] = 0
            ord12['iii7_6'] = 0
            ord12['iii7_7'] = 0
            ord12['iii7_8'] = 0
            ord12['iii7_9'] = 0
            ord12['iii7_10'] = 0
            ord12['iii7_11'] = 0

            ord12['iii8_4'] = 0
            ord12['iii8_5'] = 0
            ord12['iii8_6'] = 0
            ord12['iii8_7'] = 0
            ord12['iii8_8'] = 0
            ord12['iii8_9'] = 0
            ord12['iii8_10'] = 0
            ord12['iii8_11'] = 0

            ord12['iii9_4'] = 0
            ord12['iii9_5'] = 0
            ord12['iii9_6'] = 0
            ord12['iii9_7'] = 0
            ord12['iii9_8'] = 0
            ord12['iii9_9'] = 0
            ord12['iii9_10'] = 0
            ord12['iii9_11'] = 0

            for d in data_12:
                year = d.patient_year
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO02_006:
                    ord12['ii1_4']+=1
                    if 0 <= year <= 14:
                        ord12['ii1_5']+=1
                    if 15 <= year <= 17:
                        ord12['ii1_6']+=1
                    if 18 <= year <= 44:
                        ord12['ii1_7']+=1
                    if 45 <= year <= 49:
                        ord12['ii1_8']+=1
                    if 50 <= year:
                        ord12['ii1_9']+=1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['ii1_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO02:
                    ord12['ii2_4'] += 1
                    if 0 <= year <= 14:
                        ord12['ii2_5'] += 1
                    if 15 <= year <= 17:
                        ord12['ii2_6'] += 1
                    if 18 <= year <= 44:
                        ord12['ii2_7'] += 1
                    if 45 <= year <= 49:
                        ord12['ii2_8'] += 1
                    if 50 <= year:
                        ord12['ii2_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['ii2_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO03:
                    ord12['ii3_4'] += 1
                    if 0 <= year <= 14:
                        ord12['ii3_5'] += 1
                    if 15 <= year <= 17:
                        ord12['ii3_6'] += 1
                    if 18 <= year <= 44:
                        ord12['ii3_7'] += 1
                    if 45 <= year <= 49:
                        ord12['ii3_8'] += 1
                    if 50 <= year:
                        ord12['ii3_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['ii3_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO04:
                    ord12['ii4_4'] += 1
                    if 0 <= year <= 14:
                        ord12['ii4_5'] += 1
                    if 15 <= year <= 17:
                        ord12['ii4_6'] += 1
                    if 18 <= year <= 44:
                        ord12['ii4_7'] += 1
                    if 45 <= year <= 49:
                        ord12['ii4_8'] += 1
                    if 50 <= year:
                        ord12['ii4_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['ii4_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO05:
                    ord12['ii5_4'] += 1
                    if 0 <= year <= 14:
                        ord12['ii5_5'] += 1
                    if 15 <= year <= 17:
                        ord12['ii5_6'] += 1
                    if 18 <= year <= 44:
                        ord12['ii5_7'] += 1
                    if 45 <= year <= 49:
                        ord12['ii5_8'] += 1
                    if 50 <= year:
                        ord12['ii5_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['ii5_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO06:
                    ord12['ii6_4'] += 1
                    if 0 <= year <= 14:
                        ord12['ii6_5'] += 1
                    if 15 <= year <= 17:
                        ord12['ii6_6'] += 1
                    if 18 <= year <= 44:
                        ord12['ii6_7'] += 1
                    if 45 <= year <= 49:
                        ord12['ii6_8'] += 1
                    if 50 <= year:
                        ord12['ii6_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['ii6_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO00:
                    ord12['ii7_4'] += 1
                    if 0 <= year <= 14:
                        ord12['ii7_5'] += 1
                    if 15 <= year <= 17:
                        ord12['ii7_6'] += 1
                    if 18 <= year <= 44:
                        ord12['ii7_7'] += 1
                    if 45 <= year <= 49:
                        ord12['ii7_8'] += 1
                    if 50 <= year:
                        ord12['ii7_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['ii7_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO01:
                    ord12['ii8_4'] += 1
                    if 0 <= year <= 14:
                        ord12['ii8_5'] += 1
                    if 15 <= year <= 17:
                        ord12['ii8_6'] += 1
                    if 18 <= year <= 44:
                        ord12['ii8_7'] += 1
                    if 45 <= year <= 49:
                        ord12['ii8_8'] += 1
                    if 50 <= year:
                        ord12['ii8_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['ii8_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO07:
                    ord12['ii9_4'] += 1
                    if 0 <= year <= 14:
                        ord12['ii9_5'] += 1
                    if 15 <= year <= 17:
                        ord12['ii9_6'] += 1
                    if 18 <= year <= 44:
                        ord12['ii9_7'] += 1
                    if 45 <= year <= 49:
                        ord12['ii9_8'] += 1
                    if 50 <= year:
                        ord12['ii9_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['ii9_10'] += 1

            for d in data_22:
                year = d.patient_year
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO02_006:
                    ord12['iii1_4']+=1
                    if 0 <= year <= 14:
                        ord12['iii1_5']+=1
                    if 15 <= year <= 17:
                        ord12['iii1_6']+=1
                    if 18 <= year <= 44:
                        ord12['iii1_7']+=1
                    if 45 <= year <= 49:
                        ord12['iii1_8']+=1
                    if 50 <= year:
                        ord12['iii1_9']+=1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['iii1_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO02:
                    ord12['iii2_4'] += 1
                    if 0 <= year <= 14:
                        ord12['iii2_5'] += 1
                    if 15 <= year <= 17:
                        ord12['iii2_6'] += 1
                    if 18 <= year <= 44:
                        ord12['iii2_7'] += 1
                    if 45 <= year <= 49:
                        ord12['iii2_8'] += 1
                    if 50 <= year:
                        ord12['iii2_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['iii2_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO03:
                    ord12['iii3_4'] += 1
                    if 0 <= year <= 14:
                        ord12['iii3_5'] += 1
                    if 15 <= year <= 17:
                        ord12['iii3_6'] += 1
                    if 18 <= year <= 44:
                        ord12['iii3_7'] += 1
                    if 45 <= year <= 49:
                        ord12['iii3_8'] += 1
                    if 50 <= year:
                        ord12['iii3_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['iii3_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO04:
                    ord12['iii4_4'] += 1
                    if 0 <= year <= 14:
                        ord12['iii4_5'] += 1
                    if 15 <= year <= 17:
                        ord12['iii4_6'] += 1
                    if 18 <= year <= 44:
                        ord12['iii4_7'] += 1
                    if 45 <= year <= 49:
                        ord12['iii4_8'] += 1
                    if 50 <= year:
                        ord12['iii4_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['iii4_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO05:
                    ord12['iii5_4'] += 1
                    if 0 <= year <= 14:
                        ord12['iii5_5'] += 1
                    if 15 <= year <= 17:
                        ord12['iii5_6'] += 1
                    if 18 <= year <= 44:
                        ord12['iii5_7'] += 1
                    if 45 <= year <= 49:
                        ord12['iii5_8'] += 1
                    if 50 <= year:
                        ord12['iii5_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['iii5_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO06:
                    ord12['iii6_4'] += 1
                    if 0 <= year <= 14:
                        ord12['iii6_5'] += 1
                    if 15 <= year <= 17:
                        ord12['iii6_6'] += 1
                    if 18 <= year <= 44:
                        ord12['iii6_7'] += 1
                    if 45 <= year <= 49:
                        ord12['iii6_8'] += 1
                    if 50 <= year:
                        ord12['iii6_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['iii6_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO00:
                    ord12['iii7_4'] += 1
                    if 0 <= year <= 14:
                        ord12['iii7_5'] += 1
                    if 15 <= year <= 17:
                        ord12['iii7_6'] += 1
                    if 18 <= year <= 44:
                        ord12['iii7_7'] += 1
                    if 45 <= year <= 49:
                        ord12['iii7_8'] += 1
                    if 50 <= year:
                        ord12['iii7_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['iii7_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO01:
                    ord12['iii8_4'] += 1
                    if 0 <= year <= 14:
                        ord12['iii8_5'] += 1
                    if 15 <= year <= 17:
                        ord12['iii8_6'] += 1
                    if 18 <= year <= 44:
                        ord12['iii8_7'] += 1
                    if 45 <= year <= 49:
                        ord12['iii8_8'] += 1
                    if 50 <= year:
                        ord12['iii8_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['iii8_10'] += 1
                if d.sluchay.dskz and d.sluchay.dskz.kod in nzO07:
                    ord12['iii9_4'] += 1
                    if 0 <= year <= 14:
                        ord12['iii9_5'] += 1
                    if 15 <= year <= 17:
                        ord12['iii9_6'] += 1
                    if 18 <= year <= 44:
                        ord12['iii9_7'] += 1
                    if 45 <= year <= 49:
                        ord12['iii9_8'] += 1
                    if 50 <= year:
                        ord12['iii9_9'] += 1
                    if d.vb_a.n_ber != None and d.vb_a.n_ber != '' and d.vb_a.n_ber == 1:
                        ord12['iii9_10'] += 1
            doc.render(ord12)
            doc.save(file)
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name, {'type': 'download',
                                                                                 'text': file,
                                                                                 'name': ''})

#Нужны данные чтобы доделать
class Annual_13_1_4(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
    def create(self):
        file = self.is_file('annual_13_1_4.xlsx',forms=False)
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays(cah=True)
            temp = []
            # list_dskz_O = []
            # list_vb_a = []
            for p in patients.patients:
                if p.sluchay.vb_a:
                    vb_a = p.sluchay.vb_a
                    if vb_a.srber and vb_a.n_ber:
                        temp.append(p)
            #     if  p.vb_a != None and p.sluchay.dskz and (p.sluchay.dskz.kod[:3] in ['O02','O03,','O04','O05','O06']):
            #         list_dskz_O.append(p)
            #     if p.vb_a != None and p.sluchay.dskz and (p.sluchay.dskz.kod[:3] in ['O04']):
            #         list_vb_a.append(p)

            # dic = dict([('sheet', sheet), ('name', self.user.statistics_type.name),
            #             ('date_1', self.date_1), ('date_2', self.date_2), ('list_dskz_O', list_dskz_O),
            #             ('list_vb_a',list_vb_a),('data', patients.patients)])
            dic = dict([('sheet', sheet), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2), ('data', temp)])
            insert_sheet_AN_13_1_4(**dic)
            wb.save(self.path() + f'group_an_13_4_{self.user.user.id}.xlsx')

        async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_group_data_annual', 'text': 'Отчет cфромирован'})
        async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                      {'type': 'download_data_annual','name': 'Отчет о выбывших'})
        async_to_sync(get_channel_layer().group_send)(self.user_group_name, {'type': 'download',
                                                                             'text': self.path() + f'group_an_13_4_{self.user.user.id}.xlsx',
                                                                             'name': 'Отчет о хирур.работе отделений'})


class Annual_30_1_1(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
    def create(self):
        file = self.is_file('annual_30_1_1.xlsx', forms=False)
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            dic = dict([('sheet', sheet), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2), ('data', patients.patients)])
            insert_sheet_AN_30_1_1(**dic)
            wb.save(self.path() + f'annual_30_1_1_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name, {'type': 'download',
                                                                                 'text': self.path() + f'annual_30_1_1_{self.user.user.id}.xlsx',
                                                                                 'name': ''})
        # get_list_otd_prof()

class Annual_30_2_1(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
    def create(self):
        file = self.is_file('annual_30_2_1.xlsx', forms=False)
        if file:
            wb = load_workbook(file)
            # sheet = wb.active
            sheet1 = wb.get_sheet_by_name('Лист1')
            sheet2 = wb.get_sheet_by_name('Лист2')
            sheet3 = wb.get_sheet_by_name('Лист3')
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            dic = dict([('sheet', [sheet1,sheet2,sheet3]), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2), ('data', patients.patients)])
            insert_sheet_AN_30_2_1(**dic)
            wb.save(self.path() + f'annual_30_2_1_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name, {'type': 'download',
                                                                                 'text': self.path() + f'annual_30_2_1_{self.user.user.id}.xlsx',
                                                                                 'name': ''})
class Annual_Vra(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('annual_vra.xlsx')
        if file:
            wb = load_workbook(file)
            sheet=wb.get_sheet_by_name('Операции')
            sheet1=wb.get_sheet_by_name('Ноз.групп')
            sheet2=wb.get_sheet_by_name('ДС')
            sheet3=wb.get_sheet_by_name('Манипул.')

            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            v = self.request.get('vra').split(' ')
            vra = Vra.objects.filter(kod=v[0],naim=v[1],dateend=None)[:1][0]

            dic = dict([('sheet', [sheet,sheet1,sheet2,sheet3]), ('name', self.user.statistics_type.name),('date_1', self.date_1),
                ('date_2', self.date_2),('vra',vra),('data',patients.patients)])
            insert_sheet_VRA(**dic)
            wb.save(self.path() + f'group_an_vra_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_group_data_annual', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'group_an_vra_{self.user.user.id}.xlsx'})
class Annual_14_1_1(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('god_14_1.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            data = []
            # year = datetime.now().year
            for p in patients.patients:
                year = p.patient_year
                if p.sluchay.otd and p.sluchay.otd.naim == 'КАРДИОЛОГИЧЕСКОЕ':
                    if (year)>=18:
                        if p.sluchay.rslt and p.sluchay.rslt.id_tip not in [102,104]:
                            data.append(p)
            dic = dict([('sheet', sheet), ('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_AN_14_1_A(**dic)
            wb.save(self.path() + f'annual_14_1_1_{self.user.user.id}.xlsx')
            # print(ymkd)
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'annual_14_1_1_{self.user.user.id}.xlsx'})
class Annual_14_1_2(AnnualReportABC):
    def __init__(self,user,request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('god_14_2.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            data = []
            for p in patients.patients:
                if p.sluchay.otd and p.sluchay.otd.naim == 'КАРДИОЛОГИЧЕСКОЕ':
                    pol = p.patient.pol.id_pol if p.patient.pol and p.patient.pol.id_pol else None
                    year = p.patient_year
                    if p.sluchay.rslt and p.sluchay.rslt.id_tip not in [102, 104]:
                        if (pol == 1) and (year >= settings.OLD_M):
                            data.append(p)
                        if (pol == 2) and (year >= settings.OLD_G):
                            data.append(p)
            dic = dict([('sheet', sheet), ('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_AN_14_1_A(**dic)
            wb.save(self.path() + f'annual_14_1_2_{self.user.user.id}.xlsx')
            # print(ymkd)
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'annual_14_1_2_{self.user.user.id}.xlsx'})

class Annual_14_1_4(AnnualReportABC):
    def __init__(self, user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('14_1_4.docx',forms=True)
        if file:
            doc = DocxTemplate(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()

            data = []
            for p in patients.patients:
                if p.sluchay.otd and p.sluchay.otd.naim == 'КАРДИОЛОГИЧЕСКОЕ':
                    data.append(p)

            dic = dict([('doc',doc),('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_AN_14_1_4(**dic)
            doc.save(file)
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'annual_14_1_4_{self.user.user.id}.docx'})

class Annual_14_3_1(AnnualReportABC):
    def __init__(self, user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('god_14_3_1.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            data = []
            for p in patients.patients:
                if p.sluchay.otd and p.sluchay.otd.naim == 'КАРДИОЛОГИЧЕСКОЕ':
                    data.append(p)
            #     year = datetime.now().year - p.patient.datr.year
            #     if 0 <= year <= 17:
            #         data.append(p)

            dic = dict([('sheet', sheet), ('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2),('typ','det')])
            insert_sheet_AN_14_3_1(**dic)
            wb.save(self.path() + f'annual_14_3_1_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'annual_14_3_1_{self.user.user.id}.xlsx'})

class Annual_14_3_2(AnnualReportABC):
    def __init__(self, user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('god_14_3_1.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            dic = dict([('sheet', sheet), ('data', patients.patients), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_AN_14_3_2(**dic)
            wb.save(self.path() + f'annual_14_3_2_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,{'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'annual_14_3_2_{self.user.user.id}.xlsx'})

class Annual_14_3_3(AnnualReportABC):
    def __init__(self, user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('god_14_3_3_1.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            data = []
            for p in patients.patients:
                if p.sluchay.otd and p.sluchay.otd.naim == 'КАРДИОЛОГИЧЕСКОЕ':
                    year = p.patient_year
                    pol = p.patient.pol.id_pol if p.patient.pol and p.patient.pol.id_pol else None
                    if pol == 1:
                        if year >= settings.OLD_M:
                            data.append(p)
                    if pol == 2:
                        if year >= settings.OLD_G:
                            data.append(p)
            dic = dict([('sheet', sheet), ('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2),('typ','det')])
            insert_sheet_AN_14_3_3_1(**dic)
            wb.save(self.path() + f'annual_14_3_3_1_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'annual_14_3_3_1_{self.user.user.id}.xlsx'})

class Annual_14_3_8(AnnualReportABC):
    def __init__(self, user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('god_14_3_8.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            data = []
            for p in patients.patients:
                if p.sluchay.otd and p.sluchay.otd.naim == 'КАРДИОЛОГИЧЕСКОЕ':
                    if p.sluchay.metod_hmp != '':
                        data.append(p)
            dic = dict([('sheet', sheet), ('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_AN_14_3_8(**dic)
            wb.save(self.path() + f'annual_14_3_8_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'annual_14_3_8_{self.user.user.id}.xlsx'})

class Annual_14_4(AnnualReportABC):
    def __init__(self, user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('god_14_4.xlsx')
        if file:
            wb = load_workbook(file)
            sheet = wb.active
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            data = []
            for p in patients.patients:
                if p.sluchay.otd and p.sluchay.otd.naim == 'КАРДИОЛОГИЧЕСКОЕ':
                    if p.sluchay.rslt and p.sluchay.rslt.id_tip not in [102, 104]:
                        if p.sluchay.rslt and p.sluchay.rslt.id_tip  in [105, 106]:
                            data.append(p)
            dic = dict([('sheet', sheet), ('data', data), ('name', self.user.statistics_type.name),
                        ('date_1', self.date_1), ('date_2', self.date_2)])
            insert_sheet_AN_14_4(**dic)
            wb.save(self.path() + f'annual_14_4_{self.user.user.id}.xlsx')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'annual_14_4_{self.user.user.id}.xlsx'})

class Annual_16_1_1(AnnualReportABC):
    def __init__(self, user, request):
        super().__init__(user, request)
        self.user_group_name = 'hospital_annual_reports_%s' % user
        self.request = request
    def create(self):
        file = self.is_file('god_16_1_1.xlsx')
        if file:
            wb = load_workbook(file)
            sheet=wb.get_sheet_by_name('1000')
            os.remove(file)
            patients = PatientsData(self.date_1, self.date_2, self.user)
            patients.sluchays()
            data = []
            for p in patients.patients:
                if p.sluchay.disability and p.sluchay.disability.ot_ln == True:
                    # if p.sluchay.otd and p.sluchay.otd.naim == 'КАРДИОЛОГИЧЕСКОЕ':
                    data.append(p)
            # print(len(data))
            dic = dict([('sheet', sheet), ('data', data)])
            insert_sheet_AN_16_1_1(**dic)
            wb.save(self.path() + f'god_16_1_1_{self.user.user.id}.xls')
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'report_data', 'text': 'Отчет cфромирован'})
            async_to_sync(get_channel_layer().group_send)(self.user_group_name,
                                                          {'type': 'download',
                                                           'text': self.path() + f'god_16_1_1_{self.user.user.id}.xls'})
def AnnualReport(user,request):
    print('AnnualReport',request)
    type_fun = request.get('type_report')
    if type_fun == 'annual_pr_1':
        report = AnnualPr1(user,request)
        report.create()
    elif type_fun == 'annual_pr_2':
        report = AnnualPr2(user,request)
        report.create()
    elif type_fun == 'annual_13_1_1':
        report = Annual_13_1_1(user,request)
        report.create()
    elif type_fun == 'annual_13_1_4':
        report = Annual_13_1_4(user,request)
        report.create()
    elif type_fun == 'annual_30_1_1':
        report = Annual_30_1_1(user,request)
        report.create()
    elif type_fun == 'annual_30_2_1':
        report = Annual_30_2_1(user,request)
        report.create()
    elif type_fun == 'annual_vra':
        report = Annual_Vra(user,request)
        report.create()
    elif type_fun == 'annual_14_1_1':
        report = Annual_14_1_1(user,request)
        report.create()
    elif type_fun == 'annual_14_1_2':
        report = Annual_14_1_2(user,request)
        report.create()
    elif type_fun == 'annual_14_1_4':
        report = Annual_14_1_4(user, request)
        report.create()
    elif type_fun == 'annual_14_3_1':
        report = Annual_14_3_1(user, request)
        report.create()
    elif type_fun == 'annual_14_3_2':
        report = Annual_14_3_2(user,request)
        report.create()
    elif type_fun == 'annual_14_3_3':
        report = Annual_14_3_3(user, request)
        report.create()
    elif type_fun == 'annual_14_3_8':
        report = Annual_14_3_8(user,request)
        report.create()
    elif type_fun == 'annual_14_4':
        report = Annual_14_4(user, request)
        report.create()
    elif type_fun == 'annual_16_1_1':
        report = Annual_16_1_1(user, request)
        report.create()



