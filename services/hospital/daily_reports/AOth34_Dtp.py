from datetime import datetime
from hospital.models import Sluchay
from okb2.models import Oksm,V020,vbb_ds_prk
from services.hospital.patient import PatientsData
from django.conf import  settings
import shutil
from openpyxl import load_workbook,styles
from  openpyxl.styles import Font,Alignment,Border,Side
import os
from services.hospital.create_reestr import Create
from django.core.mail import EmailMessage
class AOth34_AOth35(PatientsData):
    def __init__(self):
        super().__init__(date_1=None,date_2=None,user=None)
    def create(self):
        border = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

        date2 = datetime.now()
        year = date2.year
        month = date2.month
        if month == 1:
            date1 = datetime(year-1,12,21)
        else:
            date1 = datetime(year,month-1,21)

        sluchay_list = Sluchay.objects.values('id').filter(datv__range=(f'{date1.year}-{date1.month}-{date1.day}',
                                                                        f'{date2.year}-{date2.month}-{20}')).exclude(update_user=None)
        # sluchay_list = Sluchay.objects.values('id').filter(datv__range=(f'{2023}-{1}-{21}',
        #                                                                 f'{2023}-{1}-{30}')).exclude(update_user=None)
        for s in sluchay_list:
            try:
                self.get_data(s['id'])
            except:
                pass
        data34 = []
        ing = []
        for p in self.patients:
            if p.sluchay.le_trv:
                le_trv = p.sluchay.le_trv
                if le_trv.t_trv and le_trv.t_trv.kod == '7':
                    data34.append(p)
            if p.patient.c_oksm and p.patient.c_oksm.kod != 643:
                if p.vds.vds and p.vds.vds.kod in ['1','0','Д']:
                    ing.append(p)

        #ДТП
        file_shoblon = '/'.join([settings.MEDIA_ROOT, 'shoblons/hospital/oth', 'a_oth_34.xlsx'])
        filename = f'{"ДТП"}.xlsx'
        file_new_dtp = '/'.join([settings.MEDIA_ROOT,'shoblons/hospital/oth',filename])
        if os.path.isfile(file_new_dtp):
            os.remove(file_new_dtp)
        shutil.copy2(file_shoblon, file_new_dtp)
        wb = load_workbook(file_new_dtp)
        sheet = wb.active
        row = 7

        sheet.cell(row=4,column=1).value = f'За период с {date1.strftime("%d.%m.%Y")} по {date2.strftime("%d.%m.%Y")} г.'

        for d in data34:
            row+=1
            sheet.cell(row=row, column=1).value = d.sluchay.nib
            fam = d.patient.fam if len(d.patient.fam) > 0 else None
            im = d.patient.im if len(d.patient.im) > 0 else None
            ot = d.patient.ot if len(d.patient.ot) > 0 else None
            sheet.cell(row=row, column=2).value = f'{fam if fam != None else ""} {im[0] if im != None else ""}.{ot[0] if ot != None else ""}'
            sheet.cell(row=row, column=3).value = d.patient.datr.strftime('%d.%m.%Y') if d.patient.datr else None
            if d.sluchay.goc and d.sluchay.goc.tip_name == 'Экстренная':
                goc = 'экс'
            elif d.sluchay.goc and d.sluchay.goc.tip_name == 'Плановая':
                goc = 'пл'
            else:
                goc = ''
            sheet.cell(row=row, column=4).value = goc
            sheet.cell(row=row, column=5).value = d.sluchay.otd.naim if d.sluchay.otd else None
            sheet.cell(row=row, column=6).value = d.sluchay.datp.strftime('%d.%m.%Y') if d.sluchay.datp else None
            le_trv = d.sluchay.le_trv
            sheet.cell(row=row, column=8).value = le_trv.details.kod if le_trv.details else None
            sheet.cell(row=row, column=9).value = 'ДОРОЖНО - ТРАНСПОРТНАЯ'
            sheet.cell(row=row, column=10).value = d.patient.rab
            for i in range(1,11):
                sheet.cell(row=row, column=i).border = border
        wb.save(file_new_dtp)

        #Иностранцы
        file_shoblon = '/'.join([settings.MEDIA_ROOT, 'shoblons/hospital/oth', 'a_oth_35.xlsx'])
        filename = f'{"Инг"}.xlsx'
        file_new_ing = '/'.join([settings.MEDIA_ROOT,'shoblons/hospital/oth',filename])
        if os.path.isfile(file_new_ing):
            os.remove(file_new_ing)
        shutil.copy2(file_shoblon, file_new_ing)
        wb = load_workbook(file_new_ing)
        sheet = wb.active
        ing_result = self.get_list_otd_prof(ing)
        row = 1
        for ing in ing_result:
            c_oksm:Oksm = Oksm.objects.filter(kod =ing[0][0]).first()
            for prof in ing[1]:
                row+=1
                sheet.cell(row=row, column=1).value = 'ГБУЗ ТО "Областная клиническая больница №2"'
                pr = V020.objects.filter(k_prname=prof[0]).first()
                idk_pr = f'{pr.idk_pr}.0'
                profil = vbb_ds_prk.objects.filter(kod_r=idk_pr).first()
                sheet.cell(row=row, column=2).value = profil.miac
                sheet.cell(row=row, column=3).value = 'средства ОМС'
                sheet.cell(row=row, column=4).value = c_oksm.naim
                sheet.cell(row=row, column=5).value = len(prof[1])
                sum = 0
                for t in prof[1]:
                    tarif = Create.tarif_sp(Create,t,None)
                    sum+=float(tarif.Tarif) if tarif != None else 0
                sheet.cell(row=row, column=6).value = float('{0:.2f}'.format(sum))
        wb.save(file_new_ing)

        # email = EmailMessage("Отчеты стационар", 'Отчеты по Дтп и Иностранцам', ' arknet@okb2-tmn.ru', ['tyktybaev_ad@okb2-tmn.ru','besedovskaya_ia@okb2-tmn.ru'])
        email = EmailMessage("Отчеты стационар", 'Отчеты по Дтп и Иностранцам', ' arknet@okb2-tmn.ru',
                             ['tyktybaev_ad@okb2-tmn.ru'])
        email.attach_file(file_new_dtp)
        email.attach_file(file_new_ing)
        email.send()

    def get_list_otd_prof(self,data):
        ing_list = []
        for d in data:
            if d.patient.c_oksm.kod not in ing_list:
                ing_list.append(d.patient.c_oksm.kod)

        c_oksm_prof_list = []
        for o in ing_list:
            temp = [[], []]
            temp[0].append(o)
            for d in data:
                if o == d.patient.c_oksm.kod:
                    if d.sluchay.le_vr and d.sluchay.le_vr.prof_k:
                        if d.sluchay.le_vr.prof_k.k_prname not in temp[1]:
                            temp[1].append(d.sluchay.le_vr.prof_k.k_prname)
            c_oksm_prof_list.append(temp)
        #
        for c_oksm_prof in c_oksm_prof_list:
            for o in range(len(c_oksm_prof[1])):
                c_oksm_prof[1][o] = [c_oksm_prof[1][o], []]

        for otds_prof in range(len(c_oksm_prof_list)):
            for d in data:
                if c_oksm_prof_list[otds_prof][0][0] == d.patient.c_oksm.kod:
                    if d.sluchay.le_vr and d.sluchay.le_vr.prof_k:
                        for p in range(len(c_oksm_prof_list[otds_prof][1])):
                            if c_oksm_prof_list[otds_prof][1][p][0] == d.sluchay.le_vr.prof_k.k_prname:
                                c_oksm_prof_list[otds_prof][1][p][1].append(d)
        return c_oksm_prof_list